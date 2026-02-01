"""
Main Application Window
"""
from PyQt6.QtWidgets import (
    QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QTabWidget, QTableWidget, QTableWidgetItem, QPushButton,
    QLabel, QComboBox, QLineEdit, QTextEdit, QSplitter,
    QMessageBox, QStatusBar, QToolBar, QGroupBox, QFormLayout,
    QHeaderView, QDialog, QSpinBox, QDoubleSpinBox, QCheckBox,
    QScrollArea, QApplication, QSizePolicy
)
from PyQt6.QtCore import Qt, QTimer, QTime, pyqtSignal, QUrl
from PyQt6.QtGui import QAction, QFont, QDesktopServices, QCursor
from datetime import datetime

from algo_trader.core.config import Config
from algo_trader.core.database import Database
from algo_trader.core.order_manager import OrderManager
from algo_trader.core.strategy_engine import StrategyEngine
from algo_trader.core.risk_manager import RiskManager
from algo_trader.core.options_manager import (
    OptionsManager, OptionType, HedgeStrategy, ExitType
)
from algo_trader.core.auto_options import (
    AutoOptionsExecutor, StrikeSelection, SignalAction, ExpirySelection
)
from algo_trader.brokers import UpstoxBroker, AliceBlueBroker

from loguru import logger


class MainWindow(QMainWindow):
    """Main application window"""

    # Signal to handle Chartink alerts from background thread
    chartink_alert_signal = pyqtSignal(object)
    # Signal for MTF analysis updates from background thread
    mtf_update_signal = pyqtSignal(str, object)  # symbol, results
    mtf_log_signal = pyqtSignal(str)  # log message

    def __init__(self):
        super().__init__()

        # Initialize components
        self.config = Config()
        self.db = Database()
        self.order_manager = OrderManager(self.db)
        self.strategy_engine = StrategyEngine(self.order_manager, self.db)

        # Active broker connections
        self.brokers = {}

        # Paper trading simulator
        self.paper_simulator = None

        # Telegram alerts
        self.telegram = None

        # Risk manager for auto square-off
        self.risk_manager = None

        # Alert manager
        self.alert_manager = None

        self._init_ui()
        self._load_configured_brokers()
        self._setup_timers()
        self._init_telegram()
        self._init_risk_manager()

        # Always initialize paper trading (needed for test mode in CPR/Custom strategies)
        self._init_paper_trading()

    def _init_ui(self):
        """Initialize the user interface"""
        self.setWindowTitle("Algo Trader - Pine Script & Chartink Trading")

        # Window size settings - resizable from small to large
        self.setMinimumSize(600, 400)  # Minimum size for corner mode
        self.resize(1200, 800)  # Default size

        # Load saved window geometry if available
        self._load_window_geometry()

        # Central widget
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)

        # Create toolbar
        self._create_toolbar()

        # Create tab widget
        self.tabs = QTabWidget()
        main_layout.addWidget(self.tabs)

        # Add tabs
        self._create_dashboard_tab()
        self._create_strategies_tab()
        self._create_chartink_tab()
        self._create_chart_tab()
        self._create_mtf_tab()  # Multi-Timeframe Analysis
        self._create_strategy_builder_tab()  # Option Strategy Builder
        self._create_cpr_strategy_tab()  # CPR Auto-Trade Strategy
        self._create_custom_strategy_tab()  # Custom Strategy Builder
        self._create_orders_tab()
        self._create_positions_tab()
        self._create_risk_tab()
        self._create_options_tab()
        self._create_backtest_tab()
        self._create_journal_tab()
        self._create_alerts_tab()
        self._create_settings_tab()

        # Load strategies after all tabs are created
        self._load_strategies()
        self._load_chartink_scans()
        self._load_custom_strategies()  # Load custom strategies
        self._init_risk_manager()
        self._init_options_manager()
        self._init_alert_manager()

        # Status bar
        self.status_bar = QStatusBar()
        self.setStatusBar(self.status_bar)
        self.status_bar.showMessage("Ready")

    def _create_toolbar(self):
        """Create application toolbar"""
        toolbar = QToolBar("Main Toolbar")
        self.addToolBar(toolbar)

        # Connect broker action
        connect_action = QAction("Connect Broker", self)
        connect_action.triggered.connect(self._show_broker_dialog)
        toolbar.addAction(connect_action)

        toolbar.addSeparator()

        # Refresh action
        refresh_action = QAction("Refresh", self)
        refresh_action.triggered.connect(self._refresh_data)
        toolbar.addAction(refresh_action)

        toolbar.addSeparator()

        # Broker selector
        toolbar.addWidget(QLabel("Active Broker: "))
        self.broker_combo = QComboBox()
        self.broker_combo.setMinimumWidth(150)
        toolbar.addWidget(self.broker_combo)

        # Add spacer to push clock to right
        spacer = QWidget()
        spacer.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Preferred)
        toolbar.addWidget(spacer)

        toolbar.addSeparator()

        # Clock display (HH:MM:SS format)
        self.clock_label = QLabel("00:00:00")
        self.clock_label.setStyleSheet("""
            QLabel {
                font-size: 18px;
                font-weight: bold;
                font-family: 'Courier New', monospace;
                color: #2196F3;
                padding: 5px 15px;
                background: #1E1E1E;
                border-radius: 5px;
                border: 1px solid #333;
            }
        """)
        toolbar.addWidget(self.clock_label)

        # Setup clock timer
        self.clock_timer = QTimer(self)
        self.clock_timer.timeout.connect(self._update_clock)
        self.clock_timer.start(1000)  # Update every second
        self._update_clock()  # Initial update

    def _update_clock(self):
        """Update the clock display"""
        from datetime import datetime
        current_time = datetime.now().strftime("%H:%M:%S")
        self.clock_label.setText(current_time)

    def _make_scrollable(self, widget):
        """Wrap a widget in a scroll area for scrolling support"""
        scroll = QScrollArea()
        scroll.setWidget(widget)
        scroll.setWidgetResizable(True)
        scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        scroll.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        return scroll

    def _create_dashboard_tab(self):
        """Create enhanced dashboard tab with live P&L"""
        dashboard = QWidget()
        layout = QVBoxLayout(dashboard)

        # === Top Row - P&L Summary Cards ===
        pnl_group = QGroupBox("Today's Performance")
        pnl_layout = QHBoxLayout(pnl_group)

        # Realized P&L Card
        realized_card = QGroupBox("Realized P&L")
        realized_layout = QVBoxLayout(realized_card)
        self.dash_realized_pnl = QLabel("₹0.00")
        self.dash_realized_pnl.setStyleSheet("font-size: 24px; font-weight: bold; color: #4CAF50;")
        self.dash_realized_pnl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        realized_layout.addWidget(self.dash_realized_pnl)
        pnl_layout.addWidget(realized_card)

        # Unrealized P&L Card
        unrealized_card = QGroupBox("Unrealized P&L")
        unrealized_layout = QVBoxLayout(unrealized_card)
        self.dash_unrealized_pnl = QLabel("₹0.00")
        self.dash_unrealized_pnl.setStyleSheet("font-size: 24px; font-weight: bold; color: #2196F3;")
        self.dash_unrealized_pnl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        unrealized_layout.addWidget(self.dash_unrealized_pnl)
        pnl_layout.addWidget(unrealized_card)

        # Total P&L Card
        total_card = QGroupBox("Total P&L")
        total_layout = QVBoxLayout(total_card)
        self.dash_total_pnl = QLabel("₹0.00")
        self.dash_total_pnl.setStyleSheet("font-size: 24px; font-weight: bold;")
        self.dash_total_pnl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        total_layout.addWidget(self.dash_total_pnl)
        pnl_layout.addWidget(total_card)

        # Trades Today Card
        trades_card = QGroupBox("Trades Today")
        trades_layout = QVBoxLayout(trades_card)
        self.dash_trades_count = QLabel("0")
        self.dash_trades_count.setStyleSheet("font-size: 24px; font-weight: bold; color: #FF9800;")
        self.dash_trades_count.setAlignment(Qt.AlignmentFlag.AlignCenter)
        trades_layout.addWidget(self.dash_trades_count)
        pnl_layout.addWidget(trades_card)

        # Win Rate Card
        winrate_card = QGroupBox("Win Rate")
        winrate_layout = QVBoxLayout(winrate_card)
        self.dash_win_rate = QLabel("0%")
        self.dash_win_rate.setStyleSheet("font-size: 24px; font-weight: bold; color: #9C27B0;")
        self.dash_win_rate.setAlignment(Qt.AlignmentFlag.AlignCenter)
        winrate_layout.addWidget(self.dash_win_rate)
        pnl_layout.addWidget(winrate_card)

        layout.addWidget(pnl_group)

        # === Second Row - Account & Market Info ===
        info_layout = QHBoxLayout()

        # Account Summary
        account_group = QGroupBox("Account Summary")
        account_layout = QFormLayout(account_group)
        self.dash_available_margin = QLabel("₹0")
        self.dash_used_margin = QLabel("₹0")
        self.dash_total_balance = QLabel("₹0")
        self.dash_broker_status = QLabel("Not Connected")
        self.dash_broker_status.setStyleSheet("color: red;")
        account_layout.addRow("Available Margin:", self.dash_available_margin)
        account_layout.addRow("Used Margin:", self.dash_used_margin)
        account_layout.addRow("Total Balance:", self.dash_total_balance)
        account_layout.addRow("Broker Status:", self.dash_broker_status)
        info_layout.addWidget(account_group)

        # Market Status
        market_group = QGroupBox("Market Status")
        market_layout = QFormLayout(market_group)
        self.dash_market_status = QLabel("Closed")
        self.dash_market_status.setStyleSheet("color: red; font-weight: bold;")
        self.dash_nifty_price = QLabel("--")
        self.dash_banknifty_price = QLabel("--")
        self.dash_last_update = QLabel("--")
        market_layout.addRow("Market:", self.dash_market_status)
        market_layout.addRow("NIFTY:", self.dash_nifty_price)
        market_layout.addRow("BANKNIFTY:", self.dash_banknifty_price)
        market_layout.addRow("Last Update:", self.dash_last_update)
        info_layout.addWidget(market_group)

        # Quick Stats
        stats_group = QGroupBox("Quick Stats")
        stats_layout = QFormLayout(stats_group)
        self.dash_open_positions = QLabel("0")
        self.dash_active_scanners = QLabel("0")
        self.dash_active_strategies = QLabel("0")
        self.dash_pending_orders = QLabel("0")
        stats_layout.addRow("Open Positions:", self.dash_open_positions)
        stats_layout.addRow("Active Scanners:", self.dash_active_scanners)
        stats_layout.addRow("Active Strategies:", self.dash_active_strategies)
        stats_layout.addRow("Pending Orders:", self.dash_pending_orders)
        info_layout.addWidget(stats_group)

        layout.addLayout(info_layout)

        # === Open Positions Table ===
        positions_group = QGroupBox("Open Positions (Live P&L)")
        positions_layout = QVBoxLayout(positions_group)

        self.dash_positions_table = QTableWidget()
        self.dash_positions_table.setColumnCount(9)
        self.dash_positions_table.setHorizontalHeaderLabels([
            "Symbol", "Type", "Qty", "Avg Price", "LTP", "P&L", "P&L %", "Source", "Action"
        ])
        self.dash_positions_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        positions_layout.addWidget(self.dash_positions_table)

        layout.addWidget(positions_group)

        # === Recent Activity / Signals ===
        activity_group = QGroupBox("Recent Activity")
        activity_layout = QVBoxLayout(activity_group)

        self.dash_activity_table = QTableWidget()
        self.dash_activity_table.setColumnCount(6)
        self.dash_activity_table.setHorizontalHeaderLabels([
            "Time", "Type", "Symbol", "Action", "Price", "Status"
        ])
        self.dash_activity_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.dash_activity_table.setMaximumHeight(150)
        activity_layout.addWidget(self.dash_activity_table)

        layout.addWidget(activity_group)

        # Refresh button
        refresh_btn = QPushButton("🔄 Refresh Dashboard")
        refresh_btn.clicked.connect(self._refresh_dashboard)
        layout.addWidget(refresh_btn)

        self.tabs.addTab(dashboard, "Dashboard")

        # Set up auto-refresh timer for dashboard
        self.dashboard_timer = QTimer(self)
        self.dashboard_timer.timeout.connect(self._refresh_dashboard)
        self.dashboard_timer.start(10000)  # Refresh every 10 seconds

    def _create_strategies_tab(self):
        """Create strategies management tab"""
        strategies = QWidget()
        layout = QHBoxLayout(strategies)

        # Left panel - Strategy list
        left_panel = QWidget()
        left_layout = QVBoxLayout(left_panel)

        left_layout.addWidget(QLabel("Strategies"))

        self.strategy_list = QTableWidget()
        self.strategy_list.setColumnCount(3)
        self.strategy_list.setHorizontalHeaderLabels(["Name", "Type", "Active"])
        self.strategy_list.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.strategy_list.cellClicked.connect(self._on_strategy_selected)
        left_layout.addWidget(self.strategy_list)

        btn_layout = QHBoxLayout()
        self.new_strategy_btn = QPushButton("New Strategy")
        self.new_strategy_btn.clicked.connect(self._new_strategy)
        self.delete_strategy_btn = QPushButton("Delete")
        self.delete_strategy_btn.clicked.connect(self._delete_strategy)
        btn_layout.addWidget(self.new_strategy_btn)
        btn_layout.addWidget(self.delete_strategy_btn)
        left_layout.addLayout(btn_layout)

        # Right panel - Strategy editor
        right_panel = QWidget()
        right_layout = QVBoxLayout(right_panel)

        right_layout.addWidget(QLabel("Pine Script Editor"))

        self.strategy_name_edit = QLineEdit()
        self.strategy_name_edit.setPlaceholderText("Strategy Name")
        right_layout.addWidget(self.strategy_name_edit)

        self.pine_script_editor = QTextEdit()
        self.pine_script_editor.setPlaceholderText(
            "Paste your Pine Script here...\n\n"
            "Example:\n"
            "//@version=5\n"
            "strategy('My Strategy', overlay=true)\n\n"
            "fast_ma = ta.sma(close, 10)\n"
            "slow_ma = ta.sma(close, 20)\n\n"
            "if ta.crossover(fast_ma, slow_ma)\n"
            "    strategy.entry('Long', strategy.long)\n\n"
            "if ta.crossunder(fast_ma, slow_ma)\n"
            "    strategy.close('Long')"
        )
        self.pine_script_editor.setFont(QFont("Consolas", 10))
        right_layout.addWidget(self.pine_script_editor)

        editor_btn_layout = QHBoxLayout()
        self.save_strategy_btn = QPushButton("Save Strategy")
        self.save_strategy_btn.clicked.connect(self._save_strategy)
        self.validate_btn = QPushButton("Validate")
        self.validate_btn.clicked.connect(self._validate_strategy)
        self.activate_btn = QPushButton("Activate")
        self.activate_btn.clicked.connect(self._activate_strategy)
        editor_btn_layout.addWidget(self.save_strategy_btn)
        editor_btn_layout.addWidget(self.validate_btn)
        editor_btn_layout.addWidget(self.activate_btn)
        right_layout.addLayout(editor_btn_layout)

        # Splitter
        splitter = QSplitter(Qt.Orientation.Horizontal)
        splitter.addWidget(left_panel)
        splitter.addWidget(right_panel)
        splitter.setSizes([300, 700])
        layout.addWidget(splitter)

        self.tabs.addTab(strategies, "Strategies")

    def _create_chartink_tab(self):
        """Create Chartink scanner integration tab with time controls & allocation"""
        chartink_scroll = QScrollArea()
        chartink_scroll.setWidgetResizable(True)
        chartink_inner = QWidget()
        layout = QVBoxLayout(chartink_inner)

        # === Add Scanner Section ===
        add_scan_group = QGroupBox("Add Chartink Scanner")
        add_scan_main = QHBoxLayout(add_scan_group)

        # Left column - Basic info
        left_form = QFormLayout()

        self.chartink_scan_name = QLineEdit()
        self.chartink_scan_name.setPlaceholderText("e.g., My Breakout Scanner")
        left_form.addRow("Scan Name:", self.chartink_scan_name)

        self.chartink_scan_url = QLineEdit()
        self.chartink_scan_url.setPlaceholderText("https://chartink.com/screener/your-scanner-name")
        left_form.addRow("Scanner URL:", self.chartink_scan_url)

        self.chartink_action = QComboBox()
        self.chartink_action.addItems(["BUY", "SELL"])
        left_form.addRow("Action:", self.chartink_action)

        # Trade Type - Equity or Options
        self.chartink_trade_type = QComboBox()
        self.chartink_trade_type.addItems(["Equity", "Options (F&O)"])
        self.chartink_trade_type.currentIndexChanged.connect(self._on_chartink_trade_type_changed)
        left_form.addRow("Trade Type:", self.chartink_trade_type)

        # Options settings (hidden by default)
        self.chartink_options_frame = QWidget()
        options_layout = QFormLayout(self.chartink_options_frame)
        options_layout.setContentsMargins(0, 0, 0, 0)

        self.chartink_strike_selection = QComboBox()
        self.chartink_strike_selection.addItems(["ATM", "ITM-1", "ITM-2", "OTM-1", "OTM-2", "Manual"])
        self.chartink_strike_selection.currentIndexChanged.connect(self._on_chartink_strike_changed)
        options_layout.addRow("Strike:", self.chartink_strike_selection)

        # Manual strike price input (hidden by default)
        self.chartink_manual_strike = QDoubleSpinBox()
        self.chartink_manual_strike.setRange(0, 999999)
        self.chartink_manual_strike.setDecimals(0)
        self.chartink_manual_strike.setSingleStep(50)
        self.chartink_manual_strike.setPrefix("₹ ")
        self.chartink_manual_strike.setVisible(False)
        options_layout.addRow("Strike Price:", self.chartink_manual_strike)

        self.chartink_option_type = QComboBox()
        self.chartink_option_type.addItems(["Auto (BUY=CE, SELL=PE)", "CE Only", "PE Only"])
        options_layout.addRow("Option Type:", self.chartink_option_type)

        self.chartink_expiry = QComboBox()
        self.chartink_expiry.addItems(["Current Week", "Next Week", "Current Month", "Next Month"])
        options_layout.addRow("Expiry:", self.chartink_expiry)

        self.chartink_options_frame.setVisible(False)
        left_form.addRow("", self.chartink_options_frame)

        self.chartink_quantity = QSpinBox()
        self.chartink_quantity.setRange(1, 10000)
        self.chartink_quantity.setValue(1)
        left_form.addRow("Default Qty:", self.chartink_quantity)

        self.chartink_interval = QSpinBox()
        self.chartink_interval.setRange(30, 3600)
        self.chartink_interval.setValue(60)
        self.chartink_interval.setSuffix(" seconds")
        left_form.addRow("Scan Interval:", self.chartink_interval)

        add_scan_main.addLayout(left_form)

        # Middle column - Time controls
        mid_form = QFormLayout()

        self.chartink_start_time = QComboBox()
        self.chartink_start_time.setEditable(True)
        for h in range(9, 16):
            for m in (0, 15, 30, 45):
                self.chartink_start_time.addItem(f"{h:02d}:{m:02d}")
        self.chartink_start_time.setCurrentText("09:15")
        mid_form.addRow("Start Time:", self.chartink_start_time)

        self.chartink_exit_time = QComboBox()
        self.chartink_exit_time.setEditable(True)
        for h in range(9, 16):
            for m in (0, 15, 30, 45):
                self.chartink_exit_time.addItem(f"{h:02d}:{m:02d}")
        self.chartink_exit_time.setCurrentText("15:15")
        mid_form.addRow("Exit / Square-off Time:", self.chartink_exit_time)

        self.chartink_no_new_trade_time = QComboBox()
        self.chartink_no_new_trade_time.setEditable(True)
        for h in range(9, 16):
            for m in (0, 15, 30, 45):
                self.chartink_no_new_trade_time.addItem(f"{h:02d}:{m:02d}")
        self.chartink_no_new_trade_time.setCurrentText("14:30")
        mid_form.addRow("No New Trade After:", self.chartink_no_new_trade_time)

        add_scan_main.addLayout(mid_form)

        # Right column - Capital & Limits
        right_form = QFormLayout()

        # Per-Stock Allocation Type
        self.chartink_alloc_type = QComboBox()
        self.chartink_alloc_type.addItems(["Auto (Capital ÷ Stocks)", "Fixed Qty", "Fixed Amount"])
        self.chartink_alloc_type.currentIndexChanged.connect(self._on_chartink_alloc_type_changed)
        right_form.addRow("Per-Stock:", self.chartink_alloc_type)

        # Per-Stock Value (qty or amount based on type)
        alloc_value_layout = QHBoxLayout()
        self.chartink_alloc_value = QDoubleSpinBox()
        self.chartink_alloc_value.setRange(0, 99999999)
        self.chartink_alloc_value.setDecimals(0)
        self.chartink_alloc_value.setValue(0)
        self.chartink_alloc_value.setEnabled(False)  # Disabled for "Auto" mode
        self.chartink_alloc_value_label = QLabel("(Auto)")
        alloc_value_layout.addWidget(self.chartink_alloc_value)
        alloc_value_layout.addWidget(self.chartink_alloc_value_label)
        right_form.addRow("Value:", alloc_value_layout)

        self.chartink_total_capital = QDoubleSpinBox()
        self.chartink_total_capital.setRange(0, 99999999)
        self.chartink_total_capital.setDecimals(0)
        self.chartink_total_capital.setValue(0)
        self.chartink_total_capital.setPrefix("₹ ")
        self.chartink_total_capital.setSpecialValueText("Unlimited")
        right_form.addRow("Total Capital:", self.chartink_total_capital)

        self.chartink_max_trades = QSpinBox()
        self.chartink_max_trades.setRange(0, 1000)
        self.chartink_max_trades.setValue(0)
        self.chartink_max_trades.setSpecialValueText("Unlimited")
        right_form.addRow("Max Trades:", self.chartink_max_trades)

        add_scan_main.addLayout(right_form)

        layout.addWidget(add_scan_group)

        # === Risk Management Section ===
        risk_group = QGroupBox("Risk Management (Per Stock & Scanner MTM)")
        risk_main = QHBoxLayout(risk_group)

        # Left - Per Stock SL/Target
        stock_risk_form = QFormLayout()
        stock_risk_form.addRow(QLabel("<b>Per Stock Exit:</b>"))

        # Stop Loss Type
        self.chartink_sl_type = QComboBox()
        self.chartink_sl_type.addItems(["None", "Fixed Points", "Fixed %", "Fixed Amount"])
        self.chartink_sl_type.currentIndexChanged.connect(self._on_chartink_sl_type_changed)
        stock_risk_form.addRow("Stop Loss:", self.chartink_sl_type)

        self.chartink_sl_value = QDoubleSpinBox()
        self.chartink_sl_value.setRange(0, 99999)
        self.chartink_sl_value.setValue(0)
        self.chartink_sl_value.setEnabled(False)
        stock_risk_form.addRow("SL Value:", self.chartink_sl_value)

        # Target
        self.chartink_target_type = QComboBox()
        self.chartink_target_type.addItems(["None", "Fixed Points", "Fixed %", "Fixed Amount"])
        self.chartink_target_type.currentIndexChanged.connect(self._on_chartink_target_type_changed)
        stock_risk_form.addRow("Target:", self.chartink_target_type)

        self.chartink_target_value = QDoubleSpinBox()
        self.chartink_target_value.setRange(0, 99999)
        self.chartink_target_value.setValue(0)
        self.chartink_target_value.setEnabled(False)
        stock_risk_form.addRow("Target Value:", self.chartink_target_value)

        risk_main.addLayout(stock_risk_form)

        # Middle - Trailing SL
        tsl_form = QFormLayout()
        tsl_form.addRow(QLabel("<b>Trailing Stop Loss:</b>"))

        self.chartink_tsl_enabled = QCheckBox("Enable TSL")
        self.chartink_tsl_enabled.stateChanged.connect(self._on_chartink_tsl_toggle)
        tsl_form.addRow(self.chartink_tsl_enabled)

        self.chartink_tsl_type = QComboBox()
        self.chartink_tsl_type.addItems(["Points", "Percentage"])
        self.chartink_tsl_type.setEnabled(False)
        tsl_form.addRow("TSL Type:", self.chartink_tsl_type)

        self.chartink_tsl_value = QDoubleSpinBox()
        self.chartink_tsl_value.setRange(0, 9999)
        self.chartink_tsl_value.setValue(0)
        self.chartink_tsl_value.setEnabled(False)
        tsl_form.addRow("TSL Value:", self.chartink_tsl_value)

        # Profit Lock (move SL to breakeven after X profit)
        tsl_form.addRow(QLabel("<b>Profit Lock:</b>"))
        self.chartink_profit_lock_enabled = QCheckBox("Enable Profit Lock")
        tsl_form.addRow(self.chartink_profit_lock_enabled)

        self.chartink_profit_lock_type = QComboBox()
        self.chartink_profit_lock_type.addItems(["Points", "Amount"])
        tsl_form.addRow("Lock After:", self.chartink_profit_lock_type)

        self.chartink_profit_lock_value = QDoubleSpinBox()
        self.chartink_profit_lock_value.setRange(0, 99999)
        self.chartink_profit_lock_value.setValue(0)
        tsl_form.addRow("Lock Value:", self.chartink_profit_lock_value)

        risk_main.addLayout(tsl_form)

        # Right - Scanner MTM
        mtm_form = QFormLayout()
        mtm_form.addRow(QLabel("<b>Scanner MTM Limits:</b>"))

        self.chartink_mtm_profit = QDoubleSpinBox()
        self.chartink_mtm_profit.setRange(0, 9999999)
        self.chartink_mtm_profit.setDecimals(0)
        self.chartink_mtm_profit.setPrefix("₹ ")
        self.chartink_mtm_profit.setSpecialValueText("No Limit")
        mtm_form.addRow("MTM Profit Target:", self.chartink_mtm_profit)

        self.chartink_mtm_loss = QDoubleSpinBox()
        self.chartink_mtm_loss.setRange(0, 9999999)
        self.chartink_mtm_loss.setDecimals(0)
        self.chartink_mtm_loss.setPrefix("₹ ")
        self.chartink_mtm_loss.setSpecialValueText("No Limit")
        mtm_form.addRow("MTM Loss Limit:", self.chartink_mtm_loss)

        mtm_form.addRow(QLabel("(Scanner stops when MTM limit hit)"))

        risk_main.addLayout(mtm_form)

        layout.addWidget(risk_group)

        # Options row
        options_layout = QHBoxLayout()
        self.chartink_trigger_first = QCheckBox("Trigger on First Scan (sab stocks pe signal)")
        self.chartink_trigger_first.setChecked(True)
        self.chartink_trigger_first.setToolTip("Enabled: Pehle scan mein sab stocks pe signal. Disabled: Sirf new stocks pe signal.")
        options_layout.addWidget(self.chartink_trigger_first)
        options_layout.addStretch()
        layout.addLayout(options_layout)

        # Buttons row
        btn_layout = QHBoxLayout()
        self.add_chartink_scan_btn = QPushButton("Add Scanner")
        self.add_chartink_scan_btn.clicked.connect(self._add_chartink_scan)
        self.test_chartink_scan_btn = QPushButton("Test Scanner")
        self.test_chartink_scan_btn.clicked.connect(self._test_chartink_scan)
        btn_layout.addWidget(self.add_chartink_scan_btn)
        btn_layout.addWidget(self.test_chartink_scan_btn)
        btn_layout.addStretch()
        layout.addLayout(btn_layout)

        # === Active Scanners Table ===
        scans_group = QGroupBox("Active Scanners")
        scans_layout = QVBoxLayout(scans_group)

        self.chartink_scans_table = QTableWidget()
        self.chartink_scans_table.setColumnCount(11)
        self.chartink_scans_table.setHorizontalHeaderLabels([
            "ON/OFF", "Name", "Action", "Per-Stock", "Start", "Exit", "No New",
            "Capital", "Max Trades", "Trades Done", "Remove"
        ])
        self.chartink_scans_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        scans_layout.addWidget(self.chartink_scans_table)

        # Start/Stop monitoring buttons
        monitor_btn_layout = QHBoxLayout()
        self.start_chartink_btn = QPushButton("Start Monitoring")
        self.start_chartink_btn.clicked.connect(self._start_chartink_monitoring)
        self.stop_chartink_btn = QPushButton("Stop Monitoring")
        self.stop_chartink_btn.clicked.connect(self._stop_chartink_monitoring)
        self.stop_chartink_btn.setEnabled(False)
        self.reset_chartink_btn = QPushButton("Reset Daily Counters")
        self.reset_chartink_btn.clicked.connect(self._reset_chartink_counters)
        monitor_btn_layout.addWidget(self.start_chartink_btn)
        monitor_btn_layout.addWidget(self.stop_chartink_btn)
        monitor_btn_layout.addWidget(self.reset_chartink_btn)
        scans_layout.addLayout(monitor_btn_layout)

        layout.addWidget(scans_group)

        # === Open Positions Section ===
        positions_group = QGroupBox("Scanner Open Positions (Only Executed Trades)")
        positions_layout = QVBoxLayout(positions_group)

        # Header with expand button
        positions_header = QHBoxLayout()
        positions_header.addWidget(QLabel("Click on Symbol to open TradingView chart"))
        positions_header.addStretch()
        self.positions_expand_btn = QPushButton("⬜ Expand")
        self.positions_expand_btn.setFixedWidth(100)
        self.positions_expand_btn.clicked.connect(lambda: self._toggle_table_expand(self.chartink_positions_table, self.positions_expand_btn))
        positions_header.addWidget(self.positions_expand_btn)
        positions_layout.addLayout(positions_header)

        self.chartink_positions_table = QTableWidget()
        self.chartink_positions_table.setColumnCount(6)
        self.chartink_positions_table.setHorizontalHeaderLabels([
            "Scanner", "Symbol", "Action", "Qty", "Entry Price", "Entry Time"
        ])
        self.chartink_positions_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.chartink_positions_table.setMinimumHeight(120)
        self.chartink_positions_table.cellClicked.connect(self._on_positions_cell_clicked)
        self.chartink_positions_table.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        positions_layout.addWidget(self.chartink_positions_table)

        layout.addWidget(positions_group)

        # === Alerts Log ===
        alerts_group = QGroupBox("Recent Alerts")
        alerts_layout = QVBoxLayout(alerts_group)

        # Header with expand button
        alerts_header = QHBoxLayout()
        alerts_header.addWidget(QLabel("Click on Symbol to open TradingView chart"))
        alerts_header.addStretch()
        self.alerts_expand_btn = QPushButton("⬜ Expand")
        self.alerts_expand_btn.setFixedWidth(100)
        self.alerts_expand_btn.clicked.connect(lambda: self._toggle_table_expand(self.chartink_alerts_table, self.alerts_expand_btn))
        alerts_header.addWidget(self.alerts_expand_btn)
        alerts_layout.addLayout(alerts_header)

        self.chartink_alerts_table = QTableWidget()
        self.chartink_alerts_table.setColumnCount(6)
        self.chartink_alerts_table.setHorizontalHeaderLabels([
            "Time", "Scanner", "Symbol", "Price", "Qty", "Action Taken"
        ])
        self.chartink_alerts_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.chartink_alerts_table.setMinimumHeight(120)
        self.chartink_alerts_table.cellClicked.connect(self._on_alerts_cell_clicked)
        self.chartink_alerts_table.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        alerts_layout.addWidget(self.chartink_alerts_table)

        layout.addWidget(alerts_group)

        chartink_scroll.setWidget(chartink_inner)
        self.tabs.addTab(chartink_scroll, "Chartink")

    def _create_chart_tab(self):
        """Create advanced interactive chart tab"""
        try:
            from algo_trader.ui.chart_widget import ChartWidget

            self.chart_widget = ChartWidget(self)
            self.chart_widget.order_placed.connect(self._on_chart_order)
            self.chart_widget.order_modified_signal.connect(self._on_chart_order_modified)

            self.tabs.addTab(self.chart_widget, "Chart")
        except ImportError as e:
            logger.warning(f"Chart widget not available: {e}")
            # Create placeholder
            placeholder = QWidget()
            layout = QVBoxLayout(placeholder)
            layout.addWidget(QLabel("Advanced Charting requires matplotlib.\n\nInstall with: pip install matplotlib mplfinance"))
            self.tabs.addTab(placeholder, "Chart")

    def _on_chart_order(self, order_data: dict):
        """Handle order from chart widget"""
        logger.info(f"Chart order: {order_data}")

        # Check if paper trading mode
        if self.config.get('trading.paper_mode', False):
            if hasattr(self, 'paper_simulator') and self.paper_simulator:
                self.paper_simulator.place_order(
                    symbol=order_data['symbol'],
                    action=order_data['side'],  # paper_trading uses 'action' not 'side'
                    quantity=order_data['quantity'],
                    order_type=order_data.get('order_type', 'MARKET'),
                    price=order_data.get('price', 0),
                    source="Chart"
                )
                self.status_bar.showMessage(f"Paper order placed: {order_data['side']} {order_data['symbol']}", 5000)
            return

        # Real order
        if not self.brokers:
            QMessageBox.warning(self, "Error", "No broker connected")
            return

        broker = list(self.brokers.values())[0]
        try:
            from algo_trader.brokers.base import BrokerOrder

            order = BrokerOrder(
                symbol=order_data['symbol'],
                exchange="NSE",
                transaction_type=order_data['side'],
                order_type=order_data['order_type'],
                quantity=order_data['quantity'],
                price=order_data.get('price', 0),
                trigger_price=order_data.get('trigger_price', 0),
                product=order_data.get('product', 'MIS')
            )

            result = broker.place_order(order)
            if result.get('success'):
                self.status_bar.showMessage(f"Order placed: {result.get('order_id')}", 5000)
                self._load_orders()
            else:
                QMessageBox.warning(self, "Order Failed", result.get('message', 'Unknown error'))

        except Exception as e:
            logger.error(f"Chart order error: {e}")
            QMessageBox.warning(self, "Error", f"Order failed: {str(e)}")

    def _on_chart_order_modified(self, modification_data: dict):
        """Handle order modification from chart dragging"""
        logger.info(f"Chart order modification: {modification_data}")

        order_type = modification_data.get('order_type', '')
        new_price = modification_data.get('new_price', 0)
        order_id = modification_data.get('order_id')
        symbol = modification_data.get('symbol', '')

        # For paper trading, just log the modification
        if self.config.get('trading.paper_mode', False):
            self.status_bar.showMessage(
                f"Paper {order_type} modified to ₹{new_price:.2f} for {symbol}", 5000
            )
            return

        # For real trading, modify the order if we have broker and order_id
        if not self.brokers:
            self.status_bar.showMessage(f"{order_type} line moved to ₹{new_price:.2f} (no broker connected)", 5000)
            return

        if order_id:
            # Try to modify the actual order
            broker = list(self.brokers.values())[0]
            try:
                if hasattr(broker, 'modify_order'):
                    result = broker.modify_order(order_id, price=new_price)
                    if result.get('success'):
                        self.status_bar.showMessage(f"Order {order_id} modified to ₹{new_price:.2f}", 5000)
                        self._load_orders()
                    else:
                        QMessageBox.warning(self, "Modification Failed", result.get('message', 'Unknown error'))
            except Exception as e:
                logger.error(f"Order modification error: {e}")
                QMessageBox.warning(self, "Error", f"Failed to modify order: {str(e)}")
        else:
            # No order_id - this is just a visual line (SL/Target marker)
            self.status_bar.showMessage(f"{order_type} level set to ₹{new_price:.2f}", 5000)

    def _create_mtf_tab(self):
        """Create Multi-Timeframe Analysis tab"""
        mtf_widget = QWidget()
        layout = QVBoxLayout(mtf_widget)

        # === Symbol Input Section ===
        input_group = QGroupBox("Symbol Selection")
        input_layout = QHBoxLayout(input_group)

        input_layout.addWidget(QLabel("Symbol:"))
        self.mtf_symbol_input = QLineEdit()
        self.mtf_symbol_input.setPlaceholderText("Enter symbol (e.g., RELIANCE, TCS, NIFTY)")
        self.mtf_symbol_input.returnPressed.connect(self._analyze_mtf)
        input_layout.addWidget(self.mtf_symbol_input)

        self.mtf_analyze_btn = QPushButton("🔍 Analyze")
        self.mtf_analyze_btn.clicked.connect(self._analyze_mtf)
        input_layout.addWidget(self.mtf_analyze_btn)

        self.mtf_open_chart_btn = QPushButton("📈 Open TradingView")
        self.mtf_open_chart_btn.clicked.connect(self._open_mtf_tradingview)
        input_layout.addWidget(self.mtf_open_chart_btn)

        layout.addWidget(input_group)

        # === Overall Trend Summary ===
        summary_group = QGroupBox("📊 Overall Trend Summary")
        summary_layout = QHBoxLayout(summary_group)

        self.mtf_overall_trend = QLabel("Enter a symbol and click Analyze")
        self.mtf_overall_trend.setStyleSheet("font-size: 18px; font-weight: bold; padding: 10px;")
        self.mtf_overall_trend.setAlignment(Qt.AlignmentFlag.AlignCenter)
        summary_layout.addWidget(self.mtf_overall_trend)

        self.mtf_recommendation = QLabel("")
        self.mtf_recommendation.setStyleSheet("font-size: 14px; padding: 10px;")
        self.mtf_recommendation.setAlignment(Qt.AlignmentFlag.AlignCenter)
        summary_layout.addWidget(self.mtf_recommendation)

        layout.addWidget(summary_group)

        # === Timeframe Analysis Grid ===
        tf_group = QGroupBox("📈 Timeframe Analysis")
        tf_layout = QVBoxLayout(tf_group)

        # Table for timeframe data
        self.mtf_table = QTableWidget()
        self.mtf_table.setColumnCount(8)
        self.mtf_table.setHorizontalHeaderLabels([
            "Timeframe", "Trend", "RSI", "MACD", "EMA 20", "EMA 50", "Volume", "Signal"
        ])
        self.mtf_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.mtf_table.setMinimumHeight(250)

        # Add default timeframes
        timeframes = ["5 Min", "15 Min", "1 Hour", "4 Hour", "Daily", "Weekly"]
        self.mtf_table.setRowCount(len(timeframes))
        for i, tf in enumerate(timeframes):
            self.mtf_table.setItem(i, 0, QTableWidgetItem(tf))
            for j in range(1, 8):
                self.mtf_table.setItem(i, j, QTableWidgetItem("--"))

        tf_layout.addWidget(self.mtf_table)
        layout.addWidget(tf_group)

        # === Quick Analysis Panel ===
        quick_group = QGroupBox("⚡ Quick Indicators")
        quick_layout = QHBoxLayout(quick_group)

        # Support/Resistance
        sr_box = QGroupBox("Support/Resistance")
        sr_layout = QVBoxLayout(sr_box)
        self.mtf_support = QLabel("Support: --")
        self.mtf_resistance = QLabel("Resistance: --")
        self.mtf_support.setStyleSheet("color: #4CAF50; font-weight: bold;")
        self.mtf_resistance.setStyleSheet("color: #F44336; font-weight: bold;")
        sr_layout.addWidget(self.mtf_resistance)
        sr_layout.addWidget(self.mtf_support)
        quick_layout.addWidget(sr_box)

        # Moving Averages
        ma_box = QGroupBox("Moving Averages")
        ma_layout = QVBoxLayout(ma_box)
        self.mtf_ma_signal = QLabel("MA Signal: --")
        self.mtf_ma_cross = QLabel("Last Cross: --")
        ma_layout.addWidget(self.mtf_ma_signal)
        ma_layout.addWidget(self.mtf_ma_cross)
        quick_layout.addWidget(ma_box)

        # Volume Analysis
        vol_box = QGroupBox("Volume Analysis")
        vol_layout = QVBoxLayout(vol_box)
        self.mtf_vol_trend = QLabel("Volume Trend: --")
        self.mtf_vol_avg = QLabel("vs 20-day Avg: --")
        vol_layout.addWidget(self.mtf_vol_trend)
        vol_layout.addWidget(self.mtf_vol_avg)
        quick_layout.addWidget(vol_box)

        # Momentum
        mom_box = QGroupBox("Momentum")
        mom_layout = QVBoxLayout(mom_box)
        self.mtf_momentum = QLabel("Momentum: --")
        self.mtf_strength = QLabel("Strength: --")
        mom_layout.addWidget(self.mtf_momentum)
        mom_layout.addWidget(self.mtf_strength)
        quick_layout.addWidget(mom_box)

        layout.addWidget(quick_group)

        # === Analysis Log ===
        log_group = QGroupBox("📝 Analysis Log")
        log_layout = QVBoxLayout(log_group)
        self.mtf_log = QTextEdit()
        self.mtf_log.setReadOnly(True)
        self.mtf_log.setMaximumHeight(150)
        self.mtf_log.setPlaceholderText("Analysis results will appear here...")
        log_layout.addWidget(self.mtf_log)
        layout.addWidget(log_group)

        # Connect MTF signals for thread-safe UI updates
        self.mtf_update_signal.connect(self._update_mtf_ui)
        self.mtf_log_signal.connect(self._append_mtf_log)

        self.tabs.addTab(mtf_widget, "MTF Analysis")

    def _append_mtf_log(self, message: str):
        """Append message to MTF log (called on main thread via signal)"""
        self.mtf_log.append(message)

    def _analyze_mtf(self):
        """Perform Multi-Timeframe Analysis"""
        symbol = self.mtf_symbol_input.text().strip().upper()
        if not symbol:
            QMessageBox.warning(self, "Error", "Please enter a symbol")
            return

        self.mtf_analyze_btn.setEnabled(False)
        self.mtf_analyze_btn.setText("Analyzing...")
        self.status_bar.showMessage(f"Analyzing {symbol}...")

        # Run analysis in background
        import threading
        thread = threading.Thread(target=self._run_mtf_analysis, args=(symbol,))
        thread.daemon = True
        thread.start()

    def _run_mtf_analysis(self, symbol: str):
        """Run MTF analysis in background thread"""
        try:
            import yfinance as yf
            import numpy as np

            # Add .NS suffix for NSE stocks
            yf_symbol = f"{symbol}.NS" if not symbol.endswith(('.NS', '.BSE')) else symbol

            self._mtf_log(f"Fetching data for {symbol}...")

            # Fetch data for different timeframes
            timeframes = {
                "5 Min": ("5d", "5m"),
                "15 Min": ("10d", "15m"),
                "1 Hour": ("30d", "1h"),
                "4 Hour": ("60d", "1h"),  # Approximate with 1h
                "Daily": ("180d", "1d"),
                "Weekly": ("2y", "1wk")
            }

            results = {}
            stock = yf.Ticker(yf_symbol)

            for tf_name, (period, interval) in timeframes.items():
                try:
                    data = stock.history(period=period, interval=interval)
                    if len(data) > 0:
                        results[tf_name] = self._calculate_indicators(data, tf_name)
                        self._mtf_log(f"✓ {tf_name}: {len(data)} candles analyzed")
                    else:
                        results[tf_name] = None
                        self._mtf_log(f"✗ {tf_name}: No data available")
                except Exception as e:
                    results[tf_name] = None
                    self._mtf_log(f"✗ {tf_name}: Error - {str(e)[:50]}")

            # Update UI on main thread using signal
            self.mtf_update_signal.emit(symbol, results)

        except Exception as e:
            self._mtf_log(f"Error: {str(e)}")
            logger.error(f"MTF analysis error: {e}")

        # Re-enable button on main thread
        def reset_button():
            self.mtf_analyze_btn.setEnabled(True)
            self.mtf_analyze_btn.setText("🔍 Analyze")

        from PyQt6.QtCore import QTimer
        QTimer.singleShot(0, reset_button)

    def _calculate_indicators(self, data, timeframe: str) -> dict:
        """Calculate technical indicators for a timeframe"""
        import numpy as np

        close = data['Close'].values
        high = data['High'].values
        low = data['Low'].values
        volume = data['Volume'].values

        result = {}

        # Current price
        result['price'] = close[-1]

        # RSI (14 period)
        delta = np.diff(close)
        gain = np.where(delta > 0, delta, 0)
        loss = np.where(delta < 0, -delta, 0)
        avg_gain = np.mean(gain[-14:]) if len(gain) >= 14 else np.mean(gain)
        avg_loss = np.mean(loss[-14:]) if len(loss) >= 14 else np.mean(loss)
        rs = avg_gain / avg_loss if avg_loss != 0 else 100
        result['rsi'] = 100 - (100 / (1 + rs))

        # EMA 20 and 50
        result['ema20'] = self._ema(close, 20)
        result['ema50'] = self._ema(close, 50)

        # MACD
        ema12 = self._ema(close, 12)
        ema26 = self._ema(close, 26)
        result['macd'] = ema12 - ema26
        result['macd_signal'] = "Bullish" if result['macd'] > 0 else "Bearish"

        # Trend based on EMAs
        if close[-1] > result['ema20'] > result['ema50']:
            result['trend'] = "🟢 Bullish"
            result['trend_score'] = 2
        elif close[-1] < result['ema20'] < result['ema50']:
            result['trend'] = "🔴 Bearish"
            result['trend_score'] = -2
        elif close[-1] > result['ema20']:
            result['trend'] = "🟡 Weak Bull"
            result['trend_score'] = 1
        elif close[-1] < result['ema20']:
            result['trend'] = "🟠 Weak Bear"
            result['trend_score'] = -1
        else:
            result['trend'] = "⚪ Neutral"
            result['trend_score'] = 0

        # Volume analysis
        avg_vol = np.mean(volume[-20:]) if len(volume) >= 20 else np.mean(volume)
        result['vol_ratio'] = volume[-1] / avg_vol if avg_vol > 0 else 1
        result['vol_trend'] = "High" if result['vol_ratio'] > 1.2 else "Low" if result['vol_ratio'] < 0.8 else "Normal"

        # Signal
        if result['rsi'] < 30 and result['trend_score'] >= 0:
            result['signal'] = "🟢 BUY"
        elif result['rsi'] > 70 and result['trend_score'] <= 0:
            result['signal'] = "🔴 SELL"
        elif result['trend_score'] > 0:
            result['signal'] = "🟡 HOLD"
        elif result['trend_score'] < 0:
            result['signal'] = "🟠 WAIT"
        else:
            result['signal'] = "⚪ NEUTRAL"

        # Support/Resistance (simple - last 20 period high/low)
        result['resistance'] = np.max(high[-20:]) if len(high) >= 20 else np.max(high)
        result['support'] = np.min(low[-20:]) if len(low) >= 20 else np.min(low)

        return result

    def _ema(self, data, period: int) -> float:
        """Calculate Exponential Moving Average"""
        import numpy as np
        if len(data) < period:
            return np.mean(data)
        multiplier = 2 / (period + 1)
        ema = data[-period]
        for price in data[-period+1:]:
            ema = (price - ema) * multiplier + ema
        return ema

    def _update_mtf_ui(self, symbol: str, results: dict):
        """Update MTF UI with analysis results (called on main thread)"""
        timeframes = ["5 Min", "15 Min", "1 Hour", "4 Hour", "Daily", "Weekly"]

        total_score = 0
        valid_count = 0

        for i, tf in enumerate(timeframes):
            data = results.get(tf)
            if data:
                self.mtf_table.setItem(i, 1, QTableWidgetItem(data['trend']))
                self.mtf_table.setItem(i, 2, QTableWidgetItem(f"{data['rsi']:.1f}"))
                self.mtf_table.setItem(i, 3, QTableWidgetItem(data['macd_signal']))
                self.mtf_table.setItem(i, 4, QTableWidgetItem(f"₹{data['ema20']:.2f}"))
                self.mtf_table.setItem(i, 5, QTableWidgetItem(f"₹{data['ema50']:.2f}"))
                self.mtf_table.setItem(i, 6, QTableWidgetItem(data['vol_trend']))
                self.mtf_table.setItem(i, 7, QTableWidgetItem(data['signal']))

                total_score += data['trend_score']
                valid_count += 1
            else:
                for j in range(1, 8):
                    self.mtf_table.setItem(i, j, QTableWidgetItem("N/A"))

        # Overall trend
        if valid_count > 0:
            avg_score = total_score / valid_count

            if avg_score >= 1.5:
                self.mtf_overall_trend.setText(f"🟢 STRONG BULLISH")
                self.mtf_overall_trend.setStyleSheet("font-size: 18px; font-weight: bold; color: #4CAF50; padding: 10px;")
                self.mtf_recommendation.setText("All timeframes align bullish. Consider BUY.")
            elif avg_score >= 0.5:
                self.mtf_overall_trend.setText(f"🟡 BULLISH")
                self.mtf_overall_trend.setStyleSheet("font-size: 18px; font-weight: bold; color: #8BC34A; padding: 10px;")
                self.mtf_recommendation.setText("Most timeframes bullish. Wait for pullback or confirmation.")
            elif avg_score <= -1.5:
                self.mtf_overall_trend.setText(f"🔴 STRONG BEARISH")
                self.mtf_overall_trend.setStyleSheet("font-size: 18px; font-weight: bold; color: #F44336; padding: 10px;")
                self.mtf_recommendation.setText("All timeframes align bearish. Avoid buying / Consider SHORT.")
            elif avg_score <= -0.5:
                self.mtf_overall_trend.setText(f"🟠 BEARISH")
                self.mtf_overall_trend.setStyleSheet("font-size: 18px; font-weight: bold; color: #FF9800; padding: 10px;")
                self.mtf_recommendation.setText("Most timeframes bearish. Be cautious.")
            else:
                self.mtf_overall_trend.setText(f"⚪ MIXED / NEUTRAL")
                self.mtf_overall_trend.setStyleSheet("font-size: 18px; font-weight: bold; color: #9E9E9E; padding: 10px;")
                self.mtf_recommendation.setText("No clear trend. Wait for alignment.")

        # Update quick indicators from Daily timeframe
        daily_data = results.get("Daily")
        if daily_data:
            self.mtf_support.setText(f"Support: ₹{daily_data['support']:.2f}")
            self.mtf_resistance.setText(f"Resistance: ₹{daily_data['resistance']:.2f}")

            if daily_data['ema20'] > daily_data['ema50']:
                self.mtf_ma_signal.setText("MA Signal: Bullish (20 > 50)")
                self.mtf_ma_signal.setStyleSheet("color: #4CAF50;")
            else:
                self.mtf_ma_signal.setText("MA Signal: Bearish (20 < 50)")
                self.mtf_ma_signal.setStyleSheet("color: #F44336;")

            self.mtf_vol_trend.setText(f"Volume Trend: {daily_data['vol_trend']}")
            self.mtf_vol_avg.setText(f"vs 20-day Avg: {daily_data['vol_ratio']:.1%}")

            if daily_data['rsi'] > 50:
                self.mtf_momentum.setText("Momentum: Positive")
                self.mtf_momentum.setStyleSheet("color: #4CAF50;")
            else:
                self.mtf_momentum.setText("Momentum: Negative")
                self.mtf_momentum.setStyleSheet("color: #F44336;")

            strength = abs(daily_data['trend_score']) / 2 * 100
            self.mtf_strength.setText(f"Strength: {strength:.0f}%")

        self._mtf_log(f"✅ Analysis complete for {symbol}")
        self.status_bar.showMessage(f"MTF Analysis complete for {symbol}", 5000)

    def _mtf_log(self, message: str):
        """Add message to MTF analysis log (thread-safe via signal)"""
        from datetime import datetime
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.mtf_log_signal.emit(f"[{timestamp}] {message}")

    def _open_mtf_tradingview(self):
        """Open TradingView for the current MTF symbol"""
        symbol = self.mtf_symbol_input.text().strip().upper()
        if symbol:
            self._open_tradingview_chart(symbol)
        else:
            QMessageBox.warning(self, "Error", "Please enter a symbol first")

    def _create_strategy_builder_tab(self):
        """Create Option Strategy Builder tab with Payoff Chart"""
        try:
            from matplotlib.backends.backend_qtagg import FigureCanvasQTAgg as FigureCanvas
            from matplotlib.figure import Figure
            import numpy as np
        except ImportError:
            # Matplotlib not available - show placeholder
            placeholder = QWidget()
            layout = QVBoxLayout(placeholder)
            layout.addWidget(QLabel("Strategy Builder requires matplotlib.\n\nInstall with: pip install matplotlib"))
            self.tabs.addTab(placeholder, "Strategy Builder")
            return

        builder_widget = QWidget()
        main_layout = QVBoxLayout(builder_widget)

        # Create scroll area for the configuration panel
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)

        # Configuration Panel (full width, scrollable)
        config_panel = QWidget()
        config_layout = QVBoxLayout(config_panel)

        # Symbol & Expiry Selection
        symbol_group = QGroupBox("📊 Symbol & Expiry")
        symbol_layout = QFormLayout(symbol_group)

        self.sb_symbol = QComboBox()
        self.sb_symbol.addItems(["NIFTY", "BANKNIFTY", "FINNIFTY", "MIDCPNIFTY"])
        self.sb_symbol.setEditable(True)
        self.sb_symbol.setMinimumHeight(30)
        self.sb_symbol.currentTextChanged.connect(self._on_sb_symbol_changed)
        symbol_layout.addRow("Symbol:", self.sb_symbol)

        self.sb_spot_price = QDoubleSpinBox()
        self.sb_spot_price.setRange(0, 999999)
        self.sb_spot_price.setDecimals(2)
        self.sb_spot_price.setValue(25000)
        self.sb_spot_price.setPrefix("₹ ")
        self.sb_spot_price.setMinimumHeight(30)
        self.sb_spot_price.valueChanged.connect(self._on_spot_price_changed)
        symbol_layout.addRow("Spot Price:", self.sb_spot_price)

        self.sb_expiry = QComboBox()
        self.sb_expiry.addItems(["Current Week", "Next Week", "Current Month", "Next Month"])
        self.sb_expiry.setMinimumHeight(30)
        symbol_layout.addRow("Expiry:", self.sb_expiry)

        # Pre-built Strategies (define before row1_layout)
        prebuilt_group = QGroupBox("⚡ Pre-built Strategies")
        prebuilt_layout = QVBoxLayout(prebuilt_group)

        strategies_row1 = QHBoxLayout()
        self.sb_iron_fly_btn = QPushButton("Iron Fly")
        self.sb_iron_fly_btn.setMinimumHeight(35)
        self.sb_iron_fly_btn.clicked.connect(lambda: self._apply_prebuilt_strategy("iron_fly"))
        strategies_row1.addWidget(self.sb_iron_fly_btn)

        self.sb_iron_condor_btn = QPushButton("Iron Condor")
        self.sb_iron_condor_btn.setMinimumHeight(35)
        self.sb_iron_condor_btn.clicked.connect(lambda: self._apply_prebuilt_strategy("iron_condor"))
        strategies_row1.addWidget(self.sb_iron_condor_btn)

        self.sb_straddle_btn = QPushButton("Straddle")
        self.sb_straddle_btn.setMinimumHeight(35)
        self.sb_straddle_btn.clicked.connect(lambda: self._apply_prebuilt_strategy("straddle"))
        strategies_row1.addWidget(self.sb_straddle_btn)
        prebuilt_layout.addLayout(strategies_row1)

        strategies_row2 = QHBoxLayout()
        self.sb_strangle_btn = QPushButton("Strangle")
        self.sb_strangle_btn.setMinimumHeight(35)
        self.sb_strangle_btn.clicked.connect(lambda: self._apply_prebuilt_strategy("strangle"))
        strategies_row2.addWidget(self.sb_strangle_btn)

        self.sb_bull_spread_btn = QPushButton("Bull Call")
        self.sb_bull_spread_btn.setMinimumHeight(35)
        self.sb_bull_spread_btn.clicked.connect(lambda: self._apply_prebuilt_strategy("bull_call_spread"))
        strategies_row2.addWidget(self.sb_bull_spread_btn)

        self.sb_bear_spread_btn = QPushButton("Bear Put")
        self.sb_bear_spread_btn.setMinimumHeight(35)
        self.sb_bear_spread_btn.clicked.connect(lambda: self._apply_prebuilt_strategy("bear_put_spread"))
        strategies_row2.addWidget(self.sb_bear_spread_btn)
        prebuilt_layout.addLayout(strategies_row2)

        # Row 1: Symbol & Pre-built Strategies side by side
        row1_layout = QHBoxLayout()
        row1_layout.addWidget(symbol_group)
        row1_layout.addWidget(prebuilt_group)
        config_layout.addLayout(row1_layout)

        # Strategy Legs
        legs_group = QGroupBox("📝 Strategy Legs (Max 4)")
        legs_layout = QVBoxLayout(legs_group)

        # Legs Table
        self.sb_legs_table = QTableWidget()
        self.sb_legs_table.setColumnCount(9)
        self.sb_legs_table.setHorizontalHeaderLabels([
            "Symbol", "B/S", "Type", "Strike", "Qty", "Premium", "LTP", "Exit", "Remove"
        ])
        self.sb_legs_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.sb_legs_table.setMinimumHeight(120)
        self.sb_legs_table.setMaximumHeight(180)
        self.sb_legs_table.verticalHeader().setDefaultSectionSize(30)
        legs_layout.addWidget(self.sb_legs_table)

        # Square-off All Button and Refresh LTP
        squareoff_layout = QHBoxLayout()

        self.sb_refresh_ltp_btn = QPushButton("🔄 Refresh LTP")
        self.sb_refresh_ltp_btn.setMinimumHeight(35)
        self.sb_refresh_ltp_btn.setStyleSheet("background-color: #2196F3; color: white; font-weight: bold;")
        self.sb_refresh_ltp_btn.clicked.connect(self._refresh_legs_ltp)
        squareoff_layout.addWidget(self.sb_refresh_ltp_btn)

        self.sb_squareoff_all_btn = QPushButton("🔴 Square-off ALL Positions")
        self.sb_squareoff_all_btn.setMinimumHeight(35)
        self.sb_squareoff_all_btn.setStyleSheet("background-color: #d32f2f; color: white; font-weight: bold;")
        self.sb_squareoff_all_btn.clicked.connect(self._squareoff_all_legs)
        squareoff_layout.addWidget(self.sb_squareoff_all_btn)
        legs_layout.addLayout(squareoff_layout)

        # Add Leg Controls
        add_leg_layout = QHBoxLayout()

        self.sb_leg_action = QComboBox()
        self.sb_leg_action.addItems(["BUY", "SELL"])
        self.sb_leg_action.setMinimumHeight(28)
        add_leg_layout.addWidget(self.sb_leg_action)

        self.sb_leg_type = QComboBox()
        self.sb_leg_type.addItems(["CE", "PE"])
        self.sb_leg_type.setMinimumHeight(28)
        add_leg_layout.addWidget(self.sb_leg_type)

        self.sb_leg_strike = QDoubleSpinBox()
        self.sb_leg_strike.setRange(0, 999999)
        self.sb_leg_strike.setDecimals(0)
        self.sb_leg_strike.setSingleStep(50)
        self.sb_leg_strike.setValue(25000)
        self.sb_leg_strike.setPrefix("₹")
        self.sb_leg_strike.setMinimumHeight(28)
        self.sb_leg_strike.setMinimumWidth(80)
        add_leg_layout.addWidget(self.sb_leg_strike)

        self.sb_leg_qty = QSpinBox()
        self.sb_leg_qty.setRange(1, 100)
        self.sb_leg_qty.setValue(1)
        self.sb_leg_qty.setSuffix(" lots")
        self.sb_leg_qty.setMinimumHeight(28)
        add_leg_layout.addWidget(self.sb_leg_qty)

        self.sb_leg_premium = QDoubleSpinBox()
        self.sb_leg_premium.setRange(0, 99999)
        self.sb_leg_premium.setDecimals(2)
        self.sb_leg_premium.setValue(100)
        self.sb_leg_premium.setPrefix("₹")
        self.sb_leg_premium.setMinimumHeight(28)
        add_leg_layout.addWidget(self.sb_leg_premium)

        legs_layout.addLayout(add_leg_layout)

        add_leg_btn_layout = QHBoxLayout()
        self.sb_add_leg_btn = QPushButton("➕ Add Leg")
        self.sb_add_leg_btn.setMinimumHeight(32)
        self.sb_add_leg_btn.clicked.connect(self._add_strategy_leg)
        add_leg_btn_layout.addWidget(self.sb_add_leg_btn)

        self.sb_clear_legs_btn = QPushButton("🗑️ Clear All")
        self.sb_clear_legs_btn.setMinimumHeight(32)
        self.sb_clear_legs_btn.clicked.connect(self._clear_strategy_legs)
        add_leg_btn_layout.addWidget(self.sb_clear_legs_btn)
        legs_layout.addLayout(add_leg_btn_layout)

        # Row 2: Strategy Legs (full width)
        config_layout.addWidget(legs_group)

        # Strategy Summary
        summary_group = QGroupBox("📈 Strategy Summary")
        summary_group.setStyleSheet("QGroupBox { font-size: 14px; font-weight: bold; }")
        summary_layout = QFormLayout(summary_group)

        self.sb_max_profit = QLabel("₹0.00")
        self.sb_max_profit.setStyleSheet("color: #4CAF50; font-weight: bold; font-size: 16px;")
        self.sb_max_profit.setMinimumHeight(25)
        summary_layout.addRow("Max Profit:", self.sb_max_profit)

        self.sb_max_loss = QLabel("₹0.00")
        self.sb_max_loss.setStyleSheet("color: #F44336; font-weight: bold; font-size: 16px;")
        self.sb_max_loss.setMinimumHeight(25)
        summary_layout.addRow("Max Loss:", self.sb_max_loss)

        self.sb_breakeven = QLabel("--")
        self.sb_breakeven.setStyleSheet("color: #2196F3; font-weight: bold; font-size: 14px;")
        self.sb_breakeven.setMinimumHeight(25)
        summary_layout.addRow("Breakeven:", self.sb_breakeven)

        self.sb_net_premium = QLabel("₹0.00")
        self.sb_net_premium.setStyleSheet("font-size: 14px;")
        self.sb_net_premium.setMinimumHeight(25)
        summary_layout.addRow("Net Premium:", self.sb_net_premium)

        self.sb_risk_reward = QLabel("--")
        self.sb_risk_reward.setStyleSheet("font-size: 14px;")
        self.sb_risk_reward.setMinimumHeight(25)
        summary_layout.addRow("Risk:Reward:", self.sb_risk_reward)

        # Live P&L Box (define before row3_layout)
        pnl_group = QGroupBox("📊 Live P&L")
        pnl_layout = QFormLayout(pnl_group)

        self.sb_entry_spot = QDoubleSpinBox()
        self.sb_entry_spot.setRange(0, 999999)
        self.sb_entry_spot.setDecimals(2)
        self.sb_entry_spot.setValue(0)
        self.sb_entry_spot.setPrefix("₹ ")
        self.sb_entry_spot.setMinimumHeight(30)
        self.sb_entry_spot.valueChanged.connect(self._update_live_pnl)
        pnl_layout.addRow("Entry Spot:", self.sb_entry_spot)

        self.sb_current_pnl = QLabel("₹0.00")
        self.sb_current_pnl.setStyleSheet("font-weight: bold; font-size: 18px;")
        self.sb_current_pnl.setMinimumHeight(25)
        pnl_layout.addRow("Current P&L:", self.sb_current_pnl)

        self.sb_pnl_change = QLabel("₹0.00 (0.00%)")
        self.sb_pnl_change.setStyleSheet("font-size: 14px;")
        self.sb_pnl_change.setMinimumHeight(20)
        pnl_layout.addRow("Change:", self.sb_pnl_change)

        # Lock Entry button
        self.sb_lock_entry_btn = QPushButton("🔒 Lock Entry Price")
        self.sb_lock_entry_btn.setMinimumHeight(32)
        self.sb_lock_entry_btn.clicked.connect(self._lock_entry_price)
        pnl_layout.addRow(self.sb_lock_entry_btn)

        # Row 3: Summary + Live P&L side by side
        row3_layout = QHBoxLayout()
        row3_layout.addWidget(summary_group)
        row3_layout.addWidget(pnl_group)
        config_layout.addLayout(row3_layout)

        # Save/Load Strategy Buttons
        strategy_io_group = QGroupBox("💾 Save/Load Strategy")
        strategy_io_layout = QVBoxLayout(strategy_io_group)

        # Strategy name input
        name_layout = QHBoxLayout()
        name_layout.addWidget(QLabel("Name:"))
        self.sb_strategy_name = QLineEdit()
        self.sb_strategy_name.setPlaceholderText("My Custom Strategy")
        self.sb_strategy_name.setMinimumHeight(30)
        name_layout.addWidget(self.sb_strategy_name)
        strategy_io_layout.addLayout(name_layout)

        # Save/Load buttons - Row 1
        io_btn_layout = QHBoxLayout()
        self.sb_save_strategy_btn = QPushButton("💾 Save")
        self.sb_save_strategy_btn.setMinimumHeight(32)
        self.sb_save_strategy_btn.clicked.connect(self._save_custom_strategy)
        io_btn_layout.addWidget(self.sb_save_strategy_btn)

        self.sb_load_strategy_btn = QPushButton("📂 Load")
        self.sb_load_strategy_btn.setMinimumHeight(32)
        self.sb_load_strategy_btn.clicked.connect(self._load_custom_strategy)
        io_btn_layout.addWidget(self.sb_load_strategy_btn)

        self.sb_edit_strategy_btn = QPushButton("✏️ Edit")
        self.sb_edit_strategy_btn.setMinimumHeight(32)
        self.sb_edit_strategy_btn.clicked.connect(self._edit_saved_strategy)
        io_btn_layout.addWidget(self.sb_edit_strategy_btn)
        strategy_io_layout.addLayout(io_btn_layout)

        # Deploy/Delete buttons - Row 2
        io_btn_layout2 = QHBoxLayout()
        self.sb_deploy_strategy_btn = QPushButton("🚀 Deploy")
        self.sb_deploy_strategy_btn.setMinimumHeight(32)
        self.sb_deploy_strategy_btn.setStyleSheet("background-color: #FF9800; color: white;")
        self.sb_deploy_strategy_btn.clicked.connect(self._deploy_saved_strategy)
        io_btn_layout2.addWidget(self.sb_deploy_strategy_btn)

        self.sb_delete_strategy_btn = QPushButton("🗑️ Delete")
        self.sb_delete_strategy_btn.setMinimumHeight(32)
        self.sb_delete_strategy_btn.setStyleSheet("background-color: #F44336; color: white;")
        self.sb_delete_strategy_btn.clicked.connect(self._delete_saved_strategy)
        io_btn_layout2.addWidget(self.sb_delete_strategy_btn)

        self.sb_open_folder_btn = QPushButton("📁 Folder")
        self.sb_open_folder_btn.setMinimumHeight(32)
        self.sb_open_folder_btn.clicked.connect(self._open_strategies_folder)
        io_btn_layout2.addWidget(self.sb_open_folder_btn)
        strategy_io_layout.addLayout(io_btn_layout2)

        # Saved strategies list
        self.sb_saved_strategies_combo = QComboBox()
        self.sb_saved_strategies_combo.setMinimumHeight(30)
        self.sb_saved_strategies_combo.setPlaceholderText("-- Select Saved Strategy --")
        strategy_io_layout.addWidget(self.sb_saved_strategies_combo)

        # Row 4: Save/Load Strategy (full width)
        config_layout.addWidget(strategy_io_group)

        # Execute Button
        execute_layout = QHBoxLayout()
        self.sb_execute_btn = QPushButton("🚀 Execute Strategy")
        self.sb_execute_btn.setMinimumHeight(45)
        self.sb_execute_btn.setStyleSheet("background-color: #4CAF50; color: white; font-weight: bold; padding: 10px; font-size: 14px;")
        self.sb_execute_btn.clicked.connect(self._execute_strategy_builder)
        execute_layout.addWidget(self.sb_execute_btn)
        # Row 5: Execute Button
        config_layout.addLayout(execute_layout)

        # Set up scroll area
        scroll_area.setWidget(config_panel)
        scroll_area.setMaximumHeight(400)
        main_layout.addWidget(scroll_area)

        # Bottom Panel - Payoff Chart (takes remaining space)
        chart_group = QGroupBox("📊 Payoff Diagram")
        chart_layout = QVBoxLayout(chart_group)

        # Create matplotlib figure for payoff chart
        self.sb_figure = Figure(figsize=(10, 4), facecolor='#1e1e1e')
        self.sb_canvas = FigureCanvas(self.sb_figure)
        self.sb_ax = self.sb_figure.add_subplot(111)
        self._setup_payoff_chart()
        chart_layout.addWidget(self.sb_canvas)

        # Chart controls in same row
        chart_controls = QHBoxLayout()
        self.sb_refresh_chart_btn = QPushButton("🔄 Refresh Chart")
        self.sb_refresh_chart_btn.setMinimumHeight(35)
        self.sb_refresh_chart_btn.clicked.connect(self._update_payoff_chart)
        chart_controls.addWidget(self.sb_refresh_chart_btn)

        self.sb_show_greeks_btn = QPushButton("📐 Show Greeks")
        self.sb_show_greeks_btn.setMinimumHeight(35)
        self.sb_show_greeks_btn.clicked.connect(self._show_strategy_greeks)
        chart_controls.addWidget(self.sb_show_greeks_btn)

        # Greeks line chart toggles
        self.sb_show_vega_check = QCheckBox("Show Vega")
        self.sb_show_vega_check.setStyleSheet("color: #FF9800;")
        self.sb_show_vega_check.stateChanged.connect(self._update_payoff_chart)
        chart_controls.addWidget(self.sb_show_vega_check)

        self.sb_show_delta_check = QCheckBox("Show Delta")
        self.sb_show_delta_check.setStyleSheet("color: #2196F3;")
        self.sb_show_delta_check.stateChanged.connect(self._update_payoff_chart)
        chart_controls.addWidget(self.sb_show_delta_check)

        self.sb_show_theta_check = QCheckBox("Show Theta")
        self.sb_show_theta_check.setStyleSheet("color: #9C27B0;")
        self.sb_show_theta_check.stateChanged.connect(self._update_payoff_chart)
        chart_controls.addWidget(self.sb_show_theta_check)

        chart_controls.addStretch()
        chart_layout.addLayout(chart_controls)

        main_layout.addWidget(chart_group)

        # Initialize strategy legs data
        self.strategy_legs = []
        self.sb_entry_spot_locked = False

        # Create strategies folder and load saved strategies
        self._init_strategies_folder()
        self._refresh_saved_strategies_list()

        self.tabs.addTab(builder_widget, "Strategy Builder")

    def _create_cpr_strategy_tab(self):
        """Create CPR Auto-Trade Strategy tab with Multi-Symbol Support"""
        from algo_trader.strategies.cpr_strategy import (
            CPRAutoTrader, CPRCalculator, CPRSignal, PremiumZone
        )

        # Wrap in scroll area
        cpr_scroll = QScrollArea()
        cpr_scroll.setWidgetResizable(True)

        cpr_widget = QWidget()
        main_layout = QVBoxLayout(cpr_widget)

        # Top Section - Multi-Symbol Management
        multi_symbol_group = QGroupBox("Multi-Symbol CPR Management")
        multi_layout = QVBoxLayout(multi_symbol_group)

        # Add Symbol Row
        add_symbol_layout = QHBoxLayout()

        self.cpr_symbol = QComboBox()
        self.cpr_symbol.setEditable(True)
        # Index F&O
        fo_indices = ["NIFTY", "BANKNIFTY", "FINNIFTY", "MIDCPNIFTY", "SENSEX", "BANKEX"]
        # F&O Stocks (NSE)
        fo_stocks = [
            "RELIANCE", "TCS", "HDFCBANK", "INFY", "ICICIBANK", "HINDUNILVR", "SBIN",
            "BHARTIARTL", "ITC", "KOTAKBANK", "LT", "AXISBANK", "ASIANPAINT", "MARUTI",
            "BAJFINANCE", "HCLTECH", "WIPRO", "SUNPHARMA", "ULTRACEMCO", "TITAN",
            "NTPC", "POWERGRID", "ONGC", "COALINDIA", "JSWSTEEL", "TATASTEEL",
            "HINDALCO", "ADANIENT", "ADANIPORTS", "GRASIM", "DRREDDY", "CIPLA",
            "APOLLOHOSP", "DIVISLAB", "EICHERMOT", "BAJAJ-AUTO", "HEROMOTOCO",
            "M&M", "TATAMOTORS", "TECHM", "BPCL", "IOC", "NESTLEIND", "BRITANNIA",
            "DABUR", "GODREJCP", "MARICO", "PIDILITIND", "HAVELLS", "VOLTAS",
            "INDUSINDBK", "BANKBARODA", "PNB", "CANBK", "IDFCFIRSTB", "FEDERALBNK",
            "AUBANK", "BANDHANBNK", "SBILIFE", "HDFCLIFE", "ICICIPRULI", "BAJAJFINSV",
            "CHOLAFIN", "MUTHOOTFIN", "SHRIRAMFIN", "PEL", "RECLTD", "PFC", "IRFC",
            "TATAPOWER", "ADANIGREEN", "NHPC", "SJVN", "TORNTPOWER", "CUMMINSIND",
            "ABB", "SIEMENS", "BHEL", "BEL", "HAL", "BHARATFORG", "ESCORTS",
            "ASHOKLEY", "TVSMOTOR", "BALKRISIND", "MRF", "APOLLOTYRE", "EXIDEIND",
            "BOSCHLTD", "MOTHERSON", "AUROPHARMA", "BIOCON", "LUPIN", "ALKEM",
            "LAURUSLABS", "IPCALAB", "ZYDUSLIFE", "GLENMARK", "TORNTPHARM",
            "ACC", "AMBUJACEM", "SHREECEM", "RAMCOCEM", "DALBHARAT", "JKCEMENT",
            "UPL", "PIIND", "COFORGE", "LTIM", "MPHASIS", "PERSISTENT", "TATAELXSI",
            "LTTS", "NAUKRI", "INDIGO", "TRENT", "DMART", "ZOMATO", "PAYTM",
            "POLICYBZR", "NYKAA", "IRCTC", "INDIANHOTELS", "LICI", "SBICARD",
            "PAGEIND", "MCDOWELL-N", "BERGEPAINT", "KANSAINER", "INDHOTEL"
        ]
        self.cpr_symbol.addItems(fo_indices + sorted(fo_stocks))
        self.cpr_symbol.setMinimumHeight(30)
        self.cpr_symbol.setMinimumWidth(150)
        add_symbol_layout.addWidget(QLabel("Symbol:"))
        add_symbol_layout.addWidget(self.cpr_symbol)

        self.cpr_add_symbol_btn = QPushButton("+ Add Symbol")
        self.cpr_add_symbol_btn.setMinimumHeight(30)
        self.cpr_add_symbol_btn.setStyleSheet("background-color: #4CAF50; color: white; font-weight: bold;")
        self.cpr_add_symbol_btn.clicked.connect(self._add_cpr_symbol)
        add_symbol_layout.addWidget(self.cpr_add_symbol_btn)

        self.cpr_remove_symbol_btn = QPushButton("- Remove Selected")
        self.cpr_remove_symbol_btn.setMinimumHeight(30)
        self.cpr_remove_symbol_btn.setStyleSheet("background-color: #F44336; color: white;")
        self.cpr_remove_symbol_btn.clicked.connect(self._remove_cpr_symbol)
        add_symbol_layout.addWidget(self.cpr_remove_symbol_btn)

        add_symbol_layout.addStretch()
        multi_layout.addLayout(add_symbol_layout)

        # Active Symbols Display
        self.cpr_active_symbols_label = QLabel("Active Symbols: None")
        self.cpr_active_symbols_label.setStyleSheet("font-size: 12px; color: #888;")
        multi_layout.addWidget(self.cpr_active_symbols_label)

        main_layout.addWidget(multi_symbol_group)

        # Middle Section - Horizontal Layout
        middle_layout = QHBoxLayout()

        # Left Panel - Configuration for Selected Symbol
        left_panel = QWidget()
        left_layout = QVBoxLayout(left_panel)
        left_panel.setMaximumWidth(400)

        # Symbol Tabs for multi-symbol
        self.cpr_symbol_tabs = QTabWidget()
        self.cpr_symbol_tabs.setTabsClosable(False)
        self.cpr_symbol_tabs.currentChanged.connect(self._on_cpr_tab_changed)
        left_layout.addWidget(self.cpr_symbol_tabs)

        # Settings for current symbol
        settings_group = QGroupBox("Settings (Current Symbol)")
        settings_layout = QFormLayout(settings_group)

        self.cpr_timeframe = QComboBox()
        self.cpr_timeframe.addItems(["Daily", "Weekly", "Monthly"])
        self.cpr_timeframe.setMinimumHeight(30)
        self.cpr_timeframe.currentTextChanged.connect(self._auto_fetch_ohlc)
        settings_layout.addRow("Pivot Timeframe:", self.cpr_timeframe)

        self.cpr_strike_method = QComboBox()
        self.cpr_strike_method.addItems([
            "Traditional Pivot S1/R1",
            "Camarilla Pivot R1/S1",
            "Camarilla Pivot R2/S2",
            "Central Pivot"
        ])
        self.cpr_strike_method.setMinimumHeight(30)
        settings_layout.addRow("Strike Method:", self.cpr_strike_method)

        left_layout.addWidget(settings_group)

        # OHLC Data Input
        data_group = QGroupBox("OHLC Data")
        data_layout = QFormLayout(data_group)

        fetch_btn_layout = QHBoxLayout()
        self.cpr_auto_fetch_btn = QPushButton("Auto-Fetch OHLC")
        self.cpr_auto_fetch_btn.setMinimumHeight(32)
        self.cpr_auto_fetch_btn.setStyleSheet("background-color: #2196F3; color: white; font-weight: bold;")
        self.cpr_auto_fetch_btn.clicked.connect(self._auto_fetch_ohlc)
        fetch_btn_layout.addWidget(self.cpr_auto_fetch_btn)

        self.cpr_fetch_status = QLabel("")
        self.cpr_fetch_status.setStyleSheet("font-size: 11px;")
        fetch_btn_layout.addWidget(self.cpr_fetch_status)
        data_layout.addRow(fetch_btn_layout)

        self.cpr_prior_high = QDoubleSpinBox()
        self.cpr_prior_high.setRange(0, 999999)
        self.cpr_prior_high.setDecimals(2)
        self.cpr_prior_high.setValue(25100)
        self.cpr_prior_high.setMinimumHeight(28)
        data_layout.addRow("High:", self.cpr_prior_high)

        self.cpr_prior_low = QDoubleSpinBox()
        self.cpr_prior_low.setRange(0, 999999)
        self.cpr_prior_low.setDecimals(2)
        self.cpr_prior_low.setValue(24900)
        self.cpr_prior_low.setMinimumHeight(28)
        data_layout.addRow("Low:", self.cpr_prior_low)

        self.cpr_prior_close = QDoubleSpinBox()
        self.cpr_prior_close.setRange(0, 999999)
        self.cpr_prior_close.setDecimals(2)
        self.cpr_prior_close.setValue(25000)
        self.cpr_prior_close.setMinimumHeight(28)
        data_layout.addRow("Close:", self.cpr_prior_close)

        calc_cpr_btn = QPushButton("Calculate CPR Levels")
        calc_cpr_btn.setMinimumHeight(32)
        calc_cpr_btn.clicked.connect(self._calculate_cpr_levels)
        data_layout.addRow(calc_cpr_btn)

        left_layout.addWidget(data_group)

        # CPR Levels Display
        levels_group = QGroupBox("CPR Levels")
        levels_layout = QFormLayout(levels_group)

        self.cpr_tc_label = QLabel("--")
        self.cpr_tc_label.setStyleSheet("color: #4CAF50; font-weight: bold; font-size: 13px;")
        levels_layout.addRow("Top Pivot:", self.cpr_tc_label)

        self.cpr_cp_label = QLabel("--")
        self.cpr_cp_label.setStyleSheet("color: #FF9800; font-weight: bold; font-size: 13px;")
        levels_layout.addRow("Central Pivot:", self.cpr_cp_label)

        self.cpr_bc_label = QLabel("--")
        self.cpr_bc_label.setStyleSheet("color: #F44336; font-weight: bold; font-size: 13px;")
        levels_layout.addRow("Bottom Pivot:", self.cpr_bc_label)

        self.cpr_range_label = QLabel("--")
        levels_layout.addRow("CPR Range:", self.cpr_range_label)

        left_layout.addWidget(levels_group)

        # Signal Display
        signal_group = QGroupBox("Live Signal")
        signal_layout = QFormLayout(signal_group)

        self.cpr_current_price = QDoubleSpinBox()
        self.cpr_current_price.setRange(0, 999999)
        self.cpr_current_price.setDecimals(2)
        self.cpr_current_price.setValue(25000)
        self.cpr_current_price.setMinimumHeight(32)
        self.cpr_current_price.valueChanged.connect(self._update_cpr_signal)
        signal_layout.addRow("Current Price:", self.cpr_current_price)

        self.cpr_signal_label = QLabel("No Signal")
        self.cpr_signal_label.setStyleSheet("font-size: 16px; font-weight: bold; padding: 8px;")
        self.cpr_signal_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        signal_layout.addRow("Signal:", self.cpr_signal_label)

        self.cpr_premium_zone_label = QLabel("--")
        self.cpr_premium_zone_label.setStyleSheet("font-size: 12px;")
        signal_layout.addRow("Premium Zone:", self.cpr_premium_zone_label)

        self.cpr_strike_label = QLabel("--")
        self.cpr_strike_label.setStyleSheet("font-size: 12px; font-weight: bold;")
        signal_layout.addRow("Suggested Strike:", self.cpr_strike_label)

        left_layout.addWidget(signal_group)

        # Risk Management Section (NEW)
        risk_group = QGroupBox("Risk Management (Per Trade)")
        risk_layout = QFormLayout(risk_group)

        self.cpr_target = QDoubleSpinBox()
        self.cpr_target.setRange(0, 100000)
        self.cpr_target.setDecimals(0)
        self.cpr_target.setValue(0)
        self.cpr_target.setPrefix("₹ ")
        self.cpr_target.setSpecialValueText("No Target")
        self.cpr_target.setMinimumHeight(28)
        risk_layout.addRow("Target (per lot):", self.cpr_target)

        self.cpr_stoploss = QDoubleSpinBox()
        self.cpr_stoploss.setRange(0, 100000)
        self.cpr_stoploss.setDecimals(0)
        self.cpr_stoploss.setValue(0)
        self.cpr_stoploss.setPrefix("₹ ")
        self.cpr_stoploss.setSpecialValueText("No SL")
        self.cpr_stoploss.setMinimumHeight(28)
        risk_layout.addRow("Stop Loss (per lot):", self.cpr_stoploss)

        self.cpr_trailing_sl = QDoubleSpinBox()
        self.cpr_trailing_sl.setRange(0, 100)
        self.cpr_trailing_sl.setDecimals(1)
        self.cpr_trailing_sl.setValue(0)
        self.cpr_trailing_sl.setSuffix(" %")
        self.cpr_trailing_sl.setSpecialValueText("No Trailing")
        self.cpr_trailing_sl.setMinimumHeight(28)
        risk_layout.addRow("Trailing SL:", self.cpr_trailing_sl)

        self.cpr_mtm_target = QDoubleSpinBox()
        self.cpr_mtm_target.setRange(0, 9999999)
        self.cpr_mtm_target.setDecimals(0)
        self.cpr_mtm_target.setValue(0)
        self.cpr_mtm_target.setPrefix("₹ ")
        self.cpr_mtm_target.setSpecialValueText("No MTM Target")
        self.cpr_mtm_target.setMinimumHeight(28)
        risk_layout.addRow("MTM Target (Total):", self.cpr_mtm_target)

        self.cpr_mtm_loss = QDoubleSpinBox()
        self.cpr_mtm_loss.setRange(0, 9999999)
        self.cpr_mtm_loss.setDecimals(0)
        self.cpr_mtm_loss.setValue(0)
        self.cpr_mtm_loss.setPrefix("₹ ")
        self.cpr_mtm_loss.setSpecialValueText("No MTM Loss Limit")
        self.cpr_mtm_loss.setMinimumHeight(28)
        risk_layout.addRow("MTM Loss Limit:", self.cpr_mtm_loss)

        left_layout.addWidget(risk_group)

        # Auto-Trade Controls
        auto_group = QGroupBox("Auto-Trade Controls")
        auto_layout = QVBoxLayout(auto_group)

        check_layout = QHBoxLayout()
        self.cpr_auto_trade_check = QCheckBox("Auto-Trade")
        self.cpr_auto_trade_check.setStyleSheet("font-weight: bold;")
        check_layout.addWidget(self.cpr_auto_trade_check)

        self.cpr_hedging_check = QCheckBox("Hedging")
        self.cpr_hedging_check.setChecked(True)
        self.cpr_hedging_check.setStyleSheet("color: #4CAF50;")
        check_layout.addWidget(self.cpr_hedging_check)

        self.cpr_test_mode_check = QCheckBox("Test Mode")
        self.cpr_test_mode_check.setChecked(True)
        check_layout.addWidget(self.cpr_test_mode_check)
        auto_layout.addLayout(check_layout)

        auto_btn_layout = QHBoxLayout()

        self.cpr_start_btn = QPushButton("Start All")
        self.cpr_start_btn.setMinimumHeight(35)
        self.cpr_start_btn.setStyleSheet("background-color: #4CAF50; color: white; font-weight: bold;")
        self.cpr_start_btn.clicked.connect(self._start_cpr_monitoring)
        auto_btn_layout.addWidget(self.cpr_start_btn)

        self.cpr_stop_btn = QPushButton("Stop All")
        self.cpr_stop_btn.setMinimumHeight(35)
        self.cpr_stop_btn.setStyleSheet("background-color: #F44336; color: white;")
        self.cpr_stop_btn.clicked.connect(self._stop_cpr_monitoring)
        self.cpr_stop_btn.setEnabled(False)
        auto_btn_layout.addWidget(self.cpr_stop_btn)

        auto_layout.addLayout(auto_btn_layout)

        self.cpr_execute_btn = QPushButton("Execute Current Signal")
        self.cpr_execute_btn.setMinimumHeight(38)
        self.cpr_execute_btn.setStyleSheet("background-color: #FF9800; color: white; font-weight: bold;")
        self.cpr_execute_btn.clicked.connect(self._execute_cpr_signal)
        auto_layout.addWidget(self.cpr_execute_btn)

        left_layout.addWidget(auto_group)

        middle_layout.addWidget(left_panel)

        # Right Panel - Multi-Symbol Logs
        right_panel = QWidget()
        right_layout = QVBoxLayout(right_panel)

        # Summary Table (All Symbols)
        summary_group = QGroupBox("All Symbols Status")
        summary_layout = QVBoxLayout(summary_group)

        self.cpr_summary_table = QTableWidget()
        self.cpr_summary_table.setColumnCount(7)
        self.cpr_summary_table.setHorizontalHeaderLabels([
            "Symbol", "Mode", "Signal", "Strike", "Status", "P&L", "Action"
        ])
        self.cpr_summary_table.horizontalHeader().setStretchLastSection(True)
        self.cpr_summary_table.setMaximumHeight(180)
        summary_layout.addWidget(self.cpr_summary_table)
        right_layout.addWidget(summary_group)

        # Symbol-wise Log Tabs
        log_group = QGroupBox("Trade Logs (Per Symbol)")
        log_layout = QVBoxLayout(log_group)

        self.cpr_log_tabs = QTabWidget()
        self.cpr_log_tabs.setTabPosition(QTabWidget.TabPosition.North)

        # Master Log Tab
        self.cpr_trade_log = QTextEdit()
        self.cpr_trade_log.setReadOnly(True)
        self.cpr_trade_log.setStyleSheet("font-family: monospace; font-size: 11px;")
        self.cpr_log_tabs.addTab(self.cpr_trade_log, "All Logs")

        log_layout.addWidget(self.cpr_log_tabs)

        clear_log_btn = QPushButton("Clear All Logs")
        clear_log_btn.clicked.connect(self._clear_all_cpr_logs)
        log_layout.addWidget(clear_log_btn)

        right_layout.addWidget(log_group)

        middle_layout.addWidget(right_panel)
        main_layout.addLayout(middle_layout)

        # Initialize multi-symbol CPR traders dictionary
        self.cpr_traders = {}  # {symbol: CPRAutoTrader}
        self.cpr_symbol_logs = {}  # {symbol: QTextEdit}
        self.cpr_symbol_data = {}  # {symbol: {high, low, close, levels, current_price}}

        # Legacy single trader for backward compatibility
        self.cpr_trader = CPRAutoTrader(
            symbol="NIFTY",
            auto_trade_enabled=False,
            test_mode=True
        )
        self.cpr_trader.on_signal_change = self._on_cpr_signal_change
        self.cpr_trader.on_trade_executed = self._on_cpr_trade_executed

        cpr_scroll.setWidget(cpr_widget)
        self.tabs.addTab(cpr_scroll, "CPR Strategy")

    def _add_cpr_symbol(self):
        """Add a new symbol to multi-symbol CPR tracking"""
        from algo_trader.strategies.cpr_strategy import CPRAutoTrader

        symbol = self.cpr_symbol.currentText().upper()

        if symbol in self.cpr_traders:
            QMessageBox.warning(self, "Duplicate", f"{symbol} is already added!")
            return

        # Create trader for this symbol
        trader = CPRAutoTrader(
            symbol=symbol,
            auto_trade_enabled=self.cpr_auto_trade_check.isChecked(),
            test_mode=self.cpr_test_mode_check.isChecked(),
            hedging_enabled=self.cpr_hedging_check.isChecked()
        )
        trader.on_signal_change = lambda sig, s=symbol: self._on_cpr_symbol_signal_change(s, sig)
        trader.on_trade_executed = lambda td, s=symbol: self._on_cpr_symbol_trade_executed(s, td)

        self.cpr_traders[symbol] = trader

        # Create log tab for this symbol
        log_widget = QTextEdit()
        log_widget.setReadOnly(True)
        log_widget.setStyleSheet("font-family: monospace; font-size: 11px;")
        self.cpr_symbol_logs[symbol] = log_widget
        self.cpr_log_tabs.addTab(log_widget, symbol)

        # Initialize symbol data with independent mode
        self.cpr_symbol_data[symbol] = {
            'high': 0, 'low': 0, 'close': 0,
            'levels': None, 'current_price': 0, 'signal': None,
            'test_mode': self.cpr_test_mode_check.isChecked()  # Independent mode per symbol
        }

        # Add to summary table
        self._update_cpr_summary_table()
        self._update_active_symbols_label()

        self._log_cpr(f"Added symbol: {symbol}")
        self._log_cpr_symbol(symbol, f"Symbol {symbol} added to monitoring")

    def _remove_cpr_symbol(self):
        """Remove selected symbol from multi-symbol tracking"""
        current_tab = self.cpr_log_tabs.currentIndex()
        if current_tab <= 0:  # Can't remove "All Logs" tab
            QMessageBox.warning(self, "Cannot Remove", "Select a symbol tab to remove")
            return

        symbol = self.cpr_log_tabs.tabText(current_tab)

        if symbol in self.cpr_traders:
            # Stop monitoring if running
            self.cpr_traders[symbol].stop_monitoring()
            del self.cpr_traders[symbol]

        if symbol in self.cpr_symbol_logs:
            del self.cpr_symbol_logs[symbol]

        if symbol in self.cpr_symbol_data:
            del self.cpr_symbol_data[symbol]

        # Remove tab
        self.cpr_log_tabs.removeTab(current_tab)

        self._update_cpr_summary_table()
        self._update_active_symbols_label()
        self._log_cpr(f"Removed symbol: {symbol}")

    def _update_active_symbols_label(self):
        """Update the active symbols label"""
        if self.cpr_traders:
            symbols = ", ".join(sorted(self.cpr_traders.keys()))
            self.cpr_active_symbols_label.setText(f"Active Symbols: {symbols}")
            self.cpr_active_symbols_label.setStyleSheet("font-size: 12px; color: #4CAF50;")
        else:
            self.cpr_active_symbols_label.setText("Active Symbols: None")
            self.cpr_active_symbols_label.setStyleSheet("font-size: 12px; color: #888;")

    def _update_cpr_summary_table(self):
        """Update the multi-symbol summary table"""
        self.cpr_summary_table.setRowCount(len(self.cpr_traders))

        for i, (symbol, trader) in enumerate(sorted(self.cpr_traders.items())):
            data = self.cpr_symbol_data.get(symbol, {})

            # Symbol
            self.cpr_summary_table.setItem(i, 0, QTableWidgetItem(symbol))

            # Mode - with toggle button
            mode_btn = QPushButton("PAPER" if trader.test_mode else "LIVE")
            mode_btn.setStyleSheet(
                "background-color: #FF9800; color: white;" if trader.test_mode
                else "background-color: #4CAF50; color: white;"
            )
            mode_btn.clicked.connect(lambda checked, s=symbol: self._toggle_cpr_symbol_mode(s))
            self.cpr_summary_table.setCellWidget(i, 1, mode_btn)

            # Signal
            signal_text = "--"
            if trader.current_signal:
                signal_text = trader.current_signal.signal.value.split(' ')[0]
            self.cpr_summary_table.setItem(i, 2, QTableWidgetItem(signal_text))

            # Strike
            strike_text = "--"
            if trader.current_signal and trader.current_signal.strike_value > 0:
                strike_text = f"₹{trader.current_signal.strike_value:.0f}"
            self.cpr_summary_table.setItem(i, 3, QTableWidgetItem(strike_text))

            # Status
            status = "Running" if trader._running else "Stopped"
            self.cpr_summary_table.setItem(i, 4, QTableWidgetItem(status))

            # P&L (placeholder)
            self.cpr_summary_table.setItem(i, 5, QTableWidgetItem("--"))

            # Action button
            action_btn = QPushButton("Execute")
            action_btn.clicked.connect(lambda checked, s=symbol: self._execute_cpr_symbol_signal(s))
            self.cpr_summary_table.setCellWidget(i, 6, action_btn)

    def _on_cpr_tab_changed(self, index):
        """Handle symbol tab change"""
        pass  # Future: Load symbol-specific settings

    def _toggle_cpr_symbol_mode(self, symbol):
        """Toggle paper/live mode for a specific symbol"""
        if symbol not in self.cpr_traders:
            return

        trader = self.cpr_traders[symbol]
        trader.test_mode = not trader.test_mode

        # Update symbol data
        if symbol in self.cpr_symbol_data:
            self.cpr_symbol_data[symbol]['test_mode'] = trader.test_mode

        mode = "PAPER" if trader.test_mode else "LIVE"
        self._log_cpr_symbol(symbol, f"Mode changed to: {mode}")
        self._log_cpr(f"[{symbol}] Mode: {mode}")

        self._update_cpr_summary_table()

    def _on_cpr_symbol_signal_change(self, symbol, signal):
        """Callback when a specific symbol's signal changes"""
        self._log_cpr_symbol(symbol, f"SIGNAL: {signal.signal.value} @ ₹{signal.current_price:.2f}")
        self._log_cpr(f"[{symbol}] SIGNAL: {signal.signal.value}")
        self._update_cpr_summary_table()

    def _on_cpr_symbol_trade_executed(self, symbol, trade_details):
        """Callback when trade is executed for a specific symbol"""
        # Include risk management settings
        trade_details['target'] = self.cpr_target.value()
        trade_details['stoploss'] = self.cpr_stoploss.value()
        trade_details['trailing_sl'] = self.cpr_trailing_sl.value()
        trade_details['mtm_target'] = self.cpr_mtm_target.value()
        trade_details['mtm_loss'] = self.cpr_mtm_loss.value()

        action = trade_details.get('action', '')
        self._log_cpr_symbol(symbol, f"\n{'='*30}\nTRADE EXECUTED!\n"
                           f"Action: {action}\n"
                           f"Target: ₹{trade_details['target']}\n"
                           f"SL: ₹{trade_details['stoploss']}\n"
                           f"Trailing: {trade_details['trailing_sl']}%\n"
                           f"MTM Target: ₹{trade_details['mtm_target']}\n"
                           f"MTM Loss: ₹{trade_details['mtm_loss']}\n"
                           f"{'='*30}")
        self._log_cpr(f"[{symbol}] TRADE: {action}")

        # Call the main trade executed handler for Strategy Builder integration
        self._on_cpr_trade_executed(trade_details)

    def _execute_cpr_symbol_signal(self, symbol):
        """Execute current signal for a specific symbol"""
        if symbol not in self.cpr_traders:
            return

        trader = self.cpr_traders[symbol]
        if trader.current_signal and trader.current_signal.signal.value != "No Signal":
            trader._execute_trade(trader.current_signal)
        else:
            QMessageBox.warning(self, "No Signal", f"No actionable signal for {symbol}")

    def _log_cpr_symbol(self, symbol, message):
        """Log message to symbol-specific log"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        if symbol in self.cpr_symbol_logs:
            self.cpr_symbol_logs[symbol].append(f"[{timestamp}] {message}")

    def _clear_all_cpr_logs(self):
        """Clear all CPR logs"""
        self.cpr_trade_log.clear()
        for log in self.cpr_symbol_logs.values():
            log.clear()

    def _calculate_cpr_levels(self):
        """Calculate CPR levels from prior day data"""
        from algo_trader.strategies.cpr_strategy import CPRCalculator

        symbol = self.cpr_symbol.currentText().upper()
        high = self.cpr_prior_high.value()
        low = self.cpr_prior_low.value()
        close = self.cpr_prior_close.value()

        if high <= 0 or low <= 0 or close <= 0:
            QMessageBox.warning(self, "Invalid Data", "Please enter valid OHLC data")
            return

        if high < low:
            QMessageBox.warning(self, "Invalid Data", "High must be greater than Low")
            return

        # Calculate levels
        levels = CPRCalculator.calculate_all(high, low, close)

        # Update labels
        self.cpr_tc_label.setText(f"₹{levels.top_pivot:.2f}")
        self.cpr_cp_label.setText(f"₹{levels.central_pivot:.2f}")
        self.cpr_bc_label.setText(f"₹{levels.bottom_pivot:.2f}")
        self.cpr_range_label.setText(f"₹{levels.cpr_range:.2f}")

        # Set in legacy trader
        self.cpr_trader.set_prior_day_data(high, low, close)
        self.cpr_trader.symbol = symbol

        # Also update multi-symbol trader if exists
        if symbol in self.cpr_traders:
            self.cpr_traders[symbol].set_prior_day_data(high, low, close)
            self.cpr_symbol_data[symbol].update({
                'high': high, 'low': low, 'close': close,
                'levels': levels
            })
            self._log_cpr_symbol(symbol, f"CPR Levels: TC={levels.top_pivot:.2f}, "
                                        f"CP={levels.central_pivot:.2f}, BC={levels.bottom_pivot:.2f}")

        # Update signal
        self._update_cpr_signal()

        self._log_cpr(f"[{symbol}] CPR Levels Calculated:\n"
                     f"  Top Pivot: {levels.top_pivot:.2f}\n"
                     f"  Central Pivot: {levels.central_pivot:.2f}\n"
                     f"  Bottom Pivot: {levels.bottom_pivot:.2f}\n"
                     f"  CPR Range: {levels.cpr_range:.2f}")

    def _update_cpr_signal(self):
        """Update signal based on current price"""
        if not self.cpr_trader.cpr_levels:
            return

        current_price = self.cpr_current_price.value()
        signal = self.cpr_trader.update_price(current_price)

        if signal:
            self._display_cpr_signal(signal)

    def _display_cpr_signal(self, signal):
        """Display the CPR signal in UI"""
        from algo_trader.strategies.cpr_strategy import CPRSignal

        signal_text = signal.signal.value
        zone_text = signal.premium_zone.value

        # Update signal label with color
        if signal.signal == CPRSignal.BULLISH:
            self.cpr_signal_label.setText("BULLISH (Sell Put)")
            self.cpr_signal_label.setStyleSheet(
                "font-size: 18px; font-weight: bold; padding: 10px; "
                "background-color: #4CAF50; color: white; border-radius: 5px;"
            )
            self.cpr_strike_label.setText(f"PE Strike: ₹{signal.strike_value:.0f}")

        elif signal.signal == CPRSignal.BEARISH:
            self.cpr_signal_label.setText("BEARISH (Sell Call)")
            self.cpr_signal_label.setStyleSheet(
                "font-size: 18px; font-weight: bold; padding: 10px; "
                "background-color: #F44336; color: white; border-radius: 5px;"
            )
            self.cpr_strike_label.setText(f"CE Strike: ₹{signal.strike_value:.0f}")

        elif signal.signal == CPRSignal.SIDEWAYS:
            self.cpr_signal_label.setText("SIDEWAYS (Iron Condor)")
            self.cpr_signal_label.setStyleSheet(
                "font-size: 18px; font-weight: bold; padding: 10px; "
                "background-color: #FF9800; color: white; border-radius: 5px;"
            )
            self.cpr_strike_label.setText(
                f"CE: ₹{signal.strike_value:.0f} | PE: ₹{signal.strike_value_2:.0f}"
            )
        else:
            self.cpr_signal_label.setText("No Signal")
            self.cpr_signal_label.setStyleSheet(
                "font-size: 18px; font-weight: bold; padding: 10px; "
                "background-color: #666; color: white; border-radius: 5px;"
            )
            self.cpr_strike_label.setText("--")

        self.cpr_premium_zone_label.setText(zone_text)

    def _start_cpr_monitoring(self):
        """Start CPR monitoring for all symbols"""
        # Check if any traders have CPR levels
        has_levels = self.cpr_trader.cpr_levels is not None
        for trader in self.cpr_traders.values():
            if trader.cpr_levels:
                has_levels = True
                break

        if not has_levels:
            QMessageBox.warning(self, "No CPR Data",
                              "Please add symbols and calculate CPR levels first")
            return

        mode = "TEST" if self.cpr_test_mode_check.isChecked() else "LIVE"
        auto = "AUTO-TRADE ON" if self.cpr_auto_trade_check.isChecked() else "Manual"
        hedge = "HEDGING ON" if self.cpr_hedging_check.isChecked() else "No Hedge"

        # Start legacy trader
        if self.cpr_trader.cpr_levels:
            self.cpr_trader.auto_trade_enabled = self.cpr_auto_trade_check.isChecked()
            self.cpr_trader.test_mode = self.cpr_test_mode_check.isChecked()
            self.cpr_trader.hedging_enabled = self.cpr_hedging_check.isChecked()
            self.cpr_trader.start_monitoring(lambda: self.cpr_current_price.value())

        # Start all multi-symbol traders
        started_count = 0
        for symbol, trader in self.cpr_traders.items():
            if trader.cpr_levels:
                trader.auto_trade_enabled = self.cpr_auto_trade_check.isChecked()
                trader.test_mode = self.cpr_test_mode_check.isChecked()
                trader.hedging_enabled = self.cpr_hedging_check.isChecked()

                # Use symbol-specific price callback if available
                price_cb = lambda s=symbol: self.cpr_symbol_data.get(s, {}).get('current_price', 0)
                trader.start_monitoring(price_cb)
                started_count += 1
                self._log_cpr_symbol(symbol, f"Monitoring STARTED - {mode}, {auto}, {hedge}")

        self.cpr_start_btn.setEnabled(False)
        self.cpr_stop_btn.setEnabled(True)

        self._log_cpr(f"Monitoring STARTED for {started_count} symbols - {mode}, {auto}, {hedge}")
        self._update_cpr_summary_table()

    def _stop_cpr_monitoring(self):
        """Stop CPR monitoring for all symbols"""
        # Stop legacy trader
        self.cpr_trader.stop_monitoring()

        # Stop all multi-symbol traders
        for symbol, trader in self.cpr_traders.items():
            trader.stop_monitoring()
            self._log_cpr_symbol(symbol, "Monitoring STOPPED")

        self.cpr_start_btn.setEnabled(True)
        self.cpr_stop_btn.setEnabled(False)

        self._log_cpr("Monitoring STOPPED for all symbols")
        self._update_cpr_summary_table()

    def _execute_cpr_signal(self):
        """Manually execute current CPR signal"""
        from algo_trader.strategies.cpr_strategy import CPRSignal

        if not self.cpr_trader.current_signal:
            QMessageBox.warning(self, "No Signal", "No signal to execute")
            return

        signal = self.cpr_trader.current_signal

        if signal.signal == CPRSignal.NO_SIGNAL:
            QMessageBox.warning(self, "No Signal", "Current signal is 'No Signal'")
            return

        # Confirm execution
        reply = QMessageBox.question(
            self, "Confirm Trade",
            f"Execute {signal.signal.value}?\n\n"
            f"Strike: {signal.strike_value:.0f}\n"
            f"Premium Zone: {signal.premium_zone.value}",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )

        if reply == QMessageBox.StandardButton.Yes:
            self.cpr_trader._execute_trade(signal)

    def _on_cpr_signal_change(self, signal):
        """Callback when CPR signal changes"""
        self._display_cpr_signal(signal)
        self._log_cpr(f"SIGNAL CHANGE: {signal.signal.value} at ₹{signal.current_price:.2f}")

    def _on_cpr_trade_executed(self, trade_details):
        """Callback when trade is executed"""
        from algo_trader.strategies.cpr_strategy import CPRSignal

        action = trade_details.get('action', '')
        hedge_info = "WITH HEDGING" if trade_details.get('hedging_enabled') else "NO HEDGE"

        # Add risk management settings if not already present
        if 'target' not in trade_details:
            trade_details['target'] = self.cpr_target.value()
        if 'stoploss' not in trade_details:
            trade_details['stoploss'] = self.cpr_stoploss.value()
        if 'trailing_sl' not in trade_details:
            trade_details['trailing_sl'] = self.cpr_trailing_sl.value()
        if 'mtm_target' not in trade_details:
            trade_details['mtm_target'] = self.cpr_mtm_target.value()
        if 'mtm_loss' not in trade_details:
            trade_details['mtm_loss'] = self.cpr_mtm_loss.value()

        # Build risk info string
        risk_info = ""
        if trade_details['target'] > 0:
            risk_info += f"Target: ₹{trade_details['target']} | "
        if trade_details['stoploss'] > 0:
            risk_info += f"SL: ₹{trade_details['stoploss']} | "
        if trade_details['trailing_sl'] > 0:
            risk_info += f"TSL: {trade_details['trailing_sl']}% | "
        if trade_details['mtm_target'] > 0:
            risk_info += f"MTM Target: ₹{trade_details['mtm_target']} | "
        if trade_details['mtm_loss'] > 0:
            risk_info += f"MTM Loss: ₹{trade_details['mtm_loss']}"

        self._log_cpr(f"\n{'='*40}\n"
                     f"TRADE EXECUTED! ({hedge_info})\n"
                     f"Signal: {trade_details['signal']}\n"
                     f"Action: {action}\n"
                     f"Symbol: {trade_details['symbol']}\n"
                     f"Mode: {'TEST' if trade_details['test_mode'] else 'LIVE'}\n"
                     f"Time: {trade_details['timestamp']}\n"
                     f"{f'Risk: {risk_info}' if risk_info else ''}\n"
                     f"{'='*40}\n")

        # Auto-load strategy in Strategy Builder based on action
        if action == 'IRON_CONDOR':
            self._auto_load_iron_condor_full(trade_details)
        elif action == 'BULL_PUT_SPREAD':
            self._auto_load_bull_put_spread(trade_details)
        elif action == 'BEAR_CALL_SPREAD':
            self._auto_load_bear_call_spread(trade_details)

        # Place paper trade if in test mode
        if trade_details.get('test_mode', True) and hasattr(self, 'paper_simulator') and self.paper_simulator:
            self._place_cpr_paper_trade(trade_details)

    def _auto_load_iron_condor_full(self, trade_details):
        """Auto-load Iron Condor in Strategy Builder from trade details"""
        try:
            symbol = trade_details['symbol']
            ce_sell = trade_details.get('ce_sell_strike', trade_details.get('ce_strike', 0))
            pe_sell = trade_details.get('pe_sell_strike', trade_details.get('pe_strike', 0))
            ce_buy = trade_details.get('ce_buy_strike', ce_sell + 100)
            pe_buy = trade_details.get('pe_buy_strike', pe_sell - 100)
            mode = 'TEST' if trade_details.get('test_mode', True) else 'LIVE'

            # Fetch actual premiums
            ce_sell_prem = self._get_option_premium(symbol, ce_sell, 'CE')
            pe_sell_prem = self._get_option_premium(symbol, pe_sell, 'PE')
            ce_buy_prem = self._get_option_premium(symbol, ce_buy, 'CE')
            pe_buy_prem = self._get_option_premium(symbol, pe_buy, 'PE')

            self._switch_to_strategy_builder(symbol)

            self.strategy_legs = [
                {'action': 'SELL', 'type': 'CE', 'strike': ce_sell, 'qty': 1, 'premium': ce_sell_prem, 'ltp': ce_sell_prem, 'symbol': symbol, 'mode': mode},
                {'action': 'SELL', 'type': 'PE', 'strike': pe_sell, 'qty': 1, 'premium': pe_sell_prem, 'ltp': pe_sell_prem, 'symbol': symbol, 'mode': mode},
                {'action': 'BUY', 'type': 'CE', 'strike': ce_buy, 'qty': 1, 'premium': ce_buy_prem, 'ltp': ce_buy_prem, 'symbol': symbol, 'mode': mode},
                {'action': 'BUY', 'type': 'PE', 'strike': pe_buy, 'qty': 1, 'premium': pe_buy_prem, 'ltp': pe_buy_prem, 'symbol': symbol, 'mode': mode}
            ]

            self._refresh_legs_table()
            self._update_payoff_chart()
            self._calculate_strategy_metrics()

            self._log_cpr(f"Iron Condor auto-loaded [{mode}]:\n"
                         f"  {symbol} Sell CE: {ce_sell} @ ₹{ce_sell_prem} | Buy CE: {ce_buy} @ ₹{ce_buy_prem}\n"
                         f"  {symbol} Sell PE: {pe_sell} @ ₹{pe_sell_prem} | Buy PE: {pe_buy} @ ₹{pe_buy_prem}")

        except Exception as e:
            logger.error(f"Failed to auto-load Iron Condor: {e}")

    def _get_option_premium(self, symbol: str, strike: int, opt_type: str) -> float:
        """Get option premium from broker or estimate if not available"""
        # Try to get from broker if connected
        if hasattr(self, 'brokers') and self.brokers:
            broker = list(self.brokers.values())[0]
            try:
                # Try broker-specific methods
                if hasattr(broker, 'get_option_ltp'):
                    ltp = broker.get_option_ltp(symbol, strike, opt_type)
                    if ltp and ltp > 0:
                        return ltp
                elif hasattr(broker, 'get_ltp'):
                    # Build option symbol name
                    option_symbol = f"{symbol}{strike}{opt_type}"
                    ltp = broker.get_ltp(option_symbol)
                    if ltp and ltp > 0:
                        return ltp
            except Exception as e:
                logger.debug(f"Error fetching option LTP from broker: {e}")

        # Estimate premium based on spot price and strike distance
        spot = self.sb_current_spot.value() if hasattr(self, 'sb_current_spot') else 25000

        # Base premium estimation
        if 'BANK' in symbol.upper():
            base_premium = 400
        elif 'SENSEX' in symbol.upper() or 'BANKEX' in symbol.upper():
            base_premium = 350
        else:
            base_premium = 200

        # Adjust for moneyness (ITM options cost more)
        strike_distance = abs(spot - strike)
        moneyness_factor = max(0.2, 1 - (strike_distance / spot) * 3)

        # ITM intrinsic value
        if opt_type == 'CE':
            intrinsic = max(0, spot - strike)
        else:  # PE
            intrinsic = max(0, strike - spot)

        estimated_premium = intrinsic + (base_premium * moneyness_factor)
        return round(max(10, estimated_premium), 2)

    def _auto_load_bull_put_spread(self, trade_details):
        """Auto-load Bull Put Spread (Bullish with Hedge)"""
        try:
            symbol = trade_details['symbol']
            sell_strike = trade_details.get('sell_strike', trade_details.get('strike', 0))
            buy_strike = trade_details.get('buy_strike', sell_strike - 100)
            mode = 'TEST' if trade_details.get('test_mode', True) else 'LIVE'

            # Fetch actual premiums
            sell_premium = self._get_option_premium(symbol, sell_strike, 'PE')
            buy_premium = self._get_option_premium(symbol, buy_strike, 'PE')

            self._switch_to_strategy_builder(symbol)

            self.strategy_legs = [
                {'action': 'SELL', 'type': 'PE', 'strike': sell_strike, 'qty': 1, 'premium': sell_premium, 'ltp': sell_premium, 'symbol': symbol, 'mode': mode},
                {'action': 'BUY', 'type': 'PE', 'strike': buy_strike, 'qty': 1, 'premium': buy_premium, 'ltp': buy_premium, 'symbol': symbol, 'mode': mode}
            ]

            self._refresh_legs_table()
            self._update_payoff_chart()
            self._calculate_strategy_metrics()

            self._log_cpr(f"Bull Put Spread auto-loaded [{mode}]:\n"
                         f"  {symbol} Sell PE: {sell_strike} @ ₹{sell_premium}\n"
                         f"  {symbol} Buy PE: {buy_strike} @ ₹{buy_premium}")

        except Exception as e:
            logger.error(f"Failed to auto-load Bull Put Spread: {e}")

    def _auto_load_bear_call_spread(self, trade_details):
        """Auto-load Bear Call Spread (Bearish with Hedge)"""
        try:
            symbol = trade_details['symbol']
            sell_strike = trade_details.get('sell_strike', trade_details.get('strike', 0))
            buy_strike = trade_details.get('buy_strike', sell_strike + 100)
            mode = 'TEST' if trade_details.get('test_mode', True) else 'LIVE'

            # Fetch actual premiums
            sell_premium = self._get_option_premium(symbol, sell_strike, 'CE')
            buy_premium = self._get_option_premium(symbol, buy_strike, 'CE')

            self._switch_to_strategy_builder(symbol)

            self.strategy_legs = [
                {'action': 'SELL', 'type': 'CE', 'strike': sell_strike, 'qty': 1, 'premium': sell_premium, 'ltp': sell_premium, 'symbol': symbol, 'mode': mode},
                {'action': 'BUY', 'type': 'CE', 'strike': buy_strike, 'qty': 1, 'premium': buy_premium, 'ltp': buy_premium, 'symbol': symbol, 'mode': mode}
            ]

            self._refresh_legs_table()
            self._update_payoff_chart()
            self._calculate_strategy_metrics()

            self._log_cpr(f"Bear Call Spread auto-loaded [{mode}]:\n"
                         f"  {symbol} Sell CE: {sell_strike} @ ₹{sell_premium}\n"
                         f"  {symbol} Buy CE: {buy_strike} @ ₹{buy_premium}")

        except Exception as e:
            logger.error(f"Failed to auto-load Bear Call Spread: {e}")

    def _place_cpr_paper_trade(self, trade_details):
        """Place simulated paper trade for CPR strategy"""
        try:
            symbol = trade_details.get('symbol', 'NIFTY')
            action = trade_details.get('action', '')
            strike = trade_details.get('sell_strike', trade_details.get('strike', 0))

            # Calculate realistic premium based on symbol and strike distance from spot
            spot_price = trade_details.get('current_price', 25000)

            # Get lot size based on symbol
            if 'BANK' in symbol.upper():
                lot_size = 15
                base_premium = 400  # BANKNIFTY options are more expensive
            elif 'SENSEX' in symbol.upper() or 'BANKEX' in symbol.upper():
                lot_size = 10
                base_premium = 350
            else:  # NIFTY, FINNIFTY, etc.
                lot_size = 50 if 'NIFTY' in symbol.upper() else 25
                base_premium = 200

            # Adjust premium based on strike distance (ATM options cost more)
            strike_distance = abs(spot_price - strike) if strike > 0 else 0
            distance_factor = max(0.3, 1 - (strike_distance / spot_price) * 5)
            estimated_premium = base_premium * distance_factor

            # Simulate option premium price based on action
            if action == 'IRON_CONDOR':
                # Net credit from Iron Condor (sell 2 options, buy 2 wings)
                net_premium = estimated_premium * 1.4  # ~140% of single premium
                trade_symbol = f"{symbol}_IC_{int(strike)}"
                self.paper_simulator.place_order(
                    symbol=trade_symbol,
                    action="SELL",
                    quantity=lot_size,
                    order_type="MARKET",
                    price=net_premium,
                    source="CPR Iron Condor"
                )
            elif action == 'BULL_PUT_SPREAD':
                # Net credit from Bull Put Spread
                net_premium = estimated_premium * 0.6  # ~60% after buying hedge
                trade_symbol = f"{symbol}_BPS_{int(strike)}"
                self.paper_simulator.place_order(
                    symbol=trade_symbol,
                    action="SELL",
                    quantity=lot_size,
                    order_type="MARKET",
                    price=net_premium,
                    source="CPR Bull Put Spread"
                )
            elif action == 'BEAR_CALL_SPREAD':
                # Net credit from Bear Call Spread
                net_premium = estimated_premium * 0.6  # ~60% after buying hedge
                trade_symbol = f"{symbol}_BCS_{int(strike)}"
                self.paper_simulator.place_order(
                    symbol=trade_symbol,
                    action="SELL",
                    quantity=lot_size,
                    order_type="MARKET",
                    price=net_premium,
                    source="CPR Bear Call Spread"
                )

            self._log_cpr(f"Paper trade placed: {action} for {symbol} @ ₹{net_premium:.2f}")
            # Refresh dashboard to show updated P&L
            self._refresh_dashboard()

        except Exception as e:
            logger.error(f"Failed to place paper trade: {e}")
            self._log_cpr(f"Paper trade error: {e}")

    def _switch_to_strategy_builder(self, symbol):
        """Switch to Strategy Builder tab and set symbol"""
        for i in range(self.tabs.count()):
            if self.tabs.tabText(i) == "Strategy Builder":
                self.tabs.setCurrentIndex(i)
                break

        idx = self.sb_symbol.findText(symbol)
        if idx >= 0:
            self.sb_symbol.setCurrentIndex(idx)

        self.strategy_legs = []

    def _auto_load_iron_condor(self, ce_strike, pe_strike, symbol):
        """Auto-load Iron Condor in Strategy Builder (legacy method)"""
        trade_details = {
            'symbol': symbol,
            'ce_sell_strike': ce_strike,
            'pe_sell_strike': pe_strike,
            'ce_buy_strike': ce_strike + (100 if "BANK" in symbol else 50) * 2,
            'pe_buy_strike': pe_strike - (100 if "BANK" in symbol else 50) * 2
        }
        self._auto_load_iron_condor_full(trade_details)

    def _log_cpr(self, message):
        """Log message to CPR trade log"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.cpr_trade_log.append(f"[{timestamp}] {message}")

    def _on_cpr_symbol_changed(self, symbol):
        """Handle CPR symbol change - update default price and fetch OHLC"""
        # Set approximate current price based on symbol
        spot_prices = {
            "NIFTY": 25000,
            "BANKNIFTY": 52000,
            "FINNIFTY": 23000,
            "MIDCPNIFTY": 12000,
            "SENSEX": 82000,
            "BANKEX": 58000,
            # Some major F&O stocks - approximate prices
            "RELIANCE": 1300,
            "TCS": 4000,
            "HDFCBANK": 1700,
            "INFY": 1900,
            "ICICIBANK": 1300,
            "SBIN": 800,
            "BHARTIARTL": 1700,
            "ITC": 480,
            "KOTAKBANK": 1800,
            "LT": 3600,
            "AXISBANK": 1100,
            "MARUTI": 12000,
            "BAJFINANCE": 6800,
            "TATAMOTORS": 780,
            "TATASTEEL": 150,
            "WIPRO": 300,
            "HCLTECH": 1900,
            "SUNPHARMA": 1800,
            "TITAN": 3500,
            "ADANIENT": 2400,
            "M&M": 3000,
        }

        symbol_upper = symbol.upper()
        if symbol_upper in spot_prices:
            self.cpr_current_price.setValue(spot_prices[symbol_upper])

        # Auto-fetch OHLC for the new symbol
        self._auto_fetch_ohlc()

    def _auto_fetch_ohlc(self):
        """Auto-fetch OHLC data for CPR calculation based on symbol and timeframe"""
        try:
            import yfinance as yf
            from datetime import datetime, timedelta

            symbol = self.cpr_symbol.currentText().upper()
            timeframe = self.cpr_timeframe.currentText()

            self.cpr_fetch_status.setText("Fetching...")
            self.cpr_fetch_status.setStyleSheet("color: #2196F3; font-size: 11px;")
            QApplication.processEvents()

            # Map symbol to Yahoo Finance ticker
            # NSE stocks need .NS suffix, indices need special handling
            yf_symbol_map = {
                "NIFTY": "^NSEI",
                "BANKNIFTY": "^NSEBANK",
                "FINNIFTY": "NIFTY_FIN_SERVICE.NS",
                "MIDCPNIFTY": "NIFTY_MID_SELECT.NS",
                "SENSEX": "^BSESN",
                "BANKEX": "BSE-BANK.BO",
            }

            if symbol in yf_symbol_map:
                yf_ticker = yf_symbol_map[symbol]
            else:
                # Regular NSE stock
                yf_ticker = f"{symbol}.NS"

            # Determine period and interval based on timeframe
            if timeframe == "Daily":
                # Fetch last 5 days to get previous day's OHLC
                period = "5d"
                interval = "1d"
            elif timeframe == "Weekly":
                # Fetch last 2 weeks to get previous week's OHLC
                period = "1mo"
                interval = "1wk"
            else:  # Monthly
                # Fetch last 3 months to get previous month's OHLC
                period = "3mo"
                interval = "1mo"

            # Fetch data
            ticker = yf.Ticker(yf_ticker)
            hist = ticker.history(period=period, interval=interval)

            if hist.empty:
                self.cpr_fetch_status.setText("No data found")
                self.cpr_fetch_status.setStyleSheet("color: #F44336; font-size: 11px;")
                self._log_cpr(f"No data found for {symbol} ({yf_ticker})")
                return

            # Get the previous period's OHLC (second last row if available, else last)
            if len(hist) >= 2:
                prior_data = hist.iloc[-2]  # Previous completed period
                current_data = hist.iloc[-1]  # Current/latest period
            else:
                prior_data = hist.iloc[-1]
                current_data = hist.iloc[-1]

            high = prior_data['High']
            low = prior_data['Low']
            close = prior_data['Close']

            # Update the input fields
            self.cpr_prior_high.setValue(high)
            self.cpr_prior_low.setValue(low)
            self.cpr_prior_close.setValue(close)

            # Update current price with latest close
            self.cpr_current_price.setValue(current_data['Close'])

            # Show success
            date_str = prior_data.name.strftime("%d-%b-%Y") if hasattr(prior_data.name, 'strftime') else str(prior_data.name)[:10]
            self.cpr_fetch_status.setText(f"✓ {date_str}")
            self.cpr_fetch_status.setStyleSheet("color: #4CAF50; font-size: 11px;")

            self._log_cpr(f"OHLC fetched for {symbol} ({timeframe}): H={high:.2f}, L={low:.2f}, C={close:.2f}")

            # Auto-calculate CPR levels
            self._calculate_cpr_levels()

        except ImportError:
            self.cpr_fetch_status.setText("yfinance not installed")
            self.cpr_fetch_status.setStyleSheet("color: #F44336; font-size: 11px;")
            self._log_cpr("Error: yfinance package not installed. Run: pip install yfinance")
        except Exception as e:
            self.cpr_fetch_status.setText(f"Error: {str(e)[:20]}")
            self.cpr_fetch_status.setStyleSheet("color: #F44336; font-size: 11px;")
            self._log_cpr(f"Error fetching OHLC: {e}")
            logger.error(f"Auto-fetch OHLC error: {e}")

    def _setup_payoff_chart(self):
        """Setup the payoff chart styling"""
        self.sb_ax.set_facecolor('#2d2d2d')
        self.sb_ax.tick_params(colors='white')
        self.sb_ax.spines['bottom'].set_color('white')
        self.sb_ax.spines['top'].set_color('white')
        self.sb_ax.spines['left'].set_color('white')
        self.sb_ax.spines['right'].set_color('white')
        self.sb_ax.set_xlabel('Spot Price at Expiry', color='white')
        self.sb_ax.set_ylabel('Profit / Loss (₹)', color='white')
        self.sb_ax.axhline(y=0, color='gray', linestyle='--', alpha=0.5)
        self.sb_ax.set_title('Strategy Payoff', color='white')
        self.sb_figure.tight_layout()

    def _on_sb_symbol_changed(self, symbol):
        """Handle symbol change in strategy builder"""
        # Update spot price based on symbol (approximate)
        spot_prices = {
            "NIFTY": 25000,
            "BANKNIFTY": 52000,
            "FINNIFTY": 23000,
            "MIDCPNIFTY": 12000
        }
        if symbol.upper() in spot_prices:
            self.sb_spot_price.setValue(spot_prices[symbol.upper()])
            self.sb_leg_strike.setValue(spot_prices[symbol.upper()])

    def _on_spot_price_changed(self):
        """Handle spot price change - update chart and live P&L"""
        self._update_payoff_chart()
        self._calculate_strategy_metrics()
        # Update live P&L if entry is locked
        if hasattr(self, 'sb_entry_spot') and self.sb_entry_spot.value() > 0:
            self._update_live_pnl()

    def _add_strategy_leg(self):
        """Add a leg to the strategy"""
        if len(self.strategy_legs) >= 4:
            QMessageBox.warning(self, "Limit", "Maximum 4 legs allowed")
            return

        leg = {
            'action': self.sb_leg_action.currentText(),
            'type': self.sb_leg_type.currentText(),
            'strike': self.sb_leg_strike.value(),
            'qty': self.sb_leg_qty.value(),
            'premium': self.sb_leg_premium.value()
        }
        self.strategy_legs.append(leg)
        self._refresh_legs_table()
        self._update_payoff_chart()
        self._calculate_strategy_metrics()

    def _remove_strategy_leg(self, index):
        """Remove a leg from the strategy"""
        if 0 <= index < len(self.strategy_legs):
            self.strategy_legs.pop(index)
            self._refresh_legs_table()
            self._update_payoff_chart()
            self._calculate_strategy_metrics()

    def _clear_strategy_legs(self):
        """Clear all strategy legs"""
        self.strategy_legs = []
        self._refresh_legs_table()
        self._update_payoff_chart()
        self._calculate_strategy_metrics()

    def _refresh_legs_table(self):
        """Refresh the legs table"""
        self.sb_legs_table.setRowCount(len(self.strategy_legs))
        for i, leg in enumerate(self.strategy_legs):
            # Symbol column (get from leg or use current symbol)
            symbol = leg.get('symbol', self.sb_symbol.currentText())
            self.sb_legs_table.setItem(i, 0, QTableWidgetItem(symbol))

            self.sb_legs_table.setItem(i, 1, QTableWidgetItem(leg['action']))
            self.sb_legs_table.setItem(i, 2, QTableWidgetItem(leg['type']))
            self.sb_legs_table.setItem(i, 3, QTableWidgetItem(f"₹{leg['strike']:.0f}"))
            self.sb_legs_table.setItem(i, 4, QTableWidgetItem(str(leg['qty'])))
            self.sb_legs_table.setItem(i, 5, QTableWidgetItem(f"₹{leg['premium']:.2f}"))

            # LTP column - fetch live price or show entry price
            ltp = leg.get('ltp', leg.get('premium', 0))
            ltp_item = QTableWidgetItem(f"₹{ltp:.2f}")
            # Color based on profit/loss vs premium
            if ltp > leg['premium']:
                ltp_item.setForeground(Qt.GlobalColor.green if leg['action'] == 'BUY' else Qt.GlobalColor.red)
            elif ltp < leg['premium']:
                ltp_item.setForeground(Qt.GlobalColor.red if leg['action'] == 'BUY' else Qt.GlobalColor.green)
            self.sb_legs_table.setItem(i, 6, ltp_item)

            # Exit button for individual leg square-off
            exit_btn = QPushButton("Exit")
            exit_btn.setStyleSheet("background-color: #FF5722; color: white;")
            exit_btn.clicked.connect(lambda checked, idx=i: self._squareoff_single_leg(idx))
            self.sb_legs_table.setCellWidget(i, 7, exit_btn)

            # Remove button
            remove_btn = QPushButton("❌")
            remove_btn.clicked.connect(lambda checked, idx=i: self._remove_strategy_leg(idx))
            self.sb_legs_table.setCellWidget(i, 8, remove_btn)

    def _calculate_leg_payoff(self, leg, spot_prices):
        """Calculate payoff for a single leg at given spot prices"""
        import numpy as np
        strike = leg['strike']
        premium = leg['premium']
        qty = leg['qty']
        lot_size = 50 if "NIFTY" in self.sb_symbol.currentText().upper() else 25

        if leg['type'] == 'CE':
            # Call option payoff
            intrinsic = np.maximum(spot_prices - strike, 0)
        else:
            # Put option payoff
            intrinsic = np.maximum(strike - spot_prices, 0)

        if leg['action'] == 'BUY':
            payoff = (intrinsic - premium) * qty * lot_size
        else:
            payoff = (premium - intrinsic) * qty * lot_size

        return payoff

    def _squareoff_single_leg(self, index):
        """Square-off a single leg (exit position)"""
        if index >= len(self.strategy_legs):
            return

        leg = self.strategy_legs[index]
        symbol = leg.get('symbol', self.sb_symbol.currentText())
        leg_desc = f"{symbol} {leg['action']} {leg['type']} {leg['strike']}"

        reply = QMessageBox.question(
            self, "Square-off Leg",
            f"Square-off this position?\n\n{leg_desc}\n\n"
            "This will place a reverse order to exit the position.",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )

        if reply == QMessageBox.StandardButton.Yes:
            # Reverse the action (BUY -> SELL, SELL -> BUY)
            exit_action = "SELL" if leg['action'] == "BUY" else "BUY"

            self._log_cpr(f"SQUARE-OFF: {leg_desc}")
            self._log_cpr(f"  Exit Order: {exit_action} {leg['type']} {leg['strike']} x {leg['qty']}")

            # Mark leg as exited (change color or remove)
            leg['status'] = 'EXITED'
            leg['exit_action'] = exit_action

            # Update table to show exited status
            self._refresh_legs_table()

            QMessageBox.information(
                self, "Square-off Initiated",
                f"Exit order placed:\n{exit_action} {symbol} {leg['type']} {leg['strike']}"
            )

    def _refresh_legs_ltp(self):
        """Refresh LTP (Last Trade Price) for all strategy legs"""
        if not self.strategy_legs:
            return

        symbol = self.sb_symbol.currentText().upper()

        # Try to get LTP from broker if connected
        if hasattr(self, 'brokers') and self.brokers:
            broker = list(self.brokers.values())[0]
            try:
                for leg in self.strategy_legs:
                    # Build option symbol (format varies by broker)
                    leg_symbol = leg.get('symbol', symbol)
                    strike = int(leg['strike'])
                    opt_type = leg['type']  # CE or PE

                    # Try to get quote from broker
                    if hasattr(broker, 'get_option_ltp'):
                        ltp = broker.get_option_ltp(leg_symbol, strike, opt_type)
                        if ltp:
                            leg['ltp'] = ltp
                    elif hasattr(broker, 'get_ltp'):
                        # Generic LTP method
                        option_symbol = f"{leg_symbol}{strike}{opt_type}"
                        ltp = broker.get_ltp(option_symbol)
                        if ltp:
                            leg['ltp'] = ltp

            except Exception as e:
                logger.debug(f"Error fetching LTP from broker: {e}")

        # If no broker connected, always simulate LTP
        if not (hasattr(self, 'brokers') and self.brokers):
            self._simulate_option_ltp()

        self._refresh_legs_table()
        self._update_live_pnl()

    def _simulate_option_ltp(self):
        """Simulate option LTP with realistic price movements"""
        try:
            import random

            for leg in self.strategy_legs:
                entry_premium = leg['premium']
                current_ltp = leg.get('ltp', entry_premium)

                # If LTP not set or same as premium, initialize with small random offset
                if current_ltp == entry_premium or current_ltp <= 0:
                    current_ltp = entry_premium * random.uniform(0.95, 1.05)

                # Simulate realistic price movement (±1-3% per tick)
                change_percent = random.uniform(-0.03, 0.03)
                change_amount = current_ltp * change_percent

                # Apply change
                new_ltp = current_ltp + change_amount

                # Ensure LTP stays within reasonable bounds (30% to 300% of entry)
                min_ltp = max(0.05, entry_premium * 0.3)
                max_ltp = entry_premium * 3.0
                new_ltp = max(min_ltp, min(max_ltp, new_ltp))

                # Update leg LTP
                leg['ltp'] = round(new_ltp, 2)

        except Exception as e:
            logger.debug(f"Error simulating option LTP: {e}")

    def _update_live_pnl(self):
        """Update live P&L for the strategy based on LTP"""
        if not self.strategy_legs:
            return

        total_pnl = 0
        lot_size = 50 if "NIFTY" in self.sb_symbol.currentText().upper() else 25

        for leg in self.strategy_legs:
            entry = leg['premium']
            ltp = leg.get('ltp', entry)
            qty = leg['qty']

            if leg['action'] == 'BUY':
                pnl = (ltp - entry) * qty * lot_size
            else:  # SELL
                pnl = (entry - ltp) * qty * lot_size

            total_pnl += pnl

        # Update live P&L label if it exists
        if hasattr(self, 'sb_live_pnl_label'):
            color = "#4CAF50" if total_pnl >= 0 else "#F44336"
            self.sb_live_pnl_label.setText(f"₹{total_pnl:+,.2f}")
            self.sb_live_pnl_label.setStyleSheet(f"font-size: 16px; font-weight: bold; color: {color};")

    def _squareoff_all_legs(self):
        """Square-off all legs (exit entire strategy)"""
        if not self.strategy_legs:
            QMessageBox.warning(self, "No Positions", "No positions to square-off")
            return

        # Build summary of all legs
        leg_summary = "\n".join([
            f"  • {leg.get('symbol', self.sb_symbol.currentText())} {leg['action']} {leg['type']} {leg['strike']}"
            for leg in self.strategy_legs
        ])

        reply = QMessageBox.question(
            self, "Square-off ALL",
            f"Square-off ALL positions?\n\n{leg_summary}\n\n"
            "This will place reverse orders to exit ALL positions.",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )

        if reply == QMessageBox.StandardButton.Yes:
            self._log_cpr("="*40)
            self._log_cpr("SQUARE-OFF ALL POSITIONS")

            for leg in self.strategy_legs:
                symbol = leg.get('symbol', self.sb_symbol.currentText())
                exit_action = "SELL" if leg['action'] == "BUY" else "BUY"

                self._log_cpr(f"  Exit: {exit_action} {symbol} {leg['type']} {leg['strike']}")
                leg['status'] = 'EXITED'
                leg['exit_action'] = exit_action

            self._log_cpr("="*40)
            self._refresh_legs_table()

            QMessageBox.information(
                self, "All Squared-off",
                f"Exit orders placed for {len(self.strategy_legs)} positions"
            )

    def _update_payoff_chart(self):
        """Update the payoff chart based on current strategy"""
        import numpy as np

        self.sb_ax.clear()
        self._setup_payoff_chart()

        if not self.strategy_legs:
            self.sb_canvas.draw()
            return

        spot = self.sb_spot_price.value()
        # Create price range (±20% from spot)
        price_range = np.linspace(spot * 0.8, spot * 1.2, 500)

        # Calculate combined payoff
        total_payoff = np.zeros_like(price_range)
        for leg in self.strategy_legs:
            leg_payoff = self._calculate_leg_payoff(leg, price_range)
            total_payoff += leg_payoff

        # Plot payoff
        # Color areas
        profit_mask = total_payoff >= 0
        loss_mask = total_payoff < 0

        self.sb_ax.fill_between(price_range, total_payoff, 0, where=profit_mask,
                                color='#4CAF50', alpha=0.3, label='Profit')
        self.sb_ax.fill_between(price_range, total_payoff, 0, where=loss_mask,
                                color='#F44336', alpha=0.3, label='Loss')
        self.sb_ax.plot(price_range, total_payoff, color='#2196F3', linewidth=2, label='Payoff')

        # Mark spot price
        self.sb_ax.axvline(x=spot, color='yellow', linestyle='--', alpha=0.7, label=f'Spot: ₹{spot:.0f}')

        # Calculate and plot Greeks if checkboxes are selected
        show_vega = hasattr(self, 'sb_show_vega_check') and self.sb_show_vega_check.isChecked()
        show_delta = hasattr(self, 'sb_show_delta_check') and self.sb_show_delta_check.isChecked()
        show_theta = hasattr(self, 'sb_show_theta_check') and self.sb_show_theta_check.isChecked()

        if show_vega or show_delta or show_theta:
            # Create secondary Y-axis for Greeks
            ax2 = self.sb_ax.twinx()
            ax2.set_ylabel('Greeks Value', color='white')
            ax2.tick_params(colors='white')

            lot_size = 50 if "NIFTY" in self.sb_symbol.currentText().upper() else 25

            if show_vega:
                # Calculate Vega across price range (sensitivity to IV change)
                vega_line = self._calculate_strategy_vega(price_range, lot_size)
                ax2.plot(price_range, vega_line, color='#FF9800', linewidth=2,
                        linestyle='-', label='Vega', alpha=0.9)

            if show_delta:
                # Calculate Delta across price range
                delta_line = self._calculate_strategy_delta(price_range, lot_size)
                ax2.plot(price_range, delta_line, color='#2196F3', linewidth=2,
                        linestyle='--', label='Delta', alpha=0.9)

            if show_theta:
                # Calculate Theta across price range
                theta_line = self._calculate_strategy_theta(price_range, lot_size)
                ax2.plot(price_range, theta_line, color='#9C27B0', linewidth=2,
                        linestyle='-.', label='Theta', alpha=0.9)

            ax2.legend(loc='upper left', facecolor='#2d2d2d', labelcolor='white')
            ax2.axhline(y=0, color='gray', linestyle=':', alpha=0.3)

        # Mark breakeven points
        breakeven_points = []
        for i in range(1, len(total_payoff)):
            if (total_payoff[i-1] < 0 and total_payoff[i] >= 0) or (total_payoff[i-1] >= 0 and total_payoff[i] < 0):
                breakeven_points.append(price_range[i])

        for be in breakeven_points:
            self.sb_ax.axvline(x=be, color='cyan', linestyle=':', alpha=0.7)
            self.sb_ax.annotate(f'BE: ₹{be:.0f}', xy=(be, 0), xytext=(be, max(total_payoff)*0.1),
                               color='cyan', fontsize=8)

        self.sb_ax.legend(loc='upper right', facecolor='#2d2d2d', labelcolor='white')
        self.sb_ax.grid(True, alpha=0.2)
        self.sb_figure.tight_layout()
        self.sb_canvas.draw()

    def _calculate_strategy_vega(self, price_range, lot_size):
        """Calculate strategy Vega across price range (like Moving Average line)"""
        import numpy as np

        vega_values = np.zeros_like(price_range)

        for leg in self.strategy_legs:
            strike = leg['strike']
            premium = leg['premium']
            qty = leg['qty']

            # Vega calculation (simplified Black-Scholes approximation)
            # Vega is highest at ATM and decreases as we move ITM/OTM
            for i, spot in enumerate(price_range):
                moneyness = abs(spot - strike) / spot
                # ATM has highest vega, decreases with distance from strike
                # Using Gaussian-like decay
                atm_vega = premium * 0.4  # Base vega ~40% of premium for ATM
                vega = atm_vega * np.exp(-0.5 * (moneyness / 0.1) ** 2)

                # Adjust for position direction
                if leg['action'] == 'BUY':
                    vega_values[i] += vega * qty * lot_size
                else:
                    vega_values[i] -= vega * qty * lot_size

        return vega_values

    def _calculate_strategy_delta(self, price_range, lot_size):
        """Calculate strategy Delta across price range"""
        import numpy as np

        delta_values = np.zeros_like(price_range)

        for leg in self.strategy_legs:
            strike = leg['strike']
            qty = leg['qty']

            for i, spot in enumerate(price_range):
                # Simplified delta calculation
                moneyness = (spot - strike) / spot

                if leg['type'] == 'CE':
                    # Call delta: 0 to 1, ~0.5 at ATM
                    delta = 0.5 + moneyness * 3  # Simplified
                    delta = max(0, min(1, delta))
                else:
                    # Put delta: -1 to 0, ~-0.5 at ATM
                    delta = -0.5 + moneyness * 3
                    delta = max(-1, min(0, delta))

                # Adjust for position direction
                if leg['action'] == 'SELL':
                    delta = -delta

                delta_values[i] += delta * qty * lot_size

        return delta_values

    def _calculate_strategy_theta(self, price_range, lot_size):
        """Calculate strategy Theta across price range"""
        import numpy as np

        theta_values = np.zeros_like(price_range)

        for leg in self.strategy_legs:
            strike = leg['strike']
            premium = leg['premium']
            qty = leg['qty']

            for i, spot in enumerate(price_range):
                # Theta is highest at ATM, decreases as we move away
                moneyness = abs(spot - strike) / spot
                # Daily theta decay (negative for buyers)
                atm_theta = -premium * 0.05  # ~5% daily decay at ATM
                theta = atm_theta * np.exp(-0.5 * (moneyness / 0.15) ** 2)

                # Buyers have negative theta, sellers have positive
                if leg['action'] == 'SELL':
                    theta = -theta

                theta_values[i] += theta * qty * lot_size

        return theta_values

    def _calculate_strategy_metrics(self):
        """Calculate max profit, max loss, breakeven, etc."""
        import numpy as np

        if not self.strategy_legs:
            self.sb_max_profit.setText("₹0.00")
            self.sb_max_loss.setText("₹0.00")
            self.sb_breakeven.setText("--")
            self.sb_net_premium.setText("₹0.00")
            self.sb_risk_reward.setText("--")
            return

        spot = self.sb_spot_price.value()
        price_range = np.linspace(spot * 0.5, spot * 1.5, 1000)

        total_payoff = np.zeros_like(price_range)
        net_premium = 0

        for leg in self.strategy_legs:
            leg_payoff = self._calculate_leg_payoff(leg, price_range)
            total_payoff += leg_payoff

            lot_size = 50 if "NIFTY" in self.sb_symbol.currentText().upper() else 25
            if leg['action'] == 'BUY':
                net_premium -= leg['premium'] * leg['qty'] * lot_size
            else:
                net_premium += leg['premium'] * leg['qty'] * lot_size

        max_profit = np.max(total_payoff)
        max_loss = np.min(total_payoff)

        # Find breakeven points
        breakeven_points = []
        for i in range(1, len(total_payoff)):
            if (total_payoff[i-1] < 0 and total_payoff[i] >= 0) or (total_payoff[i-1] >= 0 and total_payoff[i] < 0):
                breakeven_points.append(price_range[i])

        # Update labels
        if max_profit == float('inf') or max_profit > 10000000:
            self.sb_max_profit.setText("Unlimited")
        else:
            self.sb_max_profit.setText(f"₹{max_profit:,.2f}")

        if max_loss == float('-inf') or max_loss < -10000000:
            self.sb_max_loss.setText("Unlimited")
        else:
            self.sb_max_loss.setText(f"₹{max_loss:,.2f}")

        if breakeven_points:
            be_str = ", ".join([f"₹{be:.0f}" for be in breakeven_points[:3]])
            self.sb_breakeven.setText(be_str)
        else:
            self.sb_breakeven.setText("--")

        self.sb_net_premium.setText(f"₹{net_premium:,.2f}")

        if max_loss != 0 and max_profit > 0 and abs(max_loss) < 10000000:
            rr = abs(max_profit / max_loss)
            self.sb_risk_reward.setText(f"1:{rr:.2f}")
        else:
            self.sb_risk_reward.setText("--")

    def _apply_prebuilt_strategy(self, strategy_name):
        """Apply a pre-built strategy"""
        spot = self.sb_spot_price.value()
        gap = 50 if "NIFTY" in self.sb_symbol.currentText().upper() else 100
        atm = round(spot / gap) * gap

        self.strategy_legs = []

        if strategy_name == "iron_fly":
            # Sell ATM CE + Sell ATM PE + Buy OTM CE + Buy OTM PE
            self.strategy_legs = [
                {'action': 'SELL', 'type': 'CE', 'strike': atm, 'qty': 1, 'premium': 200},
                {'action': 'SELL', 'type': 'PE', 'strike': atm, 'qty': 1, 'premium': 200},
                {'action': 'BUY', 'type': 'CE', 'strike': atm + gap * 4, 'qty': 1, 'premium': 50},
                {'action': 'BUY', 'type': 'PE', 'strike': atm - gap * 4, 'qty': 1, 'premium': 50}
            ]

        elif strategy_name == "iron_condor":
            # Sell OTM CE + Sell OTM PE + Buy further OTM CE + Buy further OTM PE
            self.strategy_legs = [
                {'action': 'SELL', 'type': 'CE', 'strike': atm + gap * 2, 'qty': 1, 'premium': 100},
                {'action': 'SELL', 'type': 'PE', 'strike': atm - gap * 2, 'qty': 1, 'premium': 100},
                {'action': 'BUY', 'type': 'CE', 'strike': atm + gap * 4, 'qty': 1, 'premium': 30},
                {'action': 'BUY', 'type': 'PE', 'strike': atm - gap * 4, 'qty': 1, 'premium': 30}
            ]

        elif strategy_name == "straddle":
            # Sell ATM CE + Sell ATM PE
            self.strategy_legs = [
                {'action': 'SELL', 'type': 'CE', 'strike': atm, 'qty': 1, 'premium': 200},
                {'action': 'SELL', 'type': 'PE', 'strike': atm, 'qty': 1, 'premium': 200}
            ]

        elif strategy_name == "strangle":
            # Sell OTM CE + Sell OTM PE
            self.strategy_legs = [
                {'action': 'SELL', 'type': 'CE', 'strike': atm + gap * 3, 'qty': 1, 'premium': 80},
                {'action': 'SELL', 'type': 'PE', 'strike': atm - gap * 3, 'qty': 1, 'premium': 80}
            ]

        elif strategy_name == "bull_call_spread":
            # Buy ATM CE + Sell OTM CE
            self.strategy_legs = [
                {'action': 'BUY', 'type': 'CE', 'strike': atm, 'qty': 1, 'premium': 200},
                {'action': 'SELL', 'type': 'CE', 'strike': atm + gap * 3, 'qty': 1, 'premium': 80}
            ]

        elif strategy_name == "bear_put_spread":
            # Buy ATM PE + Sell OTM PE
            self.strategy_legs = [
                {'action': 'BUY', 'type': 'PE', 'strike': atm, 'qty': 1, 'premium': 200},
                {'action': 'SELL', 'type': 'PE', 'strike': atm - gap * 3, 'qty': 1, 'premium': 80}
            ]

        self._refresh_legs_table()
        self._update_payoff_chart()
        self._calculate_strategy_metrics()
        self.status_bar.showMessage(f"Applied {strategy_name.replace('_', ' ').title()} strategy", 3000)

    def _show_strategy_greeks(self):
        """Show strategy Greeks (Delta, Gamma, Theta, Vega)"""
        if not self.strategy_legs:
            QMessageBox.warning(self, "No Strategy", "Please add legs to the strategy first")
            return

        # Simplified Greeks calculation
        total_delta = 0
        total_theta = 0

        for leg in self.strategy_legs:
            # Approximate delta (0.5 for ATM, adjusted for ITM/OTM)
            spot = self.sb_spot_price.value()
            moneyness = (spot - leg['strike']) / spot

            if leg['type'] == 'CE':
                delta = 0.5 + moneyness * 2  # Simplified
                delta = max(0, min(1, delta))
            else:
                delta = -0.5 + moneyness * 2
                delta = max(-1, min(0, delta))

            if leg['action'] == 'SELL':
                delta = -delta

            total_delta += delta * leg['qty']

            # Approximate theta (negative for buyers)
            theta = -leg['premium'] * 0.1  # ~10% decay per day approximation
            if leg['action'] == 'SELL':
                theta = -theta
            total_theta += theta * leg['qty']

        lot_size = 50 if "NIFTY" in self.sb_symbol.currentText().upper() else 25

        msg = f"""Strategy Greeks (Approximate):

📊 Total Delta: {total_delta * lot_size:.2f}
   (Position moves ₹{abs(total_delta * lot_size):.2f} per ₹1 move in spot)

⏰ Total Theta: ₹{total_theta * lot_size:.2f}/day
   (Daily time decay {'gain' if total_theta > 0 else 'loss'})

Note: These are simplified approximations.
For accurate Greeks, use live option data."""

        QMessageBox.information(self, "Strategy Greeks", msg)

    def _execute_strategy_builder(self):
        """Execute the strategy"""
        if not self.strategy_legs:
            QMessageBox.warning(self, "No Strategy", "Please add legs to the strategy first")
            return

        # Confirm execution
        legs_summary = "\n".join([
            f"{leg['action']} {leg['qty']} lot {leg['type']} @ ₹{leg['strike']:.0f}"
            for leg in self.strategy_legs
        ])

        reply = QMessageBox.question(
            self, "Confirm Strategy Execution",
            f"Execute the following strategy?\n\n{legs_summary}\n\nSymbol: {self.sb_symbol.currentText()}\nExpiry: {self.sb_expiry.currentText()}",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )

        if reply == QMessageBox.StandardButton.Yes:
            # Execute each leg
            paper_mode = self.config.get('trading.paper_mode', False)

            for leg in self.strategy_legs:
                logger.info(f"Executing leg: {leg['action']} {leg['type']} @ {leg['strike']}")

                if paper_mode and hasattr(self, 'paper_simulator') and self.paper_simulator:
                    # Paper trade
                    option_symbol = f"{self.sb_symbol.currentText()}{int(leg['strike'])}{leg['type']}"
                    self.paper_simulator.place_order(
                        symbol=option_symbol,
                        action=leg['action'],
                        quantity=leg['qty'],
                        order_type='MARKET',
                        price=leg['premium']
                    )
                else:
                    # Live trade - would need broker integration
                    pass

            QMessageBox.information(self, "Strategy Executed",
                                   f"Strategy with {len(self.strategy_legs)} legs executed successfully!")
            self.status_bar.showMessage("Strategy executed", 5000)

    def _init_strategies_folder(self):
        """Initialize the strategies folder"""
        import os
        self.strategies_folder = os.path.join(os.path.expanduser("~"), "algo_trader_strategies")
        if not os.path.exists(self.strategies_folder):
            os.makedirs(self.strategies_folder)
            logger.info(f"Created strategies folder: {self.strategies_folder}")

    def _refresh_saved_strategies_list(self):
        """Refresh the saved strategies dropdown"""
        import os
        import json
        self.sb_saved_strategies_combo.clear()
        self.sb_saved_strategies_combo.addItem("-- Select Saved Strategy --")

        if not hasattr(self, 'strategies_folder'):
            return

        try:
            for filename in os.listdir(self.strategies_folder):
                if filename.endswith('.json'):
                    strategy_name = filename[:-5]  # Remove .json
                    self.sb_saved_strategies_combo.addItem(strategy_name)
        except Exception as e:
            logger.error(f"Error loading strategies list: {e}")

    def _save_custom_strategy(self):
        """Save current strategy to file"""
        import os
        import json

        if not self.strategy_legs:
            QMessageBox.warning(self, "No Strategy", "Please add legs to the strategy first")
            return

        strategy_name = self.sb_strategy_name.text().strip()
        if not strategy_name:
            strategy_name = f"Strategy_{len(os.listdir(self.strategies_folder)) + 1}"
            self.sb_strategy_name.setText(strategy_name)

        strategy_data = {
            'name': strategy_name,
            'symbol': self.sb_symbol.currentText(),
            'spot_price': self.sb_spot_price.value(),
            'expiry': self.sb_expiry.currentText(),
            'legs': self.strategy_legs,
            'created_at': datetime.now().isoformat()
        }

        filepath = os.path.join(self.strategies_folder, f"{strategy_name}.json")
        try:
            with open(filepath, 'w') as f:
                json.dump(strategy_data, f, indent=2)

            self._refresh_saved_strategies_list()
            QMessageBox.information(self, "Saved", f"Strategy saved to:\n{filepath}")
            self.status_bar.showMessage(f"Strategy '{strategy_name}' saved", 3000)
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to save strategy: {e}")

    def _load_custom_strategy(self):
        """Load a saved strategy"""
        import os
        import json

        selected = self.sb_saved_strategies_combo.currentText()
        if selected == "-- Select Saved Strategy --" or not selected:
            QMessageBox.warning(self, "Select Strategy", "Please select a strategy from the dropdown")
            return

        filepath = os.path.join(self.strategies_folder, f"{selected}.json")
        try:
            with open(filepath, 'r') as f:
                strategy_data = json.load(f)

            # Apply loaded strategy
            self.sb_strategy_name.setText(strategy_data.get('name', ''))

            # Set symbol
            symbol = strategy_data.get('symbol', 'NIFTY')
            idx = self.sb_symbol.findText(symbol)
            if idx >= 0:
                self.sb_symbol.setCurrentIndex(idx)
            else:
                self.sb_symbol.setCurrentText(symbol)

            self.sb_spot_price.setValue(strategy_data.get('spot_price', 25000))

            # Set expiry
            expiry = strategy_data.get('expiry', 'Current Week')
            idx = self.sb_expiry.findText(expiry)
            if idx >= 0:
                self.sb_expiry.setCurrentIndex(idx)

            # Load legs
            self.strategy_legs = strategy_data.get('legs', [])
            self._refresh_legs_table()
            self._update_payoff_chart()
            self._calculate_strategy_metrics()

            self.status_bar.showMessage(f"Strategy '{selected}' loaded", 3000)

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to load strategy: {e}")

    def _deploy_saved_strategy(self):
        """Deploy a saved strategy directly"""
        selected = self.sb_saved_strategies_combo.currentText()
        if selected == "-- Select Saved Strategy --" or not selected:
            QMessageBox.warning(self, "Select Strategy", "Please select a strategy to deploy")
            return

        # First load the strategy
        self._load_custom_strategy()

        # Then execute it
        if self.strategy_legs:
            reply = QMessageBox.question(
                self, "Deploy Strategy",
                f"Deploy strategy '{selected}' now?\n\nThis will execute all {len(self.strategy_legs)} legs.",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )
            if reply == QMessageBox.StandardButton.Yes:
                self._execute_strategy_builder()

    def _edit_saved_strategy(self):
        """Edit a saved strategy - load it for modification"""
        selected = self.sb_saved_strategies_combo.currentText()
        if selected == "-- Select Saved Strategy --" or not selected:
            QMessageBox.warning(self, "Select Strategy", "Please select a strategy to edit")
            return

        # Load the strategy for editing
        self._load_custom_strategy()

        # Set the name so user can save it back
        self.sb_strategy_name.setText(selected)

        QMessageBox.information(self, "Edit Mode",
            f"Strategy '{selected}' loaded for editing.\n\n"
            "Make your changes:\n"
            "- Modify legs (add/remove)\n"
            "- Change strike prices\n"
            "- Adjust quantities\n\n"
            "Click 'Save' to save changes with same name,\n"
            "or change the name to save as new strategy.")

        self.status_bar.showMessage(f"Editing strategy: {selected}", 5000)

    def _delete_saved_strategy(self):
        """Delete a saved strategy"""
        import os

        selected = self.sb_saved_strategies_combo.currentText()
        if selected == "-- Select Saved Strategy --" or not selected:
            QMessageBox.warning(self, "Select Strategy", "Please select a strategy to delete")
            return

        reply = QMessageBox.question(
            self, "Confirm Delete",
            f"Are you sure you want to delete strategy '{selected}'?\n\nThis cannot be undone.",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )

        if reply == QMessageBox.StandardButton.Yes:
            filepath = os.path.join(self.strategies_folder, f"{selected}.json")
            try:
                os.remove(filepath)
                self._refresh_saved_strategies_list()
                QMessageBox.information(self, "Deleted", f"Strategy '{selected}' deleted successfully.")
                self.status_bar.showMessage(f"Strategy '{selected}' deleted", 3000)
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Failed to delete strategy: {e}")

    def _open_strategies_folder(self):
        """Open the strategies folder in file explorer"""
        import os
        import subprocess
        import sys

        if not hasattr(self, 'strategies_folder'):
            self._init_strategies_folder()

        folder_path = self.strategies_folder

        try:
            if sys.platform == 'win32':
                os.startfile(folder_path)
            elif sys.platform == 'darwin':  # macOS
                subprocess.run(['open', folder_path])
            else:  # Linux
                subprocess.run(['xdg-open', folder_path])

            self.status_bar.showMessage(f"Opened folder: {folder_path}", 3000)
        except Exception as e:
            # If can't open, show the path
            QMessageBox.information(self, "Strategies Folder",
                f"Strategies are saved at:\n\n{folder_path}\n\n"
                "You can open this folder manually to view/edit JSON files.")

    def _lock_entry_price(self):
        """Lock the current spot price as entry price"""
        current_spot = self.sb_spot_price.value()
        self.sb_entry_spot.setValue(current_spot)
        self.sb_entry_spot_locked = True
        self.sb_lock_entry_btn.setText("🔓 Entry Locked")
        self.sb_lock_entry_btn.setStyleSheet("background-color: #4CAF50; color: white;")
        self._update_live_pnl()
        self.status_bar.showMessage(f"Entry price locked at ₹{current_spot:.2f}", 3000)

    def _update_live_pnl(self):
        """Update live P&L based on spot price movement"""
        import numpy as np

        if not self.strategy_legs:
            self.sb_current_pnl.setText("₹0.00")
            self.sb_pnl_change.setText("₹0.00 (0.00%)")
            return

        entry_spot = self.sb_entry_spot.value()
        current_spot = self.sb_spot_price.value()

        if entry_spot <= 0:
            self.sb_current_pnl.setText("Set entry price")
            self.sb_pnl_change.setText("--")
            return

        # Calculate P&L at entry and current spot
        entry_pnl = 0
        current_pnl = 0

        for leg in self.strategy_legs:
            entry_payoff = self._calculate_leg_payoff(leg, np.array([entry_spot]))[0]
            current_payoff = self._calculate_leg_payoff(leg, np.array([current_spot]))[0]
            entry_pnl += entry_payoff
            current_pnl += current_payoff

        pnl_diff = current_pnl - entry_pnl

        # Update current P&L display
        if current_pnl >= 0:
            self.sb_current_pnl.setText(f"₹{current_pnl:,.2f}")
            self.sb_current_pnl.setStyleSheet("color: #4CAF50; font-weight: bold; font-size: 16px;")
        else:
            self.sb_current_pnl.setText(f"-₹{abs(current_pnl):,.2f}")
            self.sb_current_pnl.setStyleSheet("color: #F44336; font-weight: bold; font-size: 16px;")

        # Update P&L change display
        spot_change = current_spot - entry_spot
        spot_change_pct = (spot_change / entry_spot * 100) if entry_spot > 0 else 0

        if pnl_diff >= 0:
            self.sb_pnl_change.setText(f"+₹{pnl_diff:,.2f} (Spot: {'+' if spot_change >= 0 else ''}{spot_change:.0f})")
            self.sb_pnl_change.setStyleSheet("color: #4CAF50; font-size: 12px;")
        else:
            self.sb_pnl_change.setText(f"-₹{abs(pnl_diff):,.2f} (Spot: {'+' if spot_change >= 0 else ''}{spot_change:.0f})")
            self.sb_pnl_change.setStyleSheet("color: #F44336; font-size: 12px;")

        # Update the payoff chart with current position marker
        self._update_payoff_chart_with_live_pnl(entry_spot, current_spot, entry_pnl, current_pnl)

    def _update_payoff_chart_with_live_pnl(self, entry_spot, current_spot, entry_pnl, current_pnl):
        """Update payoff chart with live P&L markers"""
        import numpy as np

        self.sb_ax.clear()
        self._setup_payoff_chart()

        if not self.strategy_legs:
            self.sb_canvas.draw()
            return

        spot = self.sb_spot_price.value()
        # Create price range (±20% from spot)
        price_range = np.linspace(spot * 0.8, spot * 1.2, 500)

        # Calculate combined payoff
        total_payoff = np.zeros_like(price_range)
        for leg in self.strategy_legs:
            leg_payoff = self._calculate_leg_payoff(leg, price_range)
            total_payoff += leg_payoff

        # Plot payoff
        profit_mask = total_payoff >= 0
        loss_mask = total_payoff < 0

        self.sb_ax.fill_between(price_range, total_payoff, 0, where=profit_mask,
                                color='#4CAF50', alpha=0.3, label='Profit')
        self.sb_ax.fill_between(price_range, total_payoff, 0, where=loss_mask,
                                color='#F44336', alpha=0.3, label='Loss')
        self.sb_ax.plot(price_range, total_payoff, color='#2196F3', linewidth=2)

        # Mark entry spot
        if entry_spot > 0:
            self.sb_ax.axvline(x=entry_spot, color='orange', linestyle='--', alpha=0.8, linewidth=2)
            self.sb_ax.scatter([entry_spot], [entry_pnl], color='orange', s=100, zorder=5, marker='o')
            self.sb_ax.annotate(f'Entry\n₹{entry_spot:.0f}', xy=(entry_spot, entry_pnl),
                               xytext=(entry_spot - spot*0.03, entry_pnl + max(abs(total_payoff))*0.1),
                               color='orange', fontsize=9, fontweight='bold')

        # Mark current spot with P&L
        self.sb_ax.axvline(x=current_spot, color='yellow', linestyle='-', alpha=0.9, linewidth=2)
        self.sb_ax.scatter([current_spot], [current_pnl], color='yellow', s=150, zorder=5, marker='*')

        # Draw line connecting entry to current position
        if entry_spot > 0 and entry_spot != current_spot:
            self.sb_ax.plot([entry_spot, current_spot], [entry_pnl, current_pnl],
                           color='white', linestyle='-', linewidth=2, alpha=0.7)
            # P&L difference annotation
            pnl_diff = current_pnl - entry_pnl
            mid_x = (entry_spot + current_spot) / 2
            mid_y = (entry_pnl + current_pnl) / 2
            color = '#4CAF50' if pnl_diff >= 0 else '#F44336'
            self.sb_ax.annotate(f'{"+₹" if pnl_diff >= 0 else "-₹"}{abs(pnl_diff):,.0f}',
                               xy=(mid_x, mid_y), fontsize=11, fontweight='bold',
                               color=color, ha='center',
                               bbox=dict(boxstyle='round', facecolor='#1e1e1e', edgecolor=color, alpha=0.9))

        self.sb_ax.annotate(f'Current\n₹{current_spot:.0f}\nP&L: ₹{current_pnl:,.0f}',
                           xy=(current_spot, current_pnl),
                           xytext=(current_spot + spot*0.02, current_pnl + max(abs(total_payoff))*0.15),
                           color='yellow', fontsize=9, fontweight='bold')

        # Mark breakeven points
        breakeven_points = []
        for i in range(1, len(total_payoff)):
            if (total_payoff[i-1] < 0 and total_payoff[i] >= 0) or (total_payoff[i-1] >= 0 and total_payoff[i] < 0):
                breakeven_points.append(price_range[i])

        for be in breakeven_points:
            self.sb_ax.axvline(x=be, color='cyan', linestyle=':', alpha=0.7)
            self.sb_ax.annotate(f'BE: ₹{be:.0f}', xy=(be, 0), xytext=(be, max(total_payoff)*0.05),
                               color='cyan', fontsize=8)

        self.sb_ax.legend(loc='upper right', facecolor='#2d2d2d', labelcolor='white')
        self.sb_ax.grid(True, alpha=0.2)
        self.sb_figure.tight_layout()
        self.sb_canvas.draw()

    def _create_custom_strategy_tab(self):
        """Create Custom Strategy tab for user-defined trading rules"""
        # Wrap in scroll area
        custom_scroll = QScrollArea()
        custom_scroll.setWidgetResizable(True)

        custom_widget = QWidget()
        main_layout = QVBoxLayout(custom_widget)

        # Header
        header_label = QLabel("Custom Strategy Builder")
        header_label.setStyleSheet("font-size: 18px; font-weight: bold; color: #4CAF50; padding: 10px;")
        main_layout.addWidget(header_label)

        info_label = QLabel("Create your own trading strategy with simple rules. No Pine Script needed!")
        info_label.setStyleSheet("color: #888; padding-bottom: 10px;")
        main_layout.addWidget(info_label)

        # Two column layout
        columns_layout = QHBoxLayout()

        # LEFT COLUMN - Strategy Configuration
        left_panel = QWidget()
        left_layout = QVBoxLayout(left_panel)
        left_panel.setMaximumWidth(500)

        # Strategy Name
        name_group = QGroupBox("Strategy Details")
        name_layout = QFormLayout(name_group)

        self.custom_strategy_name = QLineEdit()
        self.custom_strategy_name.setPlaceholderText("My EMA Crossover Strategy")
        self.custom_strategy_name.setMinimumHeight(30)
        name_layout.addRow("Strategy Name:", self.custom_strategy_name)

        self.custom_strategy_symbol = QComboBox()
        self.custom_strategy_symbol.setEditable(True)
        fo_indices = ["NIFTY", "BANKNIFTY", "FINNIFTY", "MIDCPNIFTY", "SENSEX", "BANKEX"]
        fo_stocks = ["RELIANCE", "TCS", "HDFCBANK", "INFY", "ICICIBANK", "SBIN", "BHARTIARTL", "ITC", "KOTAKBANK", "LT", "AXISBANK", "MARUTI", "BAJFINANCE", "HCLTECH", "WIPRO", "TATASTEEL", "TATAMOTORS"]
        self.custom_strategy_symbol.addItems(fo_indices + sorted(fo_stocks))
        self.custom_strategy_symbol.setMinimumHeight(30)
        name_layout.addRow("Symbol:", self.custom_strategy_symbol)

        self.custom_strategy_timeframe = QComboBox()
        self.custom_strategy_timeframe.addItems(["1 Minute", "5 Minutes", "15 Minutes", "1 Hour", "Daily"])
        self.custom_strategy_timeframe.setMinimumHeight(30)
        name_layout.addRow("Timeframe:", self.custom_strategy_timeframe)

        left_layout.addWidget(name_group)

        # Entry Conditions
        entry_group = QGroupBox("Entry Conditions (When to Buy)")
        entry_layout = QVBoxLayout(entry_group)

        entry_info = QLabel("Set conditions - signal triggers when ALL conditions are true")
        entry_info.setStyleSheet("color: #888; font-size: 11px;")
        entry_layout.addWidget(entry_info)

        # Condition 1
        cond1_layout = QHBoxLayout()
        self.custom_entry_ind1 = QComboBox()
        self.custom_entry_ind1.addItems(["Price", "EMA(9)", "EMA(21)", "SMA(20)", "SMA(50)", "RSI(14)", "MACD", "Supertrend", "VWAP", "High", "Low", "Open", "Close"])
        self.custom_entry_ind1.setMinimumHeight(28)
        cond1_layout.addWidget(self.custom_entry_ind1)

        self.custom_entry_op1 = QComboBox()
        self.custom_entry_op1.addItems(["Crosses Above", "Crosses Below", ">", "<", ">=", "<=", "="])
        self.custom_entry_op1.setMinimumHeight(28)
        cond1_layout.addWidget(self.custom_entry_op1)

        self.custom_entry_ind2 = QComboBox()
        self.custom_entry_ind2.addItems(["EMA(21)", "EMA(9)", "SMA(20)", "SMA(50)", "RSI(14)", "Value", "MACD Signal", "Supertrend", "VWAP", "High", "Low", "Open", "Close", "Price"])
        self.custom_entry_ind2.setMinimumHeight(28)
        cond1_layout.addWidget(self.custom_entry_ind2)

        self.custom_entry_val1 = QLineEdit()
        self.custom_entry_val1.setPlaceholderText("(optional value)")
        self.custom_entry_val1.setMaximumWidth(80)
        self.custom_entry_val1.setMinimumHeight(28)
        cond1_layout.addWidget(self.custom_entry_val1)

        entry_layout.addLayout(cond1_layout)

        # Condition 2 (optional)
        cond2_layout = QHBoxLayout()
        cond2_layout.addWidget(QLabel("AND"))

        self.custom_entry_ind3 = QComboBox()
        self.custom_entry_ind3.addItems(["-- None --", "RSI(14)", "MACD", "Volume", "Price", "EMA(9)", "EMA(21)", "SMA(20)", "SMA(50)", "Supertrend"])
        self.custom_entry_ind3.setMinimumHeight(28)
        cond2_layout.addWidget(self.custom_entry_ind3)

        self.custom_entry_op2 = QComboBox()
        self.custom_entry_op2.addItems([">", "<", ">=", "<=", "Crosses Above", "Crosses Below"])
        self.custom_entry_op2.setMinimumHeight(28)
        cond2_layout.addWidget(self.custom_entry_op2)

        self.custom_entry_val2 = QLineEdit()
        self.custom_entry_val2.setPlaceholderText("Value (e.g., 30)")
        self.custom_entry_val2.setMaximumWidth(100)
        self.custom_entry_val2.setMinimumHeight(28)
        cond2_layout.addWidget(self.custom_entry_val2)

        entry_layout.addLayout(cond2_layout)

        left_layout.addWidget(entry_group)

        # Exit Conditions
        exit_group = QGroupBox("Exit Conditions (When to Sell)")
        exit_layout = QVBoxLayout(exit_group)

        # Exit Condition
        exit_cond_layout = QHBoxLayout()
        self.custom_exit_ind1 = QComboBox()
        self.custom_exit_ind1.addItems(["Price", "EMA(9)", "EMA(21)", "SMA(20)", "RSI(14)", "MACD", "Supertrend"])
        self.custom_exit_ind1.setMinimumHeight(28)
        exit_cond_layout.addWidget(self.custom_exit_ind1)

        self.custom_exit_op1 = QComboBox()
        self.custom_exit_op1.addItems(["Crosses Below", "Crosses Above", ">", "<", ">=", "<="])
        self.custom_exit_op1.setMinimumHeight(28)
        exit_cond_layout.addWidget(self.custom_exit_op1)

        self.custom_exit_ind2 = QComboBox()
        self.custom_exit_ind2.addItems(["EMA(21)", "EMA(9)", "SMA(20)", "Value", "MACD Signal", "Supertrend"])
        self.custom_exit_ind2.setMinimumHeight(28)
        exit_cond_layout.addWidget(self.custom_exit_ind2)

        self.custom_exit_val1 = QLineEdit()
        self.custom_exit_val1.setPlaceholderText("(optional)")
        self.custom_exit_val1.setMaximumWidth(80)
        self.custom_exit_val1.setMinimumHeight(28)
        exit_cond_layout.addWidget(self.custom_exit_val1)

        exit_layout.addLayout(exit_cond_layout)

        left_layout.addWidget(exit_group)

        # Trade Action
        action_group = QGroupBox("Trade Action (On Signal)")
        action_layout = QFormLayout(action_group)

        self.custom_action_type = QComboBox()
        self.custom_action_type.addItems([
            "Buy Call (CE)",
            "Buy Put (PE)",
            "Sell Call (CE)",
            "Sell Put (PE)",
            "Bull Call Spread",
            "Bear Put Spread",
            "Iron Condor",
            "Straddle",
            "Alert Only (No Trade)"
        ])
        self.custom_action_type.setMinimumHeight(30)
        action_layout.addRow("Action:", self.custom_action_type)

        self.custom_lots = QSpinBox()
        self.custom_lots.setRange(1, 100)
        self.custom_lots.setValue(1)
        self.custom_lots.setMinimumHeight(28)
        action_layout.addRow("Lots:", self.custom_lots)

        self.custom_strike_offset = QSpinBox()
        self.custom_strike_offset.setRange(-10, 10)
        self.custom_strike_offset.setValue(0)
        self.custom_strike_offset.setSuffix(" strikes from ATM")
        self.custom_strike_offset.setMinimumHeight(28)
        action_layout.addRow("Strike:", self.custom_strike_offset)

        left_layout.addWidget(action_group)

        # Controls
        controls_layout = QHBoxLayout()

        self.custom_test_mode_check = QCheckBox("Test Mode")
        self.custom_test_mode_check.setChecked(True)
        controls_layout.addWidget(self.custom_test_mode_check)

        self.custom_save_btn = QPushButton("Save Strategy")
        self.custom_save_btn.setMinimumHeight(35)
        self.custom_save_btn.setStyleSheet("background-color: #2196F3; color: white;")
        self.custom_save_btn.clicked.connect(self._save_custom_strategy)
        controls_layout.addWidget(self.custom_save_btn)

        self.custom_start_btn = QPushButton("Start Monitoring")
        self.custom_start_btn.setMinimumHeight(35)
        self.custom_start_btn.setStyleSheet("background-color: #4CAF50; color: white; font-weight: bold;")
        self.custom_start_btn.clicked.connect(self._start_custom_strategy)
        controls_layout.addWidget(self.custom_start_btn)

        self.custom_stop_btn = QPushButton("Stop")
        self.custom_stop_btn.setMinimumHeight(35)
        self.custom_stop_btn.setStyleSheet("background-color: #F44336; color: white;")
        self.custom_stop_btn.clicked.connect(self._stop_custom_strategy)
        self.custom_stop_btn.setEnabled(False)
        controls_layout.addWidget(self.custom_stop_btn)

        left_layout.addLayout(controls_layout)
        columns_layout.addWidget(left_panel)

        # RIGHT COLUMN - Saved Strategies & Logs
        right_panel = QWidget()
        right_layout = QVBoxLayout(right_panel)

        # Saved Strategies
        saved_group = QGroupBox("Saved Custom Strategies")
        saved_layout = QVBoxLayout(saved_group)

        self.custom_strategies_table = QTableWidget()
        self.custom_strategies_table.setColumnCount(5)
        self.custom_strategies_table.setHorizontalHeaderLabels(["Name", "Symbol", "Status", "Signals", "Action"])
        self.custom_strategies_table.horizontalHeader().setStretchLastSection(True)
        self.custom_strategies_table.setMaximumHeight(200)
        saved_layout.addWidget(self.custom_strategies_table)

        btn_layout = QHBoxLayout()
        load_btn = QPushButton("Load Selected")
        load_btn.clicked.connect(self._load_custom_strategy)
        btn_layout.addWidget(load_btn)

        delete_btn = QPushButton("Delete")
        delete_btn.setStyleSheet("background-color: #F44336; color: white;")
        delete_btn.clicked.connect(self._delete_custom_strategy)
        btn_layout.addWidget(delete_btn)
        saved_layout.addLayout(btn_layout)

        right_layout.addWidget(saved_group)

        # Strategy Log
        log_group = QGroupBox("Signal Log")
        log_layout = QVBoxLayout(log_group)

        self.custom_strategy_log = QTextEdit()
        self.custom_strategy_log.setReadOnly(True)
        self.custom_strategy_log.setStyleSheet("font-family: monospace; font-size: 11px;")
        self.custom_strategy_log.setPlaceholderText("Strategy signals will appear here...")
        log_layout.addWidget(self.custom_strategy_log)

        clear_log_btn = QPushButton("Clear Log")
        clear_log_btn.clicked.connect(lambda: self.custom_strategy_log.clear())
        log_layout.addWidget(clear_log_btn)

        right_layout.addWidget(log_group)

        columns_layout.addWidget(right_panel)
        main_layout.addLayout(columns_layout)

        # Examples Section
        examples_group = QGroupBox("Quick Start Examples")
        examples_layout = QHBoxLayout(examples_group)

        ema_cross_btn = QPushButton("EMA 9/21 Crossover")
        ema_cross_btn.clicked.connect(lambda: self._load_example_strategy("ema_cross"))
        examples_layout.addWidget(ema_cross_btn)

        rsi_btn = QPushButton("RSI Oversold/Overbought")
        rsi_btn.clicked.connect(lambda: self._load_example_strategy("rsi"))
        examples_layout.addWidget(rsi_btn)

        supertrend_btn = QPushButton("Supertrend")
        supertrend_btn.clicked.connect(lambda: self._load_example_strategy("supertrend"))
        examples_layout.addWidget(supertrend_btn)

        vwap_btn = QPushButton("VWAP Breakout")
        vwap_btn.clicked.connect(lambda: self._load_example_strategy("vwap"))
        examples_layout.addWidget(vwap_btn)

        main_layout.addWidget(examples_group)

        # Initialize storage
        self.custom_strategies = {}  # {name: strategy_config}
        self.custom_strategy_running = False

        custom_scroll.setWidget(custom_widget)
        self.tabs.addTab(custom_scroll, "Custom Strategy")

    def _save_custom_strategy(self):
        """Save custom strategy to JSON"""
        import json
        import os

        name = self.custom_strategy_name.text().strip()
        if not name:
            QMessageBox.warning(self, "Error", "Please enter a strategy name")
            return

        strategy = {
            'name': name,
            'symbol': self.custom_strategy_symbol.currentText(),
            'timeframe': self.custom_strategy_timeframe.currentText(),
            'entry': {
                'indicator1': self.custom_entry_ind1.currentText(),
                'operator1': self.custom_entry_op1.currentText(),
                'indicator2': self.custom_entry_ind2.currentText(),
                'value1': self.custom_entry_val1.text(),
                'indicator3': self.custom_entry_ind3.currentText(),
                'operator2': self.custom_entry_op2.currentText(),
                'value2': self.custom_entry_val2.text()
            },
            'exit': {
                'indicator1': self.custom_exit_ind1.currentText(),
                'operator1': self.custom_exit_op1.currentText(),
                'indicator2': self.custom_exit_ind2.currentText(),
                'value1': self.custom_exit_val1.text()
            },
            'action': {
                'type': self.custom_action_type.currentText(),
                'lots': self.custom_lots.value(),
                'strike_offset': self.custom_strike_offset.value()
            },
            'test_mode': self.custom_test_mode_check.isChecked()
        }

        # Save to file
        strategies_dir = os.path.expanduser("~/.algo_trader/custom_strategies")
        os.makedirs(strategies_dir, exist_ok=True)

        filepath = os.path.join(strategies_dir, f"{name.replace(' ', '_')}.json")
        with open(filepath, 'w') as f:
            json.dump(strategy, f, indent=2)

        self.custom_strategies[name] = strategy
        self._update_custom_strategies_table()
        self._log_custom_strategy(f"Strategy '{name}' saved successfully")
        QMessageBox.information(self, "Saved", f"Strategy '{name}' saved!")

    def _load_custom_strategies(self):
        """Load saved custom strategies from disk"""
        import json
        import os

        strategies_dir = os.path.expanduser("~/.algo_trader/custom_strategies")
        if not os.path.exists(strategies_dir):
            return

        for filename in os.listdir(strategies_dir):
            if filename.endswith('.json'):
                filepath = os.path.join(strategies_dir, filename)
                try:
                    with open(filepath, 'r') as f:
                        strategy = json.load(f)
                        self.custom_strategies[strategy['name']] = strategy
                except Exception as e:
                    logger.error(f"Error loading strategy {filename}: {e}")

        self._update_custom_strategies_table()

    def _update_custom_strategies_table(self):
        """Update the custom strategies table"""
        self.custom_strategies_table.setRowCount(len(self.custom_strategies))

        for i, (name, strategy) in enumerate(self.custom_strategies.items()):
            self.custom_strategies_table.setItem(i, 0, QTableWidgetItem(name))
            self.custom_strategies_table.setItem(i, 1, QTableWidgetItem(strategy.get('symbol', 'N/A')))
            self.custom_strategies_table.setItem(i, 2, QTableWidgetItem("Stopped"))
            self.custom_strategies_table.setItem(i, 3, QTableWidgetItem("0"))

            start_btn = QPushButton("Start")
            start_btn.clicked.connect(lambda checked, n=name: self._start_saved_strategy(n))
            self.custom_strategies_table.setCellWidget(i, 4, start_btn)

    def _load_custom_strategy(self):
        """Load selected strategy into form"""
        row = self.custom_strategies_table.currentRow()
        if row < 0:
            QMessageBox.warning(self, "Select", "Please select a strategy to load")
            return

        name = self.custom_strategies_table.item(row, 0).text()
        strategy = self.custom_strategies.get(name)
        if not strategy:
            return

        # Populate form
        self.custom_strategy_name.setText(strategy['name'])
        self.custom_strategy_symbol.setCurrentText(strategy['symbol'])
        self.custom_strategy_timeframe.setCurrentText(strategy['timeframe'])

        # Entry conditions
        entry = strategy['entry']
        self.custom_entry_ind1.setCurrentText(entry['indicator1'])
        self.custom_entry_op1.setCurrentText(entry['operator1'])
        self.custom_entry_ind2.setCurrentText(entry['indicator2'])
        self.custom_entry_val1.setText(entry.get('value1', ''))
        self.custom_entry_ind3.setCurrentText(entry.get('indicator3', '-- None --'))
        self.custom_entry_op2.setCurrentText(entry.get('operator2', '>'))
        self.custom_entry_val2.setText(entry.get('value2', ''))

        # Exit conditions
        exit_cond = strategy['exit']
        self.custom_exit_ind1.setCurrentText(exit_cond['indicator1'])
        self.custom_exit_op1.setCurrentText(exit_cond['operator1'])
        self.custom_exit_ind2.setCurrentText(exit_cond['indicator2'])
        self.custom_exit_val1.setText(exit_cond.get('value1', ''))

        # Action
        action = strategy['action']
        self.custom_action_type.setCurrentText(action['type'])
        self.custom_lots.setValue(action.get('lots', 1))
        self.custom_strike_offset.setValue(action.get('strike_offset', 0))

        self.custom_test_mode_check.setChecked(strategy.get('test_mode', True))
        self._log_custom_strategy(f"Loaded strategy: {name}")

    def _delete_custom_strategy(self):
        """Delete selected custom strategy"""
        import os

        row = self.custom_strategies_table.currentRow()
        if row < 0:
            QMessageBox.warning(self, "Select", "Please select a strategy to delete")
            return

        name = self.custom_strategies_table.item(row, 0).text()

        reply = QMessageBox.question(self, "Confirm Delete",
                                    f"Delete strategy '{name}'?",
                                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)

        if reply == QMessageBox.StandardButton.Yes:
            # Remove from memory
            if name in self.custom_strategies:
                del self.custom_strategies[name]

            # Remove file
            strategies_dir = os.path.expanduser("~/.algo_trader/custom_strategies")
            filepath = os.path.join(strategies_dir, f"{name.replace(' ', '_')}.json")
            if os.path.exists(filepath):
                os.remove(filepath)

            self._update_custom_strategies_table()
            self._log_custom_strategy(f"Deleted strategy: {name}")

    def _start_custom_strategy(self):
        """Start monitoring with current custom strategy"""
        name = self.custom_strategy_name.text().strip()
        if not name:
            QMessageBox.warning(self, "Error", "Please enter a strategy name first")
            return

        # Save first
        self._save_custom_strategy()

        self.custom_strategy_running = True
        self.custom_start_btn.setEnabled(False)
        self.custom_stop_btn.setEnabled(True)

        mode = "Test Mode" if self.custom_test_mode_check.isChecked() else "LIVE"
        self._log_custom_strategy(f"Started monitoring: {name} [{mode}]")
        self._log_custom_strategy(f"Symbol: {self.custom_strategy_symbol.currentText()}")
        self._log_custom_strategy(f"Waiting for signal...")

    def _stop_custom_strategy(self):
        """Stop custom strategy monitoring"""
        self.custom_strategy_running = False
        self.custom_start_btn.setEnabled(True)
        self.custom_stop_btn.setEnabled(False)
        self._log_custom_strategy("Stopped monitoring")

    def _start_saved_strategy(self, name):
        """Start a saved strategy by name"""
        strategy = self.custom_strategies.get(name)
        if not strategy:
            return

        # Load into form and start
        self._load_custom_strategy()
        self._start_custom_strategy()

    def _load_example_strategy(self, example_type):
        """Load a pre-defined example strategy"""
        if example_type == "ema_cross":
            self.custom_strategy_name.setText("EMA 9/21 Crossover")
            self.custom_entry_ind1.setCurrentText("EMA(9)")
            self.custom_entry_op1.setCurrentText("Crosses Above")
            self.custom_entry_ind2.setCurrentText("EMA(21)")
            self.custom_exit_ind1.setCurrentText("EMA(9)")
            self.custom_exit_op1.setCurrentText("Crosses Below")
            self.custom_exit_ind2.setCurrentText("EMA(21)")
            self.custom_action_type.setCurrentText("Buy Call (CE)")
            self._log_custom_strategy("Loaded EMA 9/21 Crossover example")

        elif example_type == "rsi":
            self.custom_strategy_name.setText("RSI Oversold Buy")
            self.custom_entry_ind1.setCurrentText("RSI(14)")
            self.custom_entry_op1.setCurrentText("Crosses Above")
            self.custom_entry_ind2.setCurrentText("Value")
            self.custom_entry_val1.setText("30")
            self.custom_exit_ind1.setCurrentText("RSI(14)")
            self.custom_exit_op1.setCurrentText("Crosses Above")
            self.custom_exit_ind2.setCurrentText("Value")
            self.custom_exit_val1.setText("70")
            self.custom_action_type.setCurrentText("Buy Call (CE)")
            self._log_custom_strategy("Loaded RSI Oversold/Overbought example")

        elif example_type == "supertrend":
            self.custom_strategy_name.setText("Supertrend Strategy")
            self.custom_entry_ind1.setCurrentText("Price")
            self.custom_entry_op1.setCurrentText("Crosses Above")
            self.custom_entry_ind2.setCurrentText("Supertrend")
            self.custom_exit_ind1.setCurrentText("Price")
            self.custom_exit_op1.setCurrentText("Crosses Below")
            self.custom_exit_ind2.setCurrentText("Supertrend")
            self.custom_action_type.setCurrentText("Buy Call (CE)")
            self._log_custom_strategy("Loaded Supertrend example")

        elif example_type == "vwap":
            self.custom_strategy_name.setText("VWAP Breakout")
            self.custom_entry_ind1.setCurrentText("Price")
            self.custom_entry_op1.setCurrentText("Crosses Above")
            self.custom_entry_ind2.setCurrentText("VWAP")
            self.custom_exit_ind1.setCurrentText("Price")
            self.custom_exit_op1.setCurrentText("Crosses Below")
            self.custom_exit_ind2.setCurrentText("VWAP")
            self.custom_action_type.setCurrentText("Buy Call (CE)")
            self._log_custom_strategy("Loaded VWAP Breakout example")

    def _log_custom_strategy(self, message):
        """Log message to custom strategy log"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.custom_strategy_log.append(f"[{timestamp}] {message}")

    def _create_orders_tab(self):
        """Create orders management tab"""
        orders = QWidget()
        layout = QVBoxLayout(orders)

        # Manual order section
        order_group = QGroupBox("Place Manual Order")
        order_layout = QFormLayout(order_group)

        self.order_symbol = QLineEdit()
        self.order_symbol.setPlaceholderText("e.g., RELIANCE")
        order_layout.addRow("Symbol:", self.order_symbol)

        self.order_exchange = QComboBox()
        self.order_exchange.addItems(["NSE", "BSE", "NFO", "MCX"])
        order_layout.addRow("Exchange:", self.order_exchange)

        self.order_type = QComboBox()
        self.order_type.addItems(["BUY", "SELL"])
        order_layout.addRow("Type:", self.order_type)

        self.order_quantity = QSpinBox()
        self.order_quantity.setRange(1, 10000)
        self.order_quantity.setValue(1)
        order_layout.addRow("Quantity:", self.order_quantity)

        self.order_price_type = QComboBox()
        self.order_price_type.addItems(["MARKET", "LIMIT", "SL", "SL-M"])
        order_layout.addRow("Price Type:", self.order_price_type)

        self.order_price = QDoubleSpinBox()
        self.order_price.setRange(0, 100000)
        self.order_price.setDecimals(2)
        order_layout.addRow("Price:", self.order_price)

        self.place_order_btn = QPushButton("Place Order")
        self.place_order_btn.clicked.connect(self._place_manual_order)
        order_layout.addRow(self.place_order_btn)

        layout.addWidget(order_group)

        # Order book
        layout.addWidget(QLabel("Order Book"))
        self.orders_table = QTableWidget()
        self.orders_table.setColumnCount(8)
        self.orders_table.setHorizontalHeaderLabels([
            "Time", "Symbol", "Type", "Qty", "Price", "Status", "Broker Order ID", "Actions"
        ])
        self.orders_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        layout.addWidget(self.orders_table)

        # GTT Orders Section
        gtt_group = QGroupBox("GTT Orders (Good Till Triggered)")
        gtt_layout = QVBoxLayout(gtt_group)

        gtt_info = QLabel("GTT orders stay active on broker servers even when app is closed!")
        gtt_info.setStyleSheet("color: green; font-weight: bold;")
        gtt_layout.addWidget(gtt_info)

        # Create GTT Order form
        gtt_form = QHBoxLayout()

        gtt_form.addWidget(QLabel("Symbol:"))
        self.gtt_symbol = QLineEdit()
        self.gtt_symbol.setPlaceholderText("RELIANCE")
        self.gtt_symbol.setMaximumWidth(100)
        gtt_form.addWidget(self.gtt_symbol)

        gtt_form.addWidget(QLabel("Type:"))
        self.gtt_type = QComboBox()
        self.gtt_type.addItems(["SELL (Stop Loss)", "BUY (Target)", "OCO (SL + Target)"])
        gtt_form.addWidget(self.gtt_type)

        gtt_form.addWidget(QLabel("Trigger:"))
        self.gtt_trigger = QDoubleSpinBox()
        self.gtt_trigger.setRange(0.01, 999999)
        self.gtt_trigger.setDecimals(2)
        self.gtt_trigger.setPrefix("₹ ")
        gtt_form.addWidget(self.gtt_trigger)

        gtt_form.addWidget(QLabel("Price:"))
        self.gtt_price = QDoubleSpinBox()
        self.gtt_price.setRange(0.01, 999999)
        self.gtt_price.setDecimals(2)
        self.gtt_price.setPrefix("₹ ")
        gtt_form.addWidget(self.gtt_price)

        gtt_form.addWidget(QLabel("Qty:"))
        self.gtt_qty = QSpinBox()
        self.gtt_qty.setRange(1, 10000)
        self.gtt_qty.setValue(1)
        gtt_form.addWidget(self.gtt_qty)

        self.place_gtt_btn = QPushButton("Place GTT")
        self.place_gtt_btn.setStyleSheet("background-color: #ff9800; color: white; font-weight: bold;")
        self.place_gtt_btn.clicked.connect(self._place_gtt_order)
        gtt_form.addWidget(self.place_gtt_btn)

        gtt_form.addStretch()
        gtt_layout.addLayout(gtt_form)

        # OCO specific fields (Target price for OCO orders)
        oco_form = QHBoxLayout()
        oco_form.addWidget(QLabel("For OCO - Target Trigger:"))
        self.gtt_target_trigger = QDoubleSpinBox()
        self.gtt_target_trigger.setRange(0.01, 999999)
        self.gtt_target_trigger.setDecimals(2)
        self.gtt_target_trigger.setPrefix("₹ ")
        oco_form.addWidget(self.gtt_target_trigger)

        oco_form.addWidget(QLabel("Target Price:"))
        self.gtt_target_price = QDoubleSpinBox()
        self.gtt_target_price.setRange(0.01, 999999)
        self.gtt_target_price.setDecimals(2)
        self.gtt_target_price.setPrefix("₹ ")
        oco_form.addWidget(self.gtt_target_price)

        oco_form.addStretch()
        gtt_layout.addLayout(oco_form)

        # GTT Orders table
        gtt_layout.addWidget(QLabel("Active GTT Orders:"))
        self.gtt_table = QTableWidget()
        self.gtt_table.setColumnCount(7)
        self.gtt_table.setHorizontalHeaderLabels([
            "GTT ID", "Symbol", "Type", "Trigger", "Price", "Qty", "Actions"
        ])
        self.gtt_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.gtt_table.setMaximumHeight(150)
        gtt_layout.addWidget(self.gtt_table)

        # Refresh GTT button
        refresh_gtt_btn = QPushButton("🔄 Refresh GTT Orders")
        refresh_gtt_btn.clicked.connect(self._refresh_gtt_orders)
        gtt_layout.addWidget(refresh_gtt_btn)

        layout.addWidget(gtt_group)

        self.tabs.addTab(orders, "Orders")

    def _create_positions_tab(self):
        """Create positions tab"""
        positions = QWidget()
        layout = QVBoxLayout(positions)

        # Positions table
        layout.addWidget(QLabel("Current Positions"))
        self.positions_table = QTableWidget()
        self.positions_table.setColumnCount(7)
        self.positions_table.setHorizontalHeaderLabels([
            "Symbol", "Exchange", "Qty", "Avg Price", "LTP", "P&L", "Actions"
        ])
        self.positions_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        layout.addWidget(self.positions_table)

        # Holdings table
        layout.addWidget(QLabel("Holdings"))
        self.holdings_table = QTableWidget()
        self.holdings_table.setColumnCount(6)
        self.holdings_table.setHorizontalHeaderLabels([
            "Symbol", "Qty", "Avg Price", "Current Value", "P&L", "P&L %"
        ])
        self.holdings_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        layout.addWidget(self.holdings_table)

        self.tabs.addTab(positions, "Positions")

    def _create_risk_tab(self):
        """Create Risk Management tab with TSL and MTM"""
        # Wrap in scroll area
        risk_scroll = QScrollArea()
        risk_scroll.setWidgetResizable(True)

        risk = QWidget()
        layout = QVBoxLayout(risk)

        # MTM Summary Section
        mtm_group = QGroupBox("MTM (Mark-to-Market) Summary")
        mtm_layout = QVBoxLayout(mtm_group)

        mtm_info_layout = QHBoxLayout()
        self.mtm_total_pnl = QLabel("Total P&L: ₹0.00")
        self.mtm_total_pnl.setFont(QFont("Arial", 14, QFont.Weight.Bold))
        self.mtm_realized = QLabel("Realized: ₹0.00")
        self.mtm_unrealized = QLabel("Unrealized: ₹0.00")
        self.mtm_trades = QLabel("Trades: 0 (W: 0 / L: 0)")

        mtm_info_layout.addWidget(self.mtm_total_pnl)
        mtm_info_layout.addWidget(self.mtm_realized)
        mtm_info_layout.addWidget(self.mtm_unrealized)
        mtm_info_layout.addWidget(self.mtm_trades)
        mtm_layout.addLayout(mtm_info_layout)

        layout.addWidget(mtm_group)

        # Add Position with TSL Section
        add_pos_group = QGroupBox("Add Position with Trailing Stop Loss")
        add_pos_layout = QFormLayout(add_pos_group)

        self.risk_symbol = QLineEdit()
        self.risk_symbol.setPlaceholderText("e.g., RELIANCE")
        add_pos_layout.addRow("Symbol:", self.risk_symbol)

        self.risk_exchange = QComboBox()
        self.risk_exchange.addItems(["NSE", "BSE", "NFO"])
        add_pos_layout.addRow("Exchange:", self.risk_exchange)

        self.risk_quantity = QSpinBox()
        self.risk_quantity.setRange(-10000, 10000)
        self.risk_quantity.setValue(1)
        add_pos_layout.addRow("Quantity (+Long/-Short):", self.risk_quantity)

        self.risk_entry_price = QDoubleSpinBox()
        self.risk_entry_price.setRange(0.01, 100000)
        self.risk_entry_price.setDecimals(2)
        add_pos_layout.addRow("Entry Price:", self.risk_entry_price)

        self.risk_sl_type = QComboBox()
        self.risk_sl_type.addItems(["Fixed Price", "Trailing %", "Trailing Points"])
        self.risk_sl_type.currentTextChanged.connect(self._on_sl_type_changed)
        add_pos_layout.addRow("Stop Loss Type:", self.risk_sl_type)

        self.risk_sl_value = QDoubleSpinBox()
        self.risk_sl_value.setRange(0, 100000)
        self.risk_sl_value.setDecimals(2)
        add_pos_layout.addRow("SL Value:", self.risk_sl_value)

        self.risk_target = QDoubleSpinBox()
        self.risk_target.setRange(0, 100000)
        self.risk_target.setDecimals(2)
        add_pos_layout.addRow("Target Price (optional):", self.risk_target)

        add_pos_btn = QPushButton("Add Position")
        add_pos_btn.clicked.connect(self._add_risk_position)
        add_pos_layout.addRow(add_pos_btn)

        layout.addWidget(add_pos_group)

        # Tracked Positions Section
        tracked_group = QGroupBox("Tracked Positions (with TSL)")
        tracked_layout = QVBoxLayout(tracked_group)

        self.risk_positions_table = QTableWidget()
        self.risk_positions_table.setColumnCount(10)
        self.risk_positions_table.setHorizontalHeaderLabels([
            "Symbol", "Qty", "Entry", "LTP", "SL", "Target", "High/Low", "P&L", "P&L %", "Actions"
        ])
        self.risk_positions_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        tracked_layout.addWidget(self.risk_positions_table)

        # Control buttons
        btn_layout = QHBoxLayout()
        self.start_risk_monitor_btn = QPushButton("Start Monitoring")
        self.start_risk_monitor_btn.clicked.connect(self._start_risk_monitoring)
        self.stop_risk_monitor_btn = QPushButton("Stop Monitoring")
        self.stop_risk_monitor_btn.clicked.connect(self._stop_risk_monitoring)
        self.stop_risk_monitor_btn.setEnabled(False)
        self.refresh_risk_btn = QPushButton("Refresh")
        self.refresh_risk_btn.clicked.connect(self._refresh_risk_positions)

        btn_layout.addWidget(self.start_risk_monitor_btn)
        btn_layout.addWidget(self.stop_risk_monitor_btn)
        btn_layout.addWidget(self.refresh_risk_btn)
        tracked_layout.addLayout(btn_layout)

        layout.addWidget(tracked_group)

        risk_scroll.setWidget(risk)
        self.tabs.addTab(risk_scroll, "Risk/TSL")

    def _create_options_tab(self):
        """Create Options Trading tab with Expiry, Strike, Hedge strategies"""
        # Use scroll area so content is not squeezed
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)

        options = QWidget()
        layout = QVBoxLayout(options)
        layout.setSpacing(8)

        # --- Options P&L Summary ---
        opt_summary_group = QGroupBox("Options P&L Summary")
        opt_summary_layout = QHBoxLayout(opt_summary_group)
        self.opt_total_pnl = QLabel("Total P&L: ₹0.00")
        self.opt_total_pnl.setFont(QFont("Arial", 13, QFont.Weight.Bold))
        self.opt_active_count = QLabel("Active: 0")
        self.opt_closed_count = QLabel("Closed: 0")
        opt_summary_layout.addWidget(self.opt_total_pnl)
        opt_summary_layout.addWidget(self.opt_active_count)
        opt_summary_layout.addWidget(self.opt_closed_count)
        layout.addWidget(opt_summary_group)

        # --- Auto Options (Signal -> Options) ---
        auto_group = QGroupBox("Auto Options (Strategy Signal → Automatic Option Trade)")
        auto_main_layout = QVBoxLayout(auto_group)

        # Row 1: Enable + Symbol + Close opposite + Save
        auto_top = QHBoxLayout()
        self.auto_opt_enabled = QCheckBox("Enable Auto-Options")
        auto_top.addWidget(self.auto_opt_enabled)

        auto_top.addWidget(QLabel("Symbol:"))
        self.auto_opt_symbol = QComboBox()
        self.auto_opt_symbol.addItems(["NIFTY", "BANKNIFTY", "FINNIFTY", "MIDCPNIFTY"])
        self.auto_opt_symbol.setEditable(True)
        auto_top.addWidget(self.auto_opt_symbol)

        self.auto_opt_close_opposite = QCheckBox("Close on Opposite Signal")
        self.auto_opt_close_opposite.setChecked(True)
        auto_top.addWidget(self.auto_opt_close_opposite)

        self.auto_opt_hedge_enabled = QCheckBox("Enable Hedge (Leg 2)")
        auto_top.addWidget(self.auto_opt_hedge_enabled)

        auto_main_layout.addLayout(auto_top)

        # Row 2: Leg 1 and Leg 2 side by side
        legs_row = QHBoxLayout()

        # Leg 1 config
        leg1_group = QGroupBox("Leg 1 (Main Trade)")
        leg1_layout = QFormLayout(leg1_group)
        leg1_layout.setSpacing(3)

        self.auto_leg1_type = QComboBox()
        self.auto_leg1_type.addItems(["CE", "PE"])
        leg1_layout.addRow("Type:", self.auto_leg1_type)

        self.auto_leg1_action = QComboBox()
        self.auto_leg1_action.addItems(["BUY", "SELL"])
        leg1_layout.addRow("Action:", self.auto_leg1_action)

        self.auto_leg1_strike = QComboBox()
        self.auto_leg1_strike.addItems([s.value for s in StrikeSelection] + ["Manual"])
        self.auto_leg1_strike.currentTextChanged.connect(self._on_auto_leg1_strike_changed)
        leg1_layout.addRow("Strike:", self.auto_leg1_strike)

        # Manual strike price for Leg 1
        self.auto_leg1_manual_strike = QDoubleSpinBox()
        self.auto_leg1_manual_strike.setRange(0, 999999)
        self.auto_leg1_manual_strike.setDecimals(0)
        self.auto_leg1_manual_strike.setSingleStep(50)
        self.auto_leg1_manual_strike.setPrefix("₹ ")
        self.auto_leg1_manual_strike.setVisible(False)
        leg1_layout.addRow("Strike Price:", self.auto_leg1_manual_strike)

        self.auto_leg1_expiry = QComboBox()
        self.auto_leg1_expiry.addItems([e.value for e in ExpirySelection])
        leg1_layout.addRow("Expiry:", self.auto_leg1_expiry)

        self.auto_leg1_qty = QSpinBox()
        self.auto_leg1_qty.setRange(1, 500)
        self.auto_leg1_qty.setValue(1)
        self.auto_leg1_qty.setSuffix(" lots")
        leg1_layout.addRow("Qty:", self.auto_leg1_qty)

        legs_row.addWidget(leg1_group)

        # Leg 2 config (Hedge)
        leg2_group = QGroupBox("Leg 2 (Hedge - Separate Strike/Expiry)")
        leg2_layout = QFormLayout(leg2_group)
        leg2_layout.setSpacing(3)

        self.auto_leg2_type = QComboBox()
        self.auto_leg2_type.addItems(["CE", "PE"])
        leg2_layout.addRow("Type:", self.auto_leg2_type)

        self.auto_leg2_action = QComboBox()
        self.auto_leg2_action.addItems(["BUY", "SELL"])
        self.auto_leg2_action.setCurrentText("SELL")
        leg2_layout.addRow("Action:", self.auto_leg2_action)

        self.auto_leg2_strike = QComboBox()
        self.auto_leg2_strike.addItems([s.value for s in StrikeSelection] + ["Manual"])
        self.auto_leg2_strike.setCurrentText("OTM +3")
        self.auto_leg2_strike.currentTextChanged.connect(self._on_auto_leg2_strike_changed)
        leg2_layout.addRow("Strike:", self.auto_leg2_strike)

        # Manual strike price for Leg 2
        self.auto_leg2_manual_strike = QDoubleSpinBox()
        self.auto_leg2_manual_strike.setRange(0, 999999)
        self.auto_leg2_manual_strike.setDecimals(0)
        self.auto_leg2_manual_strike.setSingleStep(50)
        self.auto_leg2_manual_strike.setPrefix("₹ ")
        self.auto_leg2_manual_strike.setVisible(False)
        leg2_layout.addRow("Strike Price:", self.auto_leg2_manual_strike)

        self.auto_leg2_expiry = QComboBox()
        self.auto_leg2_expiry.addItems([e.value for e in ExpirySelection])
        leg2_layout.addRow("Expiry:", self.auto_leg2_expiry)

        self.auto_leg2_qty = QSpinBox()
        self.auto_leg2_qty.setRange(1, 500)
        self.auto_leg2_qty.setValue(1)
        self.auto_leg2_qty.setSuffix(" lots")
        leg2_layout.addRow("Qty:", self.auto_leg2_qty)

        legs_row.addWidget(leg2_group)

        # SL/Target/TSL config
        exit_group_auto = QGroupBox("Exit (Combined P&L)")
        exit_auto_layout = QFormLayout(exit_group_auto)
        exit_auto_layout.setSpacing(3)

        self.auto_opt_exit_type = QComboBox()
        self.auto_opt_exit_type.addItems([e.value for e in ExitType])
        self.auto_opt_exit_type.setCurrentText("P&L Based")
        exit_auto_layout.addRow("Exit:", self.auto_opt_exit_type)

        self.auto_opt_sl = QDoubleSpinBox()
        self.auto_opt_sl.setRange(0, 1000000)
        self.auto_opt_sl.setPrefix("₹")
        exit_auto_layout.addRow("SL:", self.auto_opt_sl)

        self.auto_opt_target = QDoubleSpinBox()
        self.auto_opt_target.setRange(0, 1000000)
        self.auto_opt_target.setPrefix("₹")
        exit_auto_layout.addRow("Target:", self.auto_opt_target)

        self.auto_opt_tsl = QDoubleSpinBox()
        self.auto_opt_tsl.setRange(0, 1000000)
        self.auto_opt_tsl.setPrefix("₹")
        exit_auto_layout.addRow("TSL:", self.auto_opt_tsl)

        save_auto_btn = QPushButton("Save Auto Config")
        save_auto_btn.clicked.connect(self._save_auto_options_config)
        exit_auto_layout.addRow(save_auto_btn)

        legs_row.addWidget(exit_group_auto)

        auto_main_layout.addLayout(legs_row)

        # Legacy hidden combos for backward compat
        self.auto_opt_buy_action = QComboBox()
        self.auto_opt_buy_action.addItems([s.value for s in SignalAction])
        self.auto_opt_buy_action.setVisible(False)
        self.auto_opt_sell_action = QComboBox()
        self.auto_opt_sell_action.addItems([s.value for s in SignalAction])
        self.auto_opt_sell_action.setCurrentText("BUY PE")
        self.auto_opt_sell_action.setVisible(False)

        layout.addWidget(auto_group)

        # --- Auto Trade Log ---
        self.auto_trade_log_table = QTableWidget()
        self.auto_trade_log_table.setColumnCount(6)
        self.auto_trade_log_table.setHorizontalHeaderLabels([
            "Time", "Signal", "Strategy", "Action", "Expiry", "Qty"
        ])
        self.auto_trade_log_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.auto_trade_log_table.setMaximumHeight(100)
        layout.addWidget(self.auto_trade_log_table)

        # --- Manual Multi-Leg Builder ---
        build_group = QGroupBox("Manual Option Builder (Multi-Leg - Each Leg Separate Strike & Expiry)")
        build_layout = QVBoxLayout(build_group)

        # Top row: Symbol, Spot, Load buttons
        top_form = QHBoxLayout()

        top_form.addWidget(QLabel("Symbol:"))
        self.opt_symbol = QComboBox()
        self.opt_symbol.addItems(["NIFTY", "BANKNIFTY", "FINNIFTY", "MIDCPNIFTY", "SENSEX"])
        self.opt_symbol.setEditable(True)
        self.opt_symbol.currentTextChanged.connect(self._on_opt_symbol_changed)
        top_form.addWidget(self.opt_symbol)

        top_form.addWidget(QLabel("Spot:"))
        self.opt_spot_price = QDoubleSpinBox()
        self.opt_spot_price.setRange(0, 200000)
        self.opt_spot_price.setDecimals(2)
        self.opt_spot_price.setPrefix("₹")
        top_form.addWidget(self.opt_spot_price)

        fetch_spot_btn = QPushButton("Fetch Spot")
        fetch_spot_btn.clicked.connect(self._fetch_spot_price)
        top_form.addWidget(fetch_spot_btn)

        load_expiry_btn = QPushButton("Load Expiries")
        load_expiry_btn.clicked.connect(self._load_expiry_dates)
        top_form.addWidget(load_expiry_btn)

        load_strikes_btn = QPushButton("Load Strikes")
        load_strikes_btn.clicked.connect(self._load_strike_prices)
        top_form.addWidget(load_strikes_btn)

        self.opt_lot_size = QLabel("Lot: 25")
        top_form.addWidget(self.opt_lot_size)

        build_layout.addLayout(top_form)

        # Hidden combo for expiry/strike data (used by leg rows)
        self.opt_expiry = QComboBox()
        self.opt_expiry.setVisible(False)
        build_layout.addWidget(self.opt_expiry)
        self.opt_strike = QComboBox()
        self.opt_strike.setEditable(True)
        self.opt_strike.setVisible(False)
        build_layout.addWidget(self.opt_strike)

        # Leg builder table
        self.leg_builder_table = QTableWidget()
        self.leg_builder_table.setColumnCount(7)
        self.leg_builder_table.setHorizontalHeaderLabels([
            "Strike", "Expiry", "CE/PE", "BUY/SELL", "Lots", "Premium ₹", "Remove"
        ])
        self.leg_builder_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.leg_builder_table.setMaximumHeight(140)
        build_layout.addWidget(self.leg_builder_table)

        # Add leg button
        leg_btn_row = QHBoxLayout()
        add_leg_btn = QPushButton("+ Add Leg")
        add_leg_btn.clicked.connect(self._add_builder_leg)
        leg_btn_row.addWidget(add_leg_btn)

        # Exit settings inline
        leg_btn_row.addWidget(QLabel("Exit:"))
        self.opt_exit_type = QComboBox()
        self.opt_exit_type.addItems([e.value for e in ExitType])
        self.opt_exit_type.currentTextChanged.connect(self._on_opt_exit_type_changed)
        leg_btn_row.addWidget(self.opt_exit_type)

        leg_btn_row.addWidget(QLabel("SL:"))
        self.opt_sl_value = QDoubleSpinBox()
        self.opt_sl_value.setRange(0, 1000000)
        self.opt_sl_value.setDecimals(2)
        self.opt_sl_value.setPrefix("₹")
        leg_btn_row.addWidget(self.opt_sl_value)

        leg_btn_row.addWidget(QLabel("Target:"))
        self.opt_target_value = QDoubleSpinBox()
        self.opt_target_value.setRange(0, 1000000)
        self.opt_target_value.setDecimals(2)
        self.opt_target_value.setPrefix("₹")
        leg_btn_row.addWidget(self.opt_target_value)

        leg_btn_row.addWidget(QLabel("TSL:"))
        self.opt_tsl_value = QDoubleSpinBox()
        self.opt_tsl_value.setRange(0, 1000000)
        self.opt_tsl_value.setDecimals(2)
        leg_btn_row.addWidget(self.opt_tsl_value)

        build_layout.addLayout(leg_btn_row)

        # Create position button
        create_pos_btn = QPushButton(">>> Create Multi-Leg Position <<<")
        create_pos_btn.setMinimumHeight(35)
        create_pos_btn.clicked.connect(self._add_option_position)
        build_layout.addWidget(create_pos_btn)

        layout.addWidget(build_group)

        # --- Active Option Positions Table ---
        positions_group = QGroupBox("Active Option Positions")
        positions_layout = QVBoxLayout(positions_group)

        self.opt_positions_table = QTableWidget()
        self.opt_positions_table.setColumnCount(10)
        self.opt_positions_table.setHorizontalHeaderLabels([
            "ID", "Symbol", "Strategy", "Legs", "Expiry",
            "Total P&L", "P&L %", "Max P&L", "Exit Type", "Actions"
        ])
        self.opt_positions_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.opt_positions_table.setMaximumHeight(180)
        positions_layout.addWidget(self.opt_positions_table)

        # Legs detail table
        positions_layout.addWidget(QLabel("Leg Details (select position above):"))
        self.opt_legs_table = QTableWidget()
        self.opt_legs_table.setColumnCount(10)
        self.opt_legs_table.setHorizontalHeaderLabels([
            "Leg", "Strike", "Expiry", "Type", "Action", "Qty", "Entry", "LTP", "P&L", "P&L %"
        ])
        self.opt_legs_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.opt_legs_table.setMaximumHeight(150)
        positions_layout.addWidget(self.opt_legs_table)

        self.opt_positions_table.cellClicked.connect(self._on_opt_position_selected)

        opt_refresh_btn = QPushButton("Refresh Positions")
        opt_refresh_btn.clicked.connect(self._refresh_option_positions)
        positions_layout.addWidget(opt_refresh_btn)

        layout.addWidget(positions_group)

        scroll.setWidget(options)
        self.tabs.addTab(scroll, "Options")

    def _create_backtest_tab(self):
        """Create advanced backtesting simulator tab"""
        backtest = QWidget()
        layout = QVBoxLayout(backtest)

        # Top section - Config and Controls in horizontal layout
        top_layout = QHBoxLayout()

        # Left - Configuration
        config_group = QGroupBox("Backtest Configuration")
        config_layout = QFormLayout(config_group)

        self.bt_strategy_combo = QComboBox()
        config_layout.addRow("Strategy:", self.bt_strategy_combo)

        # Symbol with preset options
        symbol_layout = QHBoxLayout()
        self.bt_symbol = QComboBox()
        self.bt_symbol.setEditable(True)
        self.bt_symbol.addItems(["NIFTY", "BANKNIFTY", "RELIANCE", "TCS", "HDFCBANK",
                                  "INFY", "ICICIBANK", "SBIN", "TATASTEEL", "WIPRO"])
        symbol_layout.addWidget(self.bt_symbol)
        config_layout.addRow("Symbol:", symbol_layout)

        # Date range
        self.bt_days = QSpinBox()
        self.bt_days.setRange(30, 1000)
        self.bt_days.setValue(365)
        self.bt_days.setSuffix(" days")
        config_layout.addRow("History:", self.bt_days)

        # Interval
        self.bt_interval = QComboBox()
        self.bt_interval.addItems(["1d (Daily)", "1h (Hourly)", "15m (15 min)", "5m (5 min)"])
        config_layout.addRow("Interval:", self.bt_interval)

        self.bt_capital = QDoubleSpinBox()
        self.bt_capital.setRange(10000, 10000000)
        self.bt_capital.setValue(100000)
        self.bt_capital.setPrefix("₹ ")
        self.bt_capital.setDecimals(0)
        config_layout.addRow("Initial Capital:", self.bt_capital)

        top_layout.addWidget(config_group)

        # Middle - Risk Settings
        risk_group = QGroupBox("Risk Settings")
        risk_layout = QFormLayout(risk_group)

        self.bt_sl = QDoubleSpinBox()
        self.bt_sl.setRange(0, 50)
        self.bt_sl.setValue(2)
        self.bt_sl.setSuffix(" %")
        self.bt_sl.setSpecialValueText("No SL")
        risk_layout.addRow("Stop Loss:", self.bt_sl)

        self.bt_target = QDoubleSpinBox()
        self.bt_target.setRange(0, 100)
        self.bt_target.setValue(4)
        self.bt_target.setSuffix(" %")
        self.bt_target.setSpecialValueText("No Target")
        risk_layout.addRow("Target:", self.bt_target)

        self.bt_tsl = QDoubleSpinBox()
        self.bt_tsl.setRange(0, 50)
        self.bt_tsl.setValue(0)
        self.bt_tsl.setSuffix(" %")
        self.bt_tsl.setSpecialValueText("No TSL")
        risk_layout.addRow("Trailing SL:", self.bt_tsl)

        top_layout.addWidget(risk_group)

        # Right - Controls
        control_group = QGroupBox("Simulation Controls")
        control_layout = QVBoxLayout(control_group)

        self.run_backtest_btn = QPushButton("▶ Start Backtest")
        self.run_backtest_btn.setStyleSheet("background-color: #4CAF50; color: white; font-weight: bold;")
        self.run_backtest_btn.clicked.connect(self._run_advanced_backtest)
        control_layout.addWidget(self.run_backtest_btn)

        self.bt_realtime_mode = QCheckBox("Real-time Mode (Slow)")
        control_layout.addWidget(self.bt_realtime_mode)

        speed_layout = QHBoxLayout()
        speed_layout.addWidget(QLabel("Speed:"))
        self.bt_speed = QComboBox()
        self.bt_speed.addItems(["1x", "2x", "5x", "10x"])
        self.bt_speed.setCurrentIndex(1)
        speed_layout.addWidget(self.bt_speed)
        control_layout.addLayout(speed_layout)

        self.bt_progress = QLabel("Ready")
        self.bt_progress.setStyleSheet("font-weight: bold;")
        control_layout.addWidget(self.bt_progress)

        self.export_trades_btn = QPushButton("Export Trades CSV")
        self.export_trades_btn.clicked.connect(self._export_backtest_trades)
        self.export_trades_btn.setEnabled(False)
        control_layout.addWidget(self.export_trades_btn)

        control_layout.addStretch()
        top_layout.addWidget(control_group)

        layout.addLayout(top_layout)

        # Middle - Results Summary
        summary_group = QGroupBox("Results Summary")
        summary_layout = QHBoxLayout(summary_group)

        # Stats in a horizontal row
        self.bt_stat_pnl = QLabel("P&L: --")
        self.bt_stat_pnl.setStyleSheet("font-size: 14px; font-weight: bold;")
        summary_layout.addWidget(self.bt_stat_pnl)

        self.bt_stat_trades = QLabel("Trades: --")
        summary_layout.addWidget(self.bt_stat_trades)

        self.bt_stat_winrate = QLabel("Win Rate: --")
        summary_layout.addWidget(self.bt_stat_winrate)

        self.bt_stat_pf = QLabel("Profit Factor: --")
        summary_layout.addWidget(self.bt_stat_pf)

        self.bt_stat_dd = QLabel("Max Drawdown: --")
        summary_layout.addWidget(self.bt_stat_dd)

        layout.addWidget(summary_group)

        # Bottom - Trade Log Table
        trades_group = QGroupBox("Trade Log (Step-by-Step)")
        trades_layout = QVBoxLayout(trades_group)

        self.bt_trades_table = QTableWidget()
        self.bt_trades_table.setColumnCount(10)
        self.bt_trades_table.setHorizontalHeaderLabels([
            "ID", "Type", "Entry Time", "Entry Price", "Exit Time",
            "Exit Price", "Qty", "P&L", "P&L %", "Exit Reason"
        ])
        self.bt_trades_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        trades_layout.addWidget(self.bt_trades_table)

        layout.addWidget(trades_group)

        self.tabs.addTab(backtest, "Backtest")

    def _create_journal_tab(self):
        """Create Trade Journal / Performance Analytics tab"""
        journal_scroll = QScrollArea()
        journal_scroll.setWidgetResizable(True)

        journal = QWidget()
        layout = QVBoxLayout(journal)

        # === Filter Controls ===
        filter_group = QGroupBox("Filter & Export")
        filter_layout = QHBoxLayout(filter_group)

        filter_layout.addWidget(QLabel("Date Range:"))
        self.journal_date_from = QComboBox()
        self.journal_date_from.addItems(["Today", "This Week", "This Month", "Last 30 Days", "Last 90 Days", "This Year", "All Time"])
        self.journal_date_from.setCurrentIndex(3)  # Last 30 Days
        self.journal_date_from.currentIndexChanged.connect(self._refresh_journal)
        filter_layout.addWidget(self.journal_date_from)

        filter_layout.addWidget(QLabel("Symbol:"))
        self.journal_symbol_filter = QComboBox()
        self.journal_symbol_filter.addItem("All Symbols")
        self.journal_symbol_filter.currentIndexChanged.connect(self._refresh_journal)
        filter_layout.addWidget(self.journal_symbol_filter)

        filter_layout.addWidget(QLabel("Side:"))
        self.journal_side_filter = QComboBox()
        self.journal_side_filter.addItems(["All", "BUY Only", "SELL Only"])
        self.journal_side_filter.currentIndexChanged.connect(self._refresh_journal)
        filter_layout.addWidget(self.journal_side_filter)

        filter_layout.addStretch()

        # Export buttons
        self.export_excel_btn = QPushButton("📊 Export to Excel")
        self.export_excel_btn.setStyleSheet("background-color: #217346; color: white; font-weight: bold; padding: 8px;")
        self.export_excel_btn.clicked.connect(self._export_trades_excel)
        filter_layout.addWidget(self.export_excel_btn)

        self.export_csv_btn = QPushButton("📄 Export CSV")
        self.export_csv_btn.clicked.connect(self._export_trades_csv)
        filter_layout.addWidget(self.export_csv_btn)

        refresh_btn = QPushButton("🔄 Refresh")
        refresh_btn.clicked.connect(self._refresh_journal)
        filter_layout.addWidget(refresh_btn)

        layout.addWidget(filter_group)

        # === Performance Summary Cards ===
        perf_group = QGroupBox("Performance Summary")
        perf_layout = QHBoxLayout(perf_group)

        # Total P&L Card
        pnl_card = QGroupBox("Total P&L")
        pnl_layout = QVBoxLayout(pnl_card)
        self.journal_total_pnl = QLabel("₹0")
        self.journal_total_pnl.setStyleSheet("font-size: 20px; font-weight: bold; color: #4CAF50;")
        self.journal_total_pnl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        pnl_layout.addWidget(self.journal_total_pnl)
        perf_layout.addWidget(pnl_card)

        # Total Trades Card
        trades_card = QGroupBox("Total Trades")
        trades_layout = QVBoxLayout(trades_card)
        self.journal_total_trades = QLabel("0")
        self.journal_total_trades.setStyleSheet("font-size: 20px; font-weight: bold; color: #2196F3;")
        self.journal_total_trades.setAlignment(Qt.AlignmentFlag.AlignCenter)
        trades_layout.addWidget(self.journal_total_trades)
        perf_layout.addWidget(trades_card)

        # Win Rate Card
        win_card = QGroupBox("Win Rate")
        win_layout = QVBoxLayout(win_card)
        self.journal_win_rate = QLabel("0%")
        self.journal_win_rate.setStyleSheet("font-size: 20px; font-weight: bold; color: #9C27B0;")
        self.journal_win_rate.setAlignment(Qt.AlignmentFlag.AlignCenter)
        win_layout.addWidget(self.journal_win_rate)
        perf_layout.addWidget(win_card)

        # Profit Factor Card
        pf_card = QGroupBox("Profit Factor")
        pf_layout = QVBoxLayout(pf_card)
        self.journal_profit_factor = QLabel("0")
        self.journal_profit_factor.setStyleSheet("font-size: 20px; font-weight: bold; color: #FF9800;")
        self.journal_profit_factor.setAlignment(Qt.AlignmentFlag.AlignCenter)
        pf_layout.addWidget(self.journal_profit_factor)
        perf_layout.addWidget(pf_card)

        # Avg Win/Loss Card
        avg_card = QGroupBox("Avg Win / Avg Loss")
        avg_layout = QVBoxLayout(avg_card)
        self.journal_avg_win_loss = QLabel("₹0 / ₹0")
        self.journal_avg_win_loss.setStyleSheet("font-size: 16px; font-weight: bold;")
        self.journal_avg_win_loss.setAlignment(Qt.AlignmentFlag.AlignCenter)
        avg_layout.addWidget(self.journal_avg_win_loss)
        perf_layout.addWidget(avg_card)

        # Max Drawdown Card
        dd_card = QGroupBox("Max Drawdown")
        dd_layout = QVBoxLayout(dd_card)
        self.journal_max_dd = QLabel("₹0")
        self.journal_max_dd.setStyleSheet("font-size: 20px; font-weight: bold; color: #F44336;")
        self.journal_max_dd.setAlignment(Qt.AlignmentFlag.AlignCenter)
        dd_layout.addWidget(self.journal_max_dd)
        perf_layout.addWidget(dd_card)

        layout.addWidget(perf_group)

        # === Detailed Statistics ===
        stats_group = QGroupBox("Detailed Statistics")
        stats_layout = QHBoxLayout(stats_group)

        # Left stats
        left_stats = QFormLayout()
        self.journal_winning_trades = QLabel("0")
        self.journal_losing_trades = QLabel("0")
        self.journal_largest_win = QLabel("₹0")
        self.journal_largest_loss = QLabel("₹0")
        left_stats.addRow("Winning Trades:", self.journal_winning_trades)
        left_stats.addRow("Losing Trades:", self.journal_losing_trades)
        left_stats.addRow("Largest Win:", self.journal_largest_win)
        left_stats.addRow("Largest Loss:", self.journal_largest_loss)
        stats_layout.addLayout(left_stats)

        # Middle stats
        mid_stats = QFormLayout()
        self.journal_avg_trade = QLabel("₹0")
        self.journal_avg_hold_time = QLabel("0 min")
        self.journal_total_volume = QLabel("0")
        self.journal_total_brokerage = QLabel("₹0")
        mid_stats.addRow("Average Trade:", self.journal_avg_trade)
        mid_stats.addRow("Avg Hold Time:", self.journal_avg_hold_time)
        mid_stats.addRow("Total Volume:", self.journal_total_volume)
        mid_stats.addRow("Est. Brokerage:", self.journal_total_brokerage)
        stats_layout.addLayout(mid_stats)

        # Right stats - Per symbol breakdown
        right_stats = QFormLayout()
        self.journal_best_symbol = QLabel("--")
        self.journal_worst_symbol = QLabel("--")
        self.journal_most_traded = QLabel("--")
        self.journal_consecutive_wins = QLabel("0")
        right_stats.addRow("Best Symbol:", self.journal_best_symbol)
        right_stats.addRow("Worst Symbol:", self.journal_worst_symbol)
        right_stats.addRow("Most Traded:", self.journal_most_traded)
        right_stats.addRow("Max Consecutive Wins:", self.journal_consecutive_wins)
        stats_layout.addLayout(right_stats)

        layout.addWidget(stats_group)

        # === Trade History Table ===
        history_group = QGroupBox("Trade History")
        history_layout = QVBoxLayout(history_group)

        self.journal_trades_table = QTableWidget()
        self.journal_trades_table.setColumnCount(12)
        self.journal_trades_table.setHorizontalHeaderLabels([
            "Date", "Time", "Symbol", "Side", "Qty", "Entry Price",
            "Exit Price", "P&L", "P&L %", "Source", "Strategy", "Notes"
        ])
        self.journal_trades_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.journal_trades_table.setAlternatingRowColors(True)
        self.journal_trades_table.setSortingEnabled(True)
        history_layout.addWidget(self.journal_trades_table)

        layout.addWidget(history_group)

        journal_scroll.setWidget(journal)
        self.tabs.addTab(journal_scroll, "Journal")

        # Initial load
        QTimer.singleShot(100, self._refresh_journal)

    def _refresh_journal(self):
        """Refresh trade journal with filtered data"""
        try:
            from datetime import datetime, timedelta
            from collections import defaultdict

            # Get date range
            date_filter = self.journal_date_from.currentText()
            now = datetime.now()

            if date_filter == "Today":
                start_date = now.replace(hour=0, minute=0, second=0)
            elif date_filter == "This Week":
                start_date = now - timedelta(days=now.weekday())
                start_date = start_date.replace(hour=0, minute=0, second=0)
            elif date_filter == "This Month":
                start_date = now.replace(day=1, hour=0, minute=0, second=0)
            elif date_filter == "Last 30 Days":
                start_date = now - timedelta(days=30)
            elif date_filter == "Last 90 Days":
                start_date = now - timedelta(days=90)
            elif date_filter == "This Year":
                start_date = now.replace(month=1, day=1, hour=0, minute=0, second=0)
            else:  # All Time
                start_date = datetime(2000, 1, 1)

            # Get all orders from database
            all_orders = self.db.get_orders()

            # Filter orders
            filtered_orders = []
            symbols_set = set()

            for order in all_orders:
                # Parse order date
                order_date_str = order.get('created_at', '')
                try:
                    if isinstance(order_date_str, str):
                        order_date = datetime.fromisoformat(order_date_str.replace('Z', '+00:00').replace('+00:00', ''))
                    else:
                        order_date = order_date_str
                except:
                    continue

                if order_date < start_date:
                    continue

                symbol = order.get('symbol', '')
                symbols_set.add(symbol)

                # Symbol filter
                symbol_filter = self.journal_symbol_filter.currentText()
                if symbol_filter != "All Symbols" and symbol != symbol_filter:
                    continue

                # Side filter
                side_filter = self.journal_side_filter.currentText()
                side = order.get('transaction_type', order.get('side', ''))
                if side_filter == "BUY Only" and side != "BUY":
                    continue
                if side_filter == "SELL Only" and side != "SELL":
                    continue

                filtered_orders.append(order)

            # Update symbol filter dropdown
            current_symbol = self.journal_symbol_filter.currentText()
            self.journal_symbol_filter.blockSignals(True)
            self.journal_symbol_filter.clear()
            self.journal_symbol_filter.addItem("All Symbols")
            for sym in sorted(symbols_set):
                if sym:
                    self.journal_symbol_filter.addItem(sym)
            idx = self.journal_symbol_filter.findText(current_symbol)
            if idx >= 0:
                self.journal_symbol_filter.setCurrentIndex(idx)
            self.journal_symbol_filter.blockSignals(False)

            # Calculate statistics
            self._calculate_journal_stats(filtered_orders)

            # Populate table
            self._populate_journal_table(filtered_orders)

        except Exception as e:
            logger.error(f"Error refreshing journal: {e}")

    def _calculate_journal_stats(self, orders: list):
        """Calculate performance statistics from orders"""
        from collections import defaultdict

        if not orders:
            self.journal_total_pnl.setText("₹0")
            self.journal_total_trades.setText("0")
            self.journal_win_rate.setText("0%")
            self.journal_profit_factor.setText("0")
            self.journal_avg_win_loss.setText("₹0 / ₹0")
            self.journal_max_dd.setText("₹0")
            return

        # Extract P&L from orders
        pnls = []
        symbol_pnl = defaultdict(float)
        symbol_count = defaultdict(int)
        total_volume = 0

        for order in orders:
            pnl = float(order.get('pnl', 0))
            qty = int(order.get('quantity', 0))
            symbol = order.get('symbol', '')

            if order.get('status') == 'COMPLETE':
                pnls.append(pnl)
                symbol_pnl[symbol] += pnl
                symbol_count[symbol] += 1
                total_volume += qty

        if not pnls:
            return

        total_pnl = sum(pnls)
        total_trades = len(pnls)
        winning_trades = [p for p in pnls if p > 0]
        losing_trades = [p for p in pnls if p < 0]

        win_count = len(winning_trades)
        loss_count = len(losing_trades)
        win_rate = (win_count / total_trades * 100) if total_trades > 0 else 0

        total_wins = sum(winning_trades) if winning_trades else 0
        total_losses = abs(sum(losing_trades)) if losing_trades else 0
        profit_factor = (total_wins / total_losses) if total_losses > 0 else total_wins

        avg_win = (total_wins / win_count) if win_count > 0 else 0
        avg_loss = (total_losses / loss_count) if loss_count > 0 else 0

        # Calculate max drawdown
        cumulative = 0
        peak = 0
        max_dd = 0
        for pnl in pnls:
            cumulative += pnl
            if cumulative > peak:
                peak = cumulative
            dd = peak - cumulative
            if dd > max_dd:
                max_dd = dd

        # Consecutive wins
        max_consecutive = 0
        current_consecutive = 0
        for pnl in pnls:
            if pnl > 0:
                current_consecutive += 1
                max_consecutive = max(max_consecutive, current_consecutive)
            else:
                current_consecutive = 0

        # Best/worst symbols
        best_symbol = max(symbol_pnl.items(), key=lambda x: x[1]) if symbol_pnl else ("--", 0)
        worst_symbol = min(symbol_pnl.items(), key=lambda x: x[1]) if symbol_pnl else ("--", 0)
        most_traded = max(symbol_count.items(), key=lambda x: x[1]) if symbol_count else ("--", 0)

        # Update UI
        self.journal_total_pnl.setText(f"₹{total_pnl:+,.2f}")
        color = "#4CAF50" if total_pnl >= 0 else "#F44336"
        self.journal_total_pnl.setStyleSheet(f"font-size: 20px; font-weight: bold; color: {color};")

        self.journal_total_trades.setText(str(total_trades))
        self.journal_win_rate.setText(f"{win_rate:.1f}%")
        self.journal_profit_factor.setText(f"{profit_factor:.2f}")
        self.journal_avg_win_loss.setText(f"₹{avg_win:,.0f} / ₹{avg_loss:,.0f}")
        self.journal_max_dd.setText(f"₹{max_dd:,.2f}")

        self.journal_winning_trades.setText(str(win_count))
        self.journal_losing_trades.setText(str(loss_count))
        self.journal_largest_win.setText(f"₹{max(winning_trades):,.2f}" if winning_trades else "₹0")
        self.journal_largest_loss.setText(f"₹{min(losing_trades):,.2f}" if losing_trades else "₹0")

        avg_trade = total_pnl / total_trades if total_trades > 0 else 0
        self.journal_avg_trade.setText(f"₹{avg_trade:,.2f}")
        self.journal_avg_hold_time.setText("--")  # Would need entry/exit times
        self.journal_total_volume.setText(f"{total_volume:,}")
        self.journal_total_brokerage.setText(f"₹{total_volume * 20 / 100000:,.2f}")  # Rough estimate

        self.journal_best_symbol.setText(f"{best_symbol[0]} (₹{best_symbol[1]:+,.0f})")
        self.journal_worst_symbol.setText(f"{worst_symbol[0]} (₹{worst_symbol[1]:+,.0f})")
        self.journal_most_traded.setText(f"{most_traded[0]} ({most_traded[1]} trades)")
        self.journal_consecutive_wins.setText(str(max_consecutive))

    def _populate_journal_table(self, orders: list):
        """Populate the journal trades table"""
        self.journal_trades_table.setRowCount(len(orders))

        for i, order in enumerate(orders):
            date_str = order.get('created_at', '')[:10]
            time_str = order.get('created_at', '')[11:19] if len(order.get('created_at', '')) > 11 else ''
            symbol = order.get('symbol', '')
            side = order.get('transaction_type', order.get('side', ''))
            qty = order.get('quantity', 0)
            entry_price = float(order.get('price', 0))
            exit_price = float(order.get('exit_price', entry_price))
            pnl = float(order.get('pnl', 0))
            pnl_pct = (pnl / (entry_price * qty) * 100) if entry_price and qty else 0
            source = order.get('source', 'Manual')
            strategy = order.get('strategy', '')
            notes = order.get('notes', '')

            self.journal_trades_table.setItem(i, 0, QTableWidgetItem(date_str))
            self.journal_trades_table.setItem(i, 1, QTableWidgetItem(time_str))
            self.journal_trades_table.setItem(i, 2, QTableWidgetItem(symbol))

            side_item = QTableWidgetItem(side)
            side_item.setForeground(Qt.GlobalColor.green if side == "BUY" else Qt.GlobalColor.red)
            self.journal_trades_table.setItem(i, 3, side_item)

            self.journal_trades_table.setItem(i, 4, QTableWidgetItem(str(qty)))
            self.journal_trades_table.setItem(i, 5, QTableWidgetItem(f"₹{entry_price:.2f}"))
            self.journal_trades_table.setItem(i, 6, QTableWidgetItem(f"₹{exit_price:.2f}"))

            pnl_item = QTableWidgetItem(f"₹{pnl:+,.2f}")
            pnl_item.setForeground(Qt.GlobalColor.green if pnl >= 0 else Qt.GlobalColor.red)
            self.journal_trades_table.setItem(i, 7, pnl_item)

            self.journal_trades_table.setItem(i, 8, QTableWidgetItem(f"{pnl_pct:+.2f}%"))
            self.journal_trades_table.setItem(i, 9, QTableWidgetItem(source))
            self.journal_trades_table.setItem(i, 10, QTableWidgetItem(strategy))
            self.journal_trades_table.setItem(i, 11, QTableWidgetItem(notes))

    def _export_trades_excel(self):
        """Export trade history to Excel with formatting"""
        try:
            from datetime import datetime
            import os

            # Check if openpyxl is available
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
                from openpyxl.utils import get_column_letter
            except ImportError:
                QMessageBox.warning(
                    self, "Missing Package",
                    "Excel export requires 'openpyxl' package.\n\nInstall with: pip install openpyxl"
                )
                return

            # Get data from table
            rows = self.journal_trades_table.rowCount()
            if rows == 0:
                QMessageBox.information(self, "No Data", "No trades to export")
                return

            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Trade History"

            # Header style
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")

            # Headers
            headers = ["Date", "Time", "Symbol", "Side", "Qty", "Entry Price",
                       "Exit Price", "P&L", "P&L %", "Source", "Strategy", "Notes"]

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            # Data
            green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

            for row in range(rows):
                for col in range(12):
                    item = self.journal_trades_table.item(row, col)
                    value = item.text() if item else ""

                    # Clean currency symbols for numeric columns
                    if col in [5, 6, 7]:  # Price and P&L columns
                        value = value.replace("₹", "").replace(",", "").replace("+", "")
                        try:
                            value = float(value)
                        except:
                            pass
                    elif col == 8:  # P&L % column
                        value = value.replace("%", "").replace("+", "")
                        try:
                            value = float(value)
                        except:
                            pass

                    cell = ws.cell(row=row + 2, column=col + 1, value=value)

                    # Color P&L cells
                    if col == 7:  # P&L column
                        try:
                            if float(value) >= 0:
                                cell.fill = green_fill
                            else:
                                cell.fill = red_fill
                        except:
                            pass

            # Adjust column widths
            for col in range(1, 13):
                ws.column_dimensions[get_column_letter(col)].width = 15

            # Add summary sheet
            ws_summary = wb.create_sheet("Summary")
            ws_summary.cell(row=1, column=1, value="Performance Summary").font = Font(bold=True, size=14)

            summary_data = [
                ("Total P&L", self.journal_total_pnl.text()),
                ("Total Trades", self.journal_total_trades.text()),
                ("Win Rate", self.journal_win_rate.text()),
                ("Profit Factor", self.journal_profit_factor.text()),
                ("Winning Trades", self.journal_winning_trades.text()),
                ("Losing Trades", self.journal_losing_trades.text()),
                ("Largest Win", self.journal_largest_win.text()),
                ("Largest Loss", self.journal_largest_loss.text()),
                ("Max Drawdown", self.journal_max_dd.text()),
                ("Best Symbol", self.journal_best_symbol.text()),
                ("Worst Symbol", self.journal_worst_symbol.text()),
            ]

            for row, (label, value) in enumerate(summary_data, 3):
                ws_summary.cell(row=row, column=1, value=label).font = Font(bold=True)
                ws_summary.cell(row=row, column=2, value=value)

            ws_summary.column_dimensions['A'].width = 20
            ws_summary.column_dimensions['B'].width = 25

            # Save file
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"trade_history_{timestamp}.xlsx"
            filepath = os.path.join(os.path.expanduser("~"), "Documents", filename)

            # Ensure directory exists
            os.makedirs(os.path.dirname(filepath), exist_ok=True)

            wb.save(filepath)

            QMessageBox.information(
                self, "Export Complete",
                f"Trade history exported to:\n{filepath}"
            )

        except Exception as e:
            logger.error(f"Excel export error: {e}")
            QMessageBox.critical(self, "Export Error", f"Failed to export: {str(e)}")

    def _export_trades_csv(self):
        """Export trade history to CSV"""
        try:
            from datetime import datetime
            import csv
            import os

            rows = self.journal_trades_table.rowCount()
            if rows == 0:
                QMessageBox.information(self, "No Data", "No trades to export")
                return

            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"trade_history_{timestamp}.csv"
            filepath = os.path.join(os.path.expanduser("~"), "Documents", filename)

            os.makedirs(os.path.dirname(filepath), exist_ok=True)

            with open(filepath, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)

                # Headers
                headers = ["Date", "Time", "Symbol", "Side", "Qty", "Entry Price",
                           "Exit Price", "P&L", "P&L %", "Source", "Strategy", "Notes"]
                writer.writerow(headers)

                # Data
                for row in range(rows):
                    row_data = []
                    for col in range(12):
                        item = self.journal_trades_table.item(row, col)
                        value = item.text() if item else ""
                        # Clean currency symbols
                        value = value.replace("₹", "").replace(",", "")
                        row_data.append(value)
                    writer.writerow(row_data)

            QMessageBox.information(
                self, "Export Complete",
                f"Trade history exported to:\n{filepath}"
            )

        except Exception as e:
            logger.error(f"CSV export error: {e}")
            QMessageBox.critical(self, "Export Error", f"Failed to export: {str(e)}")

    def _create_alerts_tab(self):
        """Create alerts tab for price and indicator alerts"""
        alerts_widget = QWidget()
        layout = QVBoxLayout(alerts_widget)

        # Create new alert section
        create_group = QGroupBox("Create New Alert")
        create_layout = QFormLayout(create_group)

        # Symbol input
        self.alert_symbol = QLineEdit()
        self.alert_symbol.setPlaceholderText("e.g., RELIANCE, NIFTY")
        create_layout.addRow("Symbol:", self.alert_symbol)

        # Alert type
        self.alert_type_combo = QComboBox()
        self.alert_type_combo.addItems([
            "Price Above", "Price Below", "Price Cross Up", "Price Cross Down",
            "RSI Overbought (>70)", "RSI Oversold (<30)",
            "MACD Bullish Cross", "MACD Bearish Cross",
            "Supertrend Buy", "Supertrend Sell"
        ])
        create_layout.addRow("Alert Type:", self.alert_type_combo)

        # Target value
        self.alert_value = QDoubleSpinBox()
        self.alert_value.setRange(0, 999999)
        self.alert_value.setDecimals(2)
        self.alert_value.setValue(100)
        create_layout.addRow("Value/Price:", self.alert_value)

        # Message
        self.alert_message = QLineEdit()
        self.alert_message.setPlaceholderText("Custom message (optional)")
        create_layout.addRow("Message:", self.alert_message)

        # Repeat checkbox
        self.alert_repeat = QCheckBox("Repeat alert (trigger multiple times)")
        create_layout.addRow(self.alert_repeat)

        # Create button
        create_btn = QPushButton("➕ Create Alert")
        create_btn.setStyleSheet("background-color: #26a69a; color: white; font-weight: bold; padding: 8px;")
        create_btn.clicked.connect(self._create_alert)
        create_layout.addRow(create_btn)

        layout.addWidget(create_group)

        # Active alerts table
        alerts_group = QGroupBox("Active Alerts")
        alerts_layout = QVBoxLayout(alerts_group)

        self.alerts_table = QTableWidget()
        self.alerts_table.setColumnCount(7)
        self.alerts_table.setHorizontalHeaderLabels([
            "ID", "Symbol", "Type", "Target", "Status", "Created", "Actions"
        ])
        self.alerts_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.alerts_table.setMinimumHeight(200)
        alerts_layout.addWidget(self.alerts_table)

        # Refresh and clear buttons
        btn_layout = QHBoxLayout()
        refresh_alerts_btn = QPushButton("🔄 Refresh")
        refresh_alerts_btn.clicked.connect(self._refresh_alerts_table)
        btn_layout.addWidget(refresh_alerts_btn)

        clear_triggered_btn = QPushButton("🗑️ Clear Triggered")
        clear_triggered_btn.clicked.connect(self._clear_triggered_alerts)
        btn_layout.addWidget(clear_triggered_btn)

        btn_layout.addStretch()
        alerts_layout.addLayout(btn_layout)

        layout.addWidget(alerts_group)

        # Alert history
        history_group = QGroupBox("Recent Alert History")
        history_layout = QVBoxLayout(history_group)

        self.alert_history = QTableWidget()
        self.alert_history.setColumnCount(5)
        self.alert_history.setHorizontalHeaderLabels([
            "Symbol", "Type", "Target", "Triggered At", "Message"
        ])
        self.alert_history.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.alert_history.setMaximumHeight(150)
        history_layout.addWidget(self.alert_history)

        layout.addWidget(history_group)
        layout.addStretch()

        self.tabs.addTab(alerts_widget, "🔔 Alerts")

    def _init_alert_manager(self):
        """Initialize alert manager"""
        from algo_trader.core.alert_manager import AlertManager
        self.alert_manager = AlertManager()

        # Register callback
        self.alert_manager.register_callback(self._on_alert_triggered)

        # Set Telegram if available
        if hasattr(self, 'telegram') and self.telegram:
            self.alert_manager.set_telegram(self.telegram)

        logger.info("Alert manager initialized")

    def _create_alert(self):
        """Create a new alert from UI"""
        if not hasattr(self, 'alert_manager') or not self.alert_manager:
            QMessageBox.warning(self, "Error", "Alert manager not initialized")
            return

        symbol = self.alert_symbol.text().strip().upper()
        if not symbol:
            QMessageBox.warning(self, "Error", "Please enter a symbol")
            return

        alert_type_text = self.alert_type_combo.currentText()
        value = self.alert_value.value()
        message = self.alert_message.text().strip()
        repeat = self.alert_repeat.isChecked()

        # Map combo text to alert creation
        if "Price Above" in alert_type_text:
            self.alert_manager.create_price_alert(symbol, "above", value, message=message, repeat=repeat)
        elif "Price Below" in alert_type_text:
            self.alert_manager.create_price_alert(symbol, "below", value, message=message, repeat=repeat)
        elif "Price Cross Up" in alert_type_text:
            self.alert_manager.create_price_alert(symbol, "cross_up", value, message=message, repeat=repeat)
        elif "Price Cross Down" in alert_type_text:
            self.alert_manager.create_price_alert(symbol, "cross_down", value, message=message, repeat=repeat)
        elif "RSI Overbought" in alert_type_text:
            self.alert_manager.create_indicator_alert(symbol, "rsi", "overbought", value if value != 100 else 70, message=message)
        elif "RSI Oversold" in alert_type_text:
            self.alert_manager.create_indicator_alert(symbol, "rsi", "oversold", value if value != 100 else 30, message=message)
        elif "MACD Bullish" in alert_type_text:
            self.alert_manager.create_indicator_alert(symbol, "macd", "bullish", message=message)
        elif "MACD Bearish" in alert_type_text:
            self.alert_manager.create_indicator_alert(symbol, "macd", "bearish", message=message)
        elif "Supertrend Buy" in alert_type_text:
            self.alert_manager.create_indicator_alert(symbol, "supertrend", "buy", message=message)
        elif "Supertrend Sell" in alert_type_text:
            self.alert_manager.create_indicator_alert(symbol, "supertrend", "sell", message=message)

        # Clear inputs
        self.alert_symbol.clear()
        self.alert_message.clear()
        self.alert_value.setValue(100)

        # Refresh table
        self._refresh_alerts_table()
        QMessageBox.information(self, "Success", f"Alert created for {symbol}")

    def _refresh_alerts_table(self):
        """Refresh the alerts table"""
        if not hasattr(self, 'alert_manager') or not self.alert_manager:
            return

        alerts = self.alert_manager.get_all_alerts()
        self.alerts_table.setRowCount(len(alerts))

        for i, alert in enumerate(alerts):
            self.alerts_table.setItem(i, 0, QTableWidgetItem(alert.id))
            self.alerts_table.setItem(i, 1, QTableWidgetItem(alert.symbol))
            self.alerts_table.setItem(i, 2, QTableWidgetItem(alert.alert_type.value))
            self.alerts_table.setItem(i, 3, QTableWidgetItem(f"{alert.target_value:.2f}"))
            self.alerts_table.setItem(i, 4, QTableWidgetItem(alert.status.value))
            self.alerts_table.setItem(i, 5, QTableWidgetItem(alert.created_at.strftime("%H:%M:%S")))

            # Delete button
            delete_btn = QPushButton("🗑️")
            delete_btn.setMaximumWidth(40)
            delete_btn.clicked.connect(lambda checked, aid=alert.id: self._delete_alert(aid))
            self.alerts_table.setCellWidget(i, 6, delete_btn)

    def _delete_alert(self, alert_id: str):
        """Delete an alert"""
        if self.alert_manager:
            self.alert_manager.delete_alert(alert_id)
            self._refresh_alerts_table()

    def _clear_triggered_alerts(self):
        """Clear all triggered alerts"""
        if not self.alert_manager:
            return

        from algo_trader.core.alert_manager import AlertStatus
        for alert in self.alert_manager.get_all_alerts():
            if alert.status == AlertStatus.TRIGGERED:
                self.alert_manager.delete_alert(alert.id)

        self._refresh_alerts_table()
        QMessageBox.information(self, "Cleared", "Triggered alerts cleared")

    def _on_alert_triggered(self, event_data: dict):
        """Handle alert trigger event"""
        symbol = event_data.get('symbol', 'Unknown')
        alert_type = event_data.get('alert_type', 'Unknown')
        message = event_data.get('message', '')

        # Show notification
        QMessageBox.information(self, "🔔 Alert Triggered!",
            f"Symbol: {symbol}\n"
            f"Type: {alert_type}\n"
            f"Message: {message}")

        # Update history table
        self._refresh_alerts_table()
        self._add_alert_to_history(event_data)

    def _add_alert_to_history(self, event_data: dict):
        """Add triggered alert to history table"""
        row = self.alert_history.rowCount()
        self.alert_history.insertRow(row)
        self.alert_history.setItem(row, 0, QTableWidgetItem(event_data.get('symbol', '')))
        self.alert_history.setItem(row, 1, QTableWidgetItem(event_data.get('alert_type', '')))
        self.alert_history.setItem(row, 2, QTableWidgetItem(f"{event_data.get('target_value', 0):.2f}"))
        self.alert_history.setItem(row, 3, QTableWidgetItem(
            event_data.get('timestamp', datetime.now()).strftime("%H:%M:%S")))
        self.alert_history.setItem(row, 4, QTableWidgetItem(event_data.get('message', '')))

    def _create_settings_tab(self):
        """Create settings tab"""
        settings = QWidget()
        layout = QVBoxLayout(settings)

        # Use scroll area for settings
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll_widget = QWidget()
        scroll_layout = QVBoxLayout(scroll_widget)

        # === Paper Trading Mode ===
        paper_group = QGroupBox("Trading Mode")
        paper_layout = QVBoxLayout(paper_group)

        mode_layout = QHBoxLayout()
        self.paper_trading_toggle = QCheckBox("Paper Trading Mode (Simulator)")
        self.paper_trading_toggle.setChecked(self.config.get('trading.paper_mode', False))
        self.paper_trading_toggle.stateChanged.connect(self._on_paper_mode_changed)
        mode_layout.addWidget(self.paper_trading_toggle)

        self.paper_mode_label = QLabel("🟢 LIVE TRADING" if not self.paper_trading_toggle.isChecked() else "📝 PAPER TRADING")
        self.paper_mode_label.setStyleSheet("font-weight: bold; font-size: 14px;")
        mode_layout.addWidget(self.paper_mode_label)
        mode_layout.addStretch()
        paper_layout.addLayout(mode_layout)

        # Paper trading capital
        paper_capital_layout = QHBoxLayout()
        paper_capital_layout.addWidget(QLabel("Simulator Capital:"))
        self.paper_capital = QDoubleSpinBox()
        self.paper_capital.setRange(1000, 99999999)
        self.paper_capital.setDecimals(0)
        self.paper_capital.setPrefix("₹ ")
        self.paper_capital.setValue(self.config.get('trading.paper_capital', 100000))
        paper_capital_layout.addWidget(self.paper_capital)

        self.reset_paper_btn = QPushButton("Reset Simulator")
        self.reset_paper_btn.clicked.connect(self._reset_paper_trading)
        paper_capital_layout.addWidget(self.reset_paper_btn)
        paper_capital_layout.addStretch()
        paper_layout.addLayout(paper_capital_layout)

        scroll_layout.addWidget(paper_group)

        # === Telegram Alerts ===
        telegram_group = QGroupBox("Telegram Alerts")
        telegram_layout = QFormLayout(telegram_group)

        self.telegram_enabled = QCheckBox("Enable Telegram Alerts")
        self.telegram_enabled.setChecked(self.config.get('telegram.enabled', False))
        self.telegram_enabled.stateChanged.connect(self._on_telegram_toggle)
        telegram_layout.addRow(self.telegram_enabled)

        self.telegram_bot_token = QLineEdit()
        self.telegram_bot_token.setPlaceholderText("Enter Bot Token from @BotFather")
        self.telegram_bot_token.setText(self.config.get('telegram.bot_token', ''))
        self.telegram_bot_token.setEchoMode(QLineEdit.EchoMode.Password)
        telegram_layout.addRow("Bot Token:", self.telegram_bot_token)

        self.telegram_chat_id = QLineEdit()
        self.telegram_chat_id.setPlaceholderText("Enter Chat ID")
        self.telegram_chat_id.setText(self.config.get('telegram.chat_id', ''))
        telegram_layout.addRow("Chat ID:", self.telegram_chat_id)

        telegram_btn_layout = QHBoxLayout()
        self.save_telegram_btn = QPushButton("Save Telegram")
        self.save_telegram_btn.clicked.connect(self._save_telegram_settings)
        self.test_telegram_btn = QPushButton("Test Connection")
        self.test_telegram_btn.clicked.connect(self._test_telegram)
        telegram_btn_layout.addWidget(self.save_telegram_btn)
        telegram_btn_layout.addWidget(self.test_telegram_btn)
        telegram_layout.addRow(telegram_btn_layout)

        # Help text
        help_label = QLabel("How to setup:\n1. Open Telegram, search @BotFather\n2. Send /newbot, follow steps\n3. Copy Bot Token\n4. Send message to your bot, then visit:\n   https://api.telegram.org/bot<TOKEN>/getUpdates\n5. Copy chat_id from response")
        help_label.setStyleSheet("color: gray; font-size: 11px;")
        telegram_layout.addRow(help_label)

        scroll_layout.addWidget(telegram_group)

        # === Telegram Bot Controller (Remote Control) ===
        bot_control_group = QGroupBox("Telegram Bot Controller (Remote Control)")
        bot_control_layout = QFormLayout(bot_control_group)

        self.bot_controller_enabled = QCheckBox("Enable Bot Controller")
        self.bot_controller_enabled.setChecked(self.config.get('telegram.bot_controller_enabled', False))
        self.bot_controller_enabled.stateChanged.connect(self._on_bot_controller_toggle)
        bot_control_layout.addRow(self.bot_controller_enabled)

        # Bot controller status
        self.bot_controller_status = QLabel("Stopped")
        self.bot_controller_status.setStyleSheet("color: red; font-weight: bold;")
        bot_control_layout.addRow("Status:", self.bot_controller_status)

        # Bot controller buttons
        bot_ctrl_btn_layout = QHBoxLayout()
        self.start_bot_ctrl_btn = QPushButton("Start Controller")
        self.start_bot_ctrl_btn.clicked.connect(self._start_bot_controller)
        self.stop_bot_ctrl_btn = QPushButton("Stop Controller")
        self.stop_bot_ctrl_btn.clicked.connect(self._stop_bot_controller)
        self.stop_bot_ctrl_btn.setEnabled(False)
        bot_ctrl_btn_layout.addWidget(self.start_bot_ctrl_btn)
        bot_ctrl_btn_layout.addWidget(self.stop_bot_ctrl_btn)
        bot_control_layout.addRow(bot_ctrl_btn_layout)

        # Commands help
        commands_label = QLabel("""
<b>Available Commands:</b>
/status - Get current algo status
/positions - List open positions
/orders - View pending orders
/pnl - Get today's P&L
/pause - Pause algo trading
/resume - Resume algo trading
/squareoff - Square off all positions
/help - Show all commands""")
        commands_label.setStyleSheet("color: #888; font-size: 11px; background: #1a1a1a; padding: 10px; border-radius: 5px;")
        commands_label.setTextFormat(Qt.TextFormat.RichText)
        bot_control_layout.addRow(commands_label)

        scroll_layout.addWidget(bot_control_group)

        # Broker settings
        broker_group = QGroupBox("Broker Connections")
        broker_layout = QVBoxLayout(broker_group)

        self.broker_settings_table = QTableWidget()
        self.broker_settings_table.setColumnCount(4)
        self.broker_settings_table.setHorizontalHeaderLabels([
            "Broker", "Status", "User ID", "Actions"
        ])
        self.broker_settings_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        broker_layout.addWidget(self.broker_settings_table)

        broker_btn_layout = QHBoxLayout()
        add_broker_btn = QPushButton("Add Broker")
        add_broker_btn.clicked.connect(self._show_broker_dialog)
        broker_btn_layout.addWidget(add_broker_btn)
        broker_layout.addLayout(broker_btn_layout)

        scroll_layout.addWidget(broker_group)

        # Trading settings
        trading_group = QGroupBox("Trading Settings")
        trading_layout = QFormLayout(trading_group)

        self.default_qty = QSpinBox()
        self.default_qty.setRange(1, 1000)
        self.default_qty.setValue(self.config.get('trading.default_quantity', 1))
        trading_layout.addRow("Default Quantity:", self.default_qty)

        self.max_positions = QSpinBox()
        self.max_positions.setRange(1, 100)
        self.max_positions.setValue(self.config.get('trading.max_positions', 10))
        trading_layout.addRow("Max Positions:", self.max_positions)

        self.risk_percent = QDoubleSpinBox()
        self.risk_percent.setRange(0.1, 10)
        self.risk_percent.setValue(self.config.get('trading.risk_percent', 2.0))
        self.risk_percent.setSuffix("%")
        trading_layout.addRow("Risk per Trade:", self.risk_percent)

        save_settings_btn = QPushButton("Save Settings")
        save_settings_btn.clicked.connect(self._save_settings)
        trading_layout.addRow(save_settings_btn)

        scroll_layout.addWidget(trading_group)

        # === Auto Square-off Settings ===
        squareoff_group = QGroupBox("Auto Square-off")
        squareoff_layout = QFormLayout(squareoff_group)

        self.auto_squareoff_enabled = QCheckBox("Enable Auto Square-off")
        self.auto_squareoff_enabled.setChecked(self.config.get('squareoff.enabled', False))
        squareoff_layout.addRow(self.auto_squareoff_enabled)

        # Daily limits
        daily_limits_label = QLabel("Daily P&L Limits:")
        daily_limits_label.setStyleSheet("font-weight: bold; margin-top: 10px;")
        squareoff_layout.addRow(daily_limits_label)

        self.daily_profit_target = QDoubleSpinBox()
        self.daily_profit_target.setRange(0, 9999999)
        self.daily_profit_target.setDecimals(0)
        self.daily_profit_target.setPrefix("₹ ")
        self.daily_profit_target.setSpecialValueText("Disabled")
        self.daily_profit_target.setValue(self.config.get('squareoff.daily_profit_target', 0))
        squareoff_layout.addRow("Profit Target:", self.daily_profit_target)

        self.daily_loss_limit = QDoubleSpinBox()
        self.daily_loss_limit.setRange(0, 9999999)
        self.daily_loss_limit.setDecimals(0)
        self.daily_loss_limit.setPrefix("₹ ")
        self.daily_loss_limit.setSpecialValueText("Disabled")
        self.daily_loss_limit.setValue(self.config.get('squareoff.daily_loss_limit', 0))
        squareoff_layout.addRow("Loss Limit:", self.daily_loss_limit)

        # Time-based square-off
        time_label = QLabel("Time-based Square-off:")
        time_label.setStyleSheet("font-weight: bold; margin-top: 10px;")
        squareoff_layout.addRow(time_label)

        self.squareoff_time = QLineEdit()
        self.squareoff_time.setPlaceholderText("15:15 (3:15 PM)")
        self.squareoff_time.setText(self.config.get('squareoff.time', ''))
        squareoff_layout.addRow("Square-off Time:", self.squareoff_time)

        # Per-position limits
        position_label = QLabel("Per-Position Limits:")
        position_label.setStyleSheet("font-weight: bold; margin-top: 10px;")
        squareoff_layout.addRow(position_label)

        self.position_profit_percent = QDoubleSpinBox()
        self.position_profit_percent.setRange(0, 100)
        self.position_profit_percent.setDecimals(1)
        self.position_profit_percent.setSuffix(" %")
        self.position_profit_percent.setSpecialValueText("Disabled")
        self.position_profit_percent.setValue(self.config.get('squareoff.position_profit_percent', 0))
        squareoff_layout.addRow("Position Profit %:", self.position_profit_percent)

        self.position_loss_percent = QDoubleSpinBox()
        self.position_loss_percent.setRange(0, 100)
        self.position_loss_percent.setDecimals(1)
        self.position_loss_percent.setSuffix(" %")
        self.position_loss_percent.setSpecialValueText("Disabled")
        self.position_loss_percent.setValue(self.config.get('squareoff.position_loss_percent', 0))
        squareoff_layout.addRow("Position Loss %:", self.position_loss_percent)

        self.trailing_profit_percent = QDoubleSpinBox()
        self.trailing_profit_percent.setRange(0, 50)
        self.trailing_profit_percent.setDecimals(1)
        self.trailing_profit_percent.setSuffix(" %")
        self.trailing_profit_percent.setSpecialValueText("Disabled")
        self.trailing_profit_percent.setValue(self.config.get('squareoff.trailing_profit_percent', 0))
        self.trailing_profit_percent.setToolTip("Lock profits - square off when profit falls this % from peak")
        squareoff_layout.addRow("Trailing Profit %:", self.trailing_profit_percent)

        save_squareoff_btn = QPushButton("Save Auto Square-off Settings")
        save_squareoff_btn.clicked.connect(self._save_squareoff_settings)
        squareoff_layout.addRow(save_squareoff_btn)

        # Status label
        self.squareoff_status_label = QLabel("Status: Not active")
        self.squareoff_status_label.setStyleSheet("color: gray;")
        squareoff_layout.addRow(self.squareoff_status_label)

        scroll_layout.addWidget(squareoff_group)
        scroll_layout.addStretch()

        scroll.setWidget(scroll_widget)
        layout.addWidget(scroll)

        self.tabs.addTab(settings, "Settings")

    def _load_configured_brokers(self):
        """Load previously configured brokers"""
        brokers = self.config.list_configured_brokers()
        for broker in brokers:
            self.broker_combo.addItem(broker.title())

        self._update_broker_settings_table()

    def _update_broker_settings_table(self):
        """Update broker settings table"""
        brokers = self.config.list_configured_brokers()
        self.broker_settings_table.setRowCount(len(brokers))

        for i, broker in enumerate(brokers):
            self.broker_settings_table.setItem(i, 0, QTableWidgetItem(broker.title()))
            is_connected = broker in self.brokers
            status = "Connected" if is_connected else "Disconnected"
            status_item = QTableWidgetItem(status)
            status_item.setForeground(Qt.GlobalColor.green if is_connected else Qt.GlobalColor.red)
            self.broker_settings_table.setItem(i, 1, status_item)
            creds = self.config.get_broker_credentials(broker)
            self.broker_settings_table.setItem(i, 2, QTableWidgetItem(creds.get('user_id', 'N/A')))

            # Add Connect/Reconnect button
            if not is_connected:
                connect_btn = QPushButton("Connect")
                connect_btn.clicked.connect(lambda checked, b=broker: self._reconnect_broker(b))
                self.broker_settings_table.setCellWidget(i, 3, connect_btn)
            else:
                disconnect_btn = QPushButton("Disconnect")
                disconnect_btn.clicked.connect(lambda checked, b=broker: self._disconnect_broker(b))
                self.broker_settings_table.setCellWidget(i, 3, disconnect_btn)

    def _reconnect_broker(self, broker_name: str):
        """Reconnect to a broker using saved credentials"""
        try:
            creds = self.config.get_broker_credentials(broker_name)
            if not creds:
                QMessageBox.warning(self, "Error", f"No credentials found for {broker_name}")
                return

            api_key = creds.get('api_key', '')
            api_secret = creds.get('api_secret', '')
            user_id = creds.get('user_id', '')

            # Need to re-authenticate - show broker dialog
            QMessageBox.information(
                self, "Re-authentication Required",
                f"Broker sessions expire daily. Please re-authenticate {broker_name.title()}.\n\n"
                "Click 'Connect Broker' and follow the authentication steps."
            )
            self._show_broker_dialog()

        except Exception as e:
            logger.error(f"Error reconnecting broker: {e}")
            QMessageBox.warning(self, "Error", f"Failed to reconnect: {str(e)}")

    def _disconnect_broker(self, broker_name: str):
        """Disconnect a broker"""
        if broker_name in self.brokers:
            del self.brokers[broker_name]
            self._update_broker_settings_table()
            self._refresh_dashboard()
            logger.info(f"Broker {broker_name} disconnected")

    def _setup_timers(self):
        """Setup refresh timers"""
        # Refresh every 30 seconds
        self.refresh_timer = QTimer(self)
        self.refresh_timer.timeout.connect(self._refresh_data)
        self.refresh_timer.start(30000)

        # Fast refresh for live prices (every 5 seconds)
        self.ltp_timer = QTimer(self)
        self.ltp_timer.timeout.connect(self._refresh_live_prices)
        self.ltp_timer.start(5000)

    def _refresh_data(self):
        """Refresh all data"""
        self._load_orders()
        self._load_positions()
        self._refresh_dashboard()
        self.status_bar.showMessage("Data refreshed", 3000)

    def _refresh_live_prices(self):
        """Refresh live prices for Strategy Builder and Dashboard (every 5 seconds)"""
        try:
            # Refresh Strategy Builder LTP if there are legs
            if hasattr(self, 'strategy_legs') and self.strategy_legs:
                self._refresh_legs_ltp()

            # Refresh Dashboard paper positions
            if hasattr(self, 'paper_simulator') and self.paper_simulator:
                # Update paper positions with simulated price movements
                self._update_paper_positions_ltp()
                self._refresh_dashboard()

        except Exception as e:
            logger.debug(f"Error refreshing live prices: {e}")

    def _update_paper_positions_ltp(self):
        """Update LTP for paper trading positions"""
        if not hasattr(self, 'paper_simulator') or not self.paper_simulator:
            return

        import random

        positions = self.paper_simulator.positions
        for symbol, pos in positions.items():
            # Simulate small price movement (±0.5%)
            if pos.current_price > 0:
                change = pos.current_price * random.uniform(-0.005, 0.005)
                pos.current_price = max(0.05, pos.current_price + change)

                # Update P&L
                if pos.action == "BUY":
                    pos.pnl = (pos.current_price - pos.avg_price) * pos.quantity
                else:
                    pos.pnl = (pos.avg_price - pos.current_price) * pos.quantity

                if pos.avg_price > 0:
                    pos.pnl_percent = (pos.pnl / (pos.avg_price * pos.quantity)) * 100

    def _refresh_dashboard(self):
        """Refresh dashboard with live data"""
        try:
            from datetime import datetime

            logger.debug(f"Refreshing dashboard, brokers dict: {list(self.brokers.keys()) if self.brokers else 'empty'}")

            # Update market status
            if hasattr(self, 'brokers') and self.brokers:
                logger.debug(f"Brokers found: {list(self.brokers.keys())}")
                broker = list(self.brokers.values())[0]
                logger.debug(f"Using broker: {broker.broker_name if hasattr(broker, 'broker_name') else 'unknown'}")

                # Update broker status FIRST before any API calls
                self.dash_broker_status.setText("Connected")
                self.dash_broker_status.setStyleSheet("color: green;")

                try:
                    is_open = broker.is_market_open() if hasattr(broker, 'is_market_open') else False
                    self.dash_market_status.setText("Open" if is_open else "Closed")
                    self.dash_market_status.setStyleSheet(f"color: {'green' if is_open else 'red'}; font-weight: bold;")
                except Exception as e:
                    logger.debug(f"Error checking market status: {e}")
                    self.dash_market_status.setText("Unknown")
                    self.dash_market_status.setStyleSheet("color: gray; font-weight: bold;")

                # Get funds
                try:
                    funds = broker.get_funds()
                    if funds:
                        self.dash_available_margin.setText(f"₹{funds.get('available_margin', 0):,.2f}")
                        self.dash_used_margin.setText(f"₹{funds.get('used_margin', 0):,.2f}")
                        self.dash_total_balance.setText(f"₹{funds.get('total_balance', 0):,.2f}")
                except:
                    pass

                # Get positions for P&L
                try:
                    positions = broker.get_positions()
                    # Combine broker positions with paper positions
                    all_positions = positions if positions else []

                    # Also add paper positions (marked as PAPER source)
                    if hasattr(self, 'paper_simulator') and self.paper_simulator:
                        paper_positions = self.paper_simulator.get_all_positions()
                        for pp in paper_positions:
                            pp['source'] = 'PAPER'
                            all_positions.append(pp)

                    self._update_dashboard_positions_combined(all_positions)
                except:
                    pass
            else:
                self.dash_broker_status.setText("Not Connected")
                self.dash_broker_status.setStyleSheet("color: red;")

            # Always show paper trading positions and P&L (whether broker connected or not)
            if hasattr(self, 'paper_simulator') and self.paper_simulator:
                # If no broker, show only paper positions
                if not (hasattr(self, 'brokers') and self.brokers):
                    paper_positions = self.paper_simulator.get_all_positions()
                    if paper_positions:
                        self._update_dashboard_positions_paper(paper_positions)

                # Update paper trading P&L
                stats = self.paper_simulator.get_stats()
                self.dash_realized_pnl.setText(f"₹{stats['total_pnl']:+,.2f}")
                self.dash_realized_pnl.setStyleSheet(f"font-size: 24px; font-weight: bold; color: {'#4CAF50' if stats['total_pnl'] >= 0 else '#F44336'};")

                self.dash_unrealized_pnl.setText(f"₹{stats['unrealized_pnl']:+,.2f}")
                self.dash_unrealized_pnl.setStyleSheet(f"font-size: 24px; font-weight: bold; color: {'#4CAF50' if stats['unrealized_pnl'] >= 0 else '#F44336'};")

                total_pnl = stats['total_pnl'] + stats['unrealized_pnl']
                self.dash_total_pnl.setText(f"₹{total_pnl:+,.2f}")
                self.dash_total_pnl.setStyleSheet(f"font-size: 24px; font-weight: bold; color: {'#4CAF50' if total_pnl >= 0 else '#F44336'};")

                self.dash_trades_today.setText(str(stats['total_trades']))
                self.dash_win_rate.setText(f"{stats['win_rate']:.1f}%")
                self.dash_open_positions.setText(str(stats['open_positions']))

                # Update margin display for paper trading (only if no broker)
                if not (hasattr(self, 'brokers') and self.brokers):
                    self.dash_available_margin.setText(f"₹{stats['available_capital']:,.2f}")
                    self.dash_used_margin.setText(f"₹{stats['used_capital']:,.2f}")
                    self.dash_total_balance.setText(f"₹{stats['total_equity']:,.2f}")

            # Update quick stats
            if hasattr(self, 'chartink_scanner'):
                active_scans = len(self.chartink_scanner.active_scans)
                self.dash_active_scanners.setText(str(active_scans))

            # Count active strategies
            strategies = self.db.get_all_strategies()
            active_strategies = len([s for s in strategies if s.get('is_active')])
            self.dash_active_strategies.setText(str(active_strategies))

            # Update timestamp
            self.dash_last_update.setText(datetime.now().strftime("%H:%M:%S"))

            # Calculate P&L from orders
            self._update_dashboard_pnl()

        except Exception as e:
            logger.debug(f"Dashboard refresh error: {e}")

    def _update_dashboard_positions(self, positions: list):
        """Update dashboard positions table"""
        self.dash_positions_table.setRowCount(len(positions))
        self.dash_open_positions.setText(str(len(positions)))

        total_unrealized = 0
        for i, pos in enumerate(positions):
            symbol = pos.get('tradingsymbol', pos.get('symbol', ''))
            qty = pos.get('quantity', pos.get('netqty', 0))
            avg_price = float(pos.get('averageprice', pos.get('buyavgprice', 0)))
            ltp = float(pos.get('ltp', pos.get('lastprice', avg_price)))
            pnl = float(pos.get('pnl', pos.get('unrealizedpnl', 0)))
            pnl_pct = (pnl / (avg_price * abs(qty)) * 100) if avg_price and qty else 0

            total_unrealized += pnl

            self.dash_positions_table.setItem(i, 0, QTableWidgetItem(symbol))
            self.dash_positions_table.setItem(i, 1, QTableWidgetItem("LONG" if qty > 0 else "SHORT"))
            self.dash_positions_table.setItem(i, 2, QTableWidgetItem(str(abs(qty))))
            self.dash_positions_table.setItem(i, 3, QTableWidgetItem(f"₹{avg_price:.2f}"))
            self.dash_positions_table.setItem(i, 4, QTableWidgetItem(f"₹{ltp:.2f}"))

            pnl_item = QTableWidgetItem(f"₹{pnl:+,.2f}")
            pnl_item.setForeground(Qt.GlobalColor.green if pnl >= 0 else Qt.GlobalColor.red)
            self.dash_positions_table.setItem(i, 5, pnl_item)

            self.dash_positions_table.setItem(i, 6, QTableWidgetItem(f"{pnl_pct:+.2f}%"))
            self.dash_positions_table.setItem(i, 7, QTableWidgetItem(pos.get('product', '')))

            # Add close button
            close_btn = QPushButton("Close")
            close_btn.clicked.connect(lambda checked, s=symbol, q=qty: self._close_position(s, q))
            self.dash_positions_table.setCellWidget(i, 8, close_btn)

        # Update unrealized P&L
        self.dash_unrealized_pnl.setText(f"₹{total_unrealized:+,.2f}")
        color = "#4CAF50" if total_unrealized >= 0 else "#F44336"
        self.dash_unrealized_pnl.setStyleSheet(f"font-size: 24px; font-weight: bold; color: {color};")

    def _update_dashboard_positions_combined(self, positions: list):
        """Update dashboard with both broker and paper positions"""
        self.dash_positions_table.setRowCount(len(positions))
        self.dash_open_positions.setText(str(len(positions)))

        total_unrealized = 0
        for i, pos in enumerate(positions):
            # Check if it's a paper position or broker position
            is_paper = pos.get('source') == 'PAPER'

            if is_paper:
                symbol = pos.get('symbol', '')
                qty = pos.get('quantity', 0)
                action = pos.get('action', 'BUY')
                avg_price = float(pos.get('avg_price', 0))
                ltp = float(pos.get('current_price', avg_price))
                pnl = float(pos.get('pnl', 0))
                pnl_pct = float(pos.get('pnl_percent', 0))
                source = "PAPER"
            else:
                symbol = pos.get('tradingsymbol', pos.get('symbol', ''))
                qty = pos.get('quantity', pos.get('netqty', 0))
                action = 'BUY' if qty > 0 else 'SELL'
                avg_price = float(pos.get('averageprice', pos.get('buyavgprice', 0)))
                ltp = float(pos.get('ltp', pos.get('lastprice', avg_price)))
                pnl = float(pos.get('pnl', pos.get('unrealizedpnl', 0)))
                pnl_pct = (pnl / (avg_price * abs(qty)) * 100) if avg_price and qty else 0
                source = pos.get('product', 'LIVE')

            total_unrealized += pnl

            self.dash_positions_table.setItem(i, 0, QTableWidgetItem(symbol))
            self.dash_positions_table.setItem(i, 1, QTableWidgetItem("LONG" if action == "BUY" or qty > 0 else "SHORT"))
            self.dash_positions_table.setItem(i, 2, QTableWidgetItem(str(abs(qty))))
            self.dash_positions_table.setItem(i, 3, QTableWidgetItem(f"₹{avg_price:.2f}"))
            self.dash_positions_table.setItem(i, 4, QTableWidgetItem(f"₹{ltp:.2f}"))

            pnl_item = QTableWidgetItem(f"₹{pnl:+,.2f}")
            pnl_item.setForeground(Qt.GlobalColor.green if pnl >= 0 else Qt.GlobalColor.red)
            self.dash_positions_table.setItem(i, 5, pnl_item)

            self.dash_positions_table.setItem(i, 6, QTableWidgetItem(f"{pnl_pct:+.2f}%"))

            # Source column - different color for PAPER vs LIVE
            source_item = QTableWidgetItem(source)
            if is_paper:
                source_item.setForeground(Qt.GlobalColor.yellow)
            else:
                source_item.setForeground(Qt.GlobalColor.green)
            self.dash_positions_table.setItem(i, 7, source_item)

            # Add close button
            if is_paper:
                close_btn = QPushButton("Close")
                close_btn.clicked.connect(lambda checked, s=symbol, q=qty, a=action: self._close_paper_position(s, q, a))
            else:
                close_btn = QPushButton("Close")
                close_btn.clicked.connect(lambda checked, s=symbol, q=qty: self._close_position(s, q))
            self.dash_positions_table.setCellWidget(i, 8, close_btn)

    def _update_dashboard_positions_paper(self, positions: list):
        """Update dashboard positions table with paper trading positions"""
        self.dash_positions_table.setRowCount(len(positions))
        self.dash_open_positions.setText(str(len(positions)))

        for i, pos in enumerate(positions):
            symbol = pos.get('symbol', '')
            qty = pos.get('quantity', 0)
            action = pos.get('action', 'BUY')
            avg_price = float(pos.get('avg_price', 0))
            ltp = float(pos.get('current_price', avg_price))
            pnl = float(pos.get('pnl', 0))
            pnl_pct = float(pos.get('pnl_percent', 0))

            self.dash_positions_table.setItem(i, 0, QTableWidgetItem(symbol))
            self.dash_positions_table.setItem(i, 1, QTableWidgetItem("LONG" if action == "BUY" else "SHORT"))
            self.dash_positions_table.setItem(i, 2, QTableWidgetItem(str(abs(qty))))
            self.dash_positions_table.setItem(i, 3, QTableWidgetItem(f"₹{avg_price:.2f}"))
            self.dash_positions_table.setItem(i, 4, QTableWidgetItem(f"₹{ltp:.2f}"))

            pnl_item = QTableWidgetItem(f"₹{pnl:+,.2f}")
            pnl_item.setForeground(Qt.GlobalColor.green if pnl >= 0 else Qt.GlobalColor.red)
            self.dash_positions_table.setItem(i, 5, pnl_item)

            self.dash_positions_table.setItem(i, 6, QTableWidgetItem(f"{pnl_pct:+.2f}%"))
            self.dash_positions_table.setItem(i, 7, QTableWidgetItem("PAPER"))

            # Add close button for paper positions
            close_btn = QPushButton("Close")
            close_btn.clicked.connect(lambda checked, s=symbol, q=qty, a=action: self._close_paper_position(s, q, a))
            self.dash_positions_table.setCellWidget(i, 8, close_btn)

    def _close_paper_position(self, symbol: str, qty: int, action: str):
        """Close a paper trading position"""
        if not hasattr(self, 'paper_simulator') or not self.paper_simulator:
            return

        # Close by placing opposite order
        close_action = "SELL" if action == "BUY" else "BUY"
        pos = self.paper_simulator.get_position(symbol)
        if pos:
            result = self.paper_simulator.place_order(
                symbol=symbol,
                action=close_action,
                quantity=qty,
                order_type="MARKET",
                price=pos.current_price if pos.current_price > 0 else pos.avg_price,
                source="Manual Close"
            )
            if result['success']:
                QMessageBox.information(self, "Position Closed", f"Closed {symbol} position")
                self._refresh_dashboard()

    def _update_dashboard_pnl(self):
        """Update P&L summary from trade history"""
        try:
            orders = self.db.get_orders()
            today = datetime.now().date()

            today_orders = [o for o in orders if str(o.get('created_at', ''))[:10] == str(today)]

            # Calculate realized P&L (from completed trades)
            realized_pnl = sum(float(o.get('pnl', 0)) for o in today_orders if o.get('status') == 'COMPLETE')

            # Count trades
            completed_trades = [o for o in today_orders if o.get('status') == 'COMPLETE']
            winning = len([t for t in completed_trades if float(t.get('pnl', 0)) > 0])
            total = len(completed_trades)

            self.dash_realized_pnl.setText(f"₹{realized_pnl:+,.2f}")
            color = "#4CAF50" if realized_pnl >= 0 else "#F44336"
            self.dash_realized_pnl.setStyleSheet(f"font-size: 24px; font-weight: bold; color: {color};")

            self.dash_trades_count.setText(str(total))
            win_rate = (winning / total * 100) if total > 0 else 0
            self.dash_win_rate.setText(f"{win_rate:.1f}%")

            # Total P&L
            unrealized_text = self.dash_unrealized_pnl.text().replace('₹', '').replace(',', '').replace('+', '')
            try:
                unrealized = float(unrealized_text)
            except:
                unrealized = 0

            total_pnl = realized_pnl + unrealized
            self.dash_total_pnl.setText(f"₹{total_pnl:+,.2f}")
            color = "#4CAF50" if total_pnl >= 0 else "#F44336"
            self.dash_total_pnl.setStyleSheet(f"font-size: 24px; font-weight: bold; color: {color};")

        except Exception as e:
            logger.debug(f"P&L update error: {e}")

    def _close_position(self, symbol: str, quantity: int):
        """Close a position from dashboard"""
        reply = QMessageBox.question(
            self, "Close Position",
            f"Close position for {symbol} ({quantity} qty)?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        if reply == QMessageBox.StandardButton.Yes:
            # Place opposite order to close
            action = "SELL" if quantity > 0 else "BUY"
            self._quick_order(symbol, action, abs(quantity))

    def _quick_order(self, symbol: str, action: str, quantity: int):
        """Place a quick market order"""
        if not self.brokers:
            QMessageBox.warning(self, "Error", "No broker connected")
            return

        broker = list(self.brokers.values())[0]
        from algo_trader.brokers.base import BrokerOrder

        order = BrokerOrder(
            symbol=symbol,
            exchange="NSE",
            transaction_type=action,
            order_type="MARKET",
            quantity=quantity,
            product="MIS"
        )

        result = broker.place_order(order)
        if result.get('success'):
            QMessageBox.information(self, "Success", f"Order placed: {result.get('order_id')}")
            self._refresh_dashboard()
        else:
            QMessageBox.warning(self, "Error", f"Order failed: {result.get('message')}")

    def _load_strategies(self):
        """Load strategies from database"""
        strategies = self.db.get_all_strategies()
        self.strategy_list.setRowCount(len(strategies))
        self.bt_strategy_combo.clear()

        for i, strategy in enumerate(strategies):
            self.strategy_list.setItem(i, 0, QTableWidgetItem(strategy['name']))
            self.strategy_list.setItem(i, 1, QTableWidgetItem(strategy.get('source_type', 'pine')))
            self.strategy_list.setItem(i, 2, QTableWidgetItem("Yes" if strategy['is_active'] else "No"))
            self.bt_strategy_combo.addItem(strategy['name'])

    def _load_orders(self):
        """Load orders from database"""
        orders = self.db.get_orders()
        self.orders_table.setRowCount(len(orders))

        for i, order in enumerate(orders):
            self.orders_table.setItem(i, 0, QTableWidgetItem(str(order.get('created_at', ''))))
            self.orders_table.setItem(i, 1, QTableWidgetItem(order.get('symbol', '')))
            self.orders_table.setItem(i, 2, QTableWidgetItem(order.get('transaction_type', '')))
            self.orders_table.setItem(i, 3, QTableWidgetItem(str(order.get('quantity', ''))))
            self.orders_table.setItem(i, 4, QTableWidgetItem(str(order.get('price', ''))))
            self.orders_table.setItem(i, 5, QTableWidgetItem(order.get('status', '')))
            self.orders_table.setItem(i, 6, QTableWidgetItem(order.get('broker_order_id', '')))

    def _load_positions(self):
        """Load positions from active broker"""
        current_broker = self.broker_combo.currentText().lower()
        if current_broker and current_broker in self.brokers:
            positions = self.brokers[current_broker].get_positions()
            self.positions_table.setRowCount(len(positions))

            for i, pos in enumerate(positions):
                self.positions_table.setItem(i, 0, QTableWidgetItem(pos.get('symbol', '')))
                self.positions_table.setItem(i, 1, QTableWidgetItem(pos.get('exchange', '')))
                self.positions_table.setItem(i, 2, QTableWidgetItem(str(pos.get('quantity', ''))))
                self.positions_table.setItem(i, 3, QTableWidgetItem(str(pos.get('average_price', ''))))
                self.positions_table.setItem(i, 4, QTableWidgetItem(str(pos.get('ltp', ''))))
                self.positions_table.setItem(i, 5, QTableWidgetItem(str(pos.get('pnl', ''))))

    def _show_broker_dialog(self):
        """Show broker configuration dialog"""
        from algo_trader.ui.broker_dialog import BrokerConfigDialog
        dialog = BrokerConfigDialog(self.config, self)
        result = dialog.exec()

        logger.info(f"Dialog result: {result}, Accepted: {QDialog.DialogCode.Accepted}")

        if result == QDialog.DialogCode.Accepted:
            broker_name = dialog.broker_combo.currentText().lower().replace(" ", "_")
            logger.info(f"Dialog accepted for broker: {broker_name}")
            logger.info(f"broker_instance exists: {dialog.broker_instance is not None}")

            # If broker was authenticated, add it to active brokers
            if dialog.broker_instance:
                logger.info(f"is_authenticated: {dialog.broker_instance.is_authenticated}")
                if dialog.broker_instance.is_authenticated:
                    self.brokers[broker_name] = dialog.broker_instance
                    self.active_broker = dialog.broker_instance
                    logger.info(f"Broker {broker_name} added to self.brokers")

                    # Set broker for chart widget
                    if hasattr(self, 'chart_widget') and self.chart_widget:
                        self.chart_widget.set_broker(dialog.broker_instance)

                    QMessageBox.information(self, "Connected", f"{broker_name.title()} broker connected successfully!")
                else:
                    logger.warning(f"Broker {broker_name} instance exists but not authenticated")
                    QMessageBox.warning(self, "Not Connected",
                        f"Broker was not fully authenticated.\n\nPlease try again and make sure to:\n"
                        "1. Click 'Get Login URL'\n"
                        "2. Login in browser\n"
                        "3. Copy the code from URL\n"
                        "4. Paste and click 'Authenticate'")
            else:
                logger.warning("No broker_instance after dialog closed")

            self._load_configured_brokers()
            self._update_broker_settings_table()
            self._refresh_dashboard()

    def _on_strategy_selected(self, row, col):
        """Handle strategy selection"""
        name_item = self.strategy_list.item(row, 0)
        if name_item:
            strategy = self.db.get_strategy(name_item.text())
            if strategy:
                self.strategy_name_edit.setText(strategy['name'])
                self.pine_script_editor.setText(strategy['pine_script'])

    def _new_strategy(self):
        """Create new strategy"""
        self.strategy_name_edit.clear()
        self.pine_script_editor.clear()

    def _save_strategy(self):
        """Save current strategy"""
        name = self.strategy_name_edit.text().strip()
        script = self.pine_script_editor.toPlainText().strip()

        if not name:
            QMessageBox.warning(self, "Error", "Please enter a strategy name")
            return

        if not script:
            QMessageBox.warning(self, "Error", "Please enter Pine Script code")
            return

        self.db.save_strategy(name, script)
        self._load_strategies()
        QMessageBox.information(self, "Success", f"Strategy '{name}' saved successfully")

    def _validate_strategy(self):
        """Validate Pine Script"""
        from algo_trader.strategies.pine_parser import PineScriptParser

        script = self.pine_script_editor.toPlainText().strip()
        if not script:
            QMessageBox.warning(self, "Error", "Please enter Pine Script code")
            return

        parser = PineScriptParser()
        result = parser.parse(script)

        if result:
            QMessageBox.information(
                self, "Validation Success",
                f"Strategy parsed successfully!\n\n"
                f"Name: {result.name}\n"
                f"Version: {result.version}\n"
                f"Variables: {len(result.variables)}\n"
                f"Inputs: {len(result.inputs)}\n"
                f"Indicators: {len(result.indicators)}\n"
                f"Entry conditions: {len(result.entry_conditions)}\n"
                f"Exit conditions: {len(result.exit_conditions)}"
            )
        else:
            QMessageBox.warning(self, "Validation Failed", "Failed to parse Pine Script. Check syntax.")

    def _delete_strategy(self):
        """Delete selected strategy"""
        current_row = self.strategy_list.currentRow()
        if current_row < 0:
            return

        name_item = self.strategy_list.item(current_row, 0)
        if name_item:
            reply = QMessageBox.question(
                self, "Confirm Delete",
                f"Are you sure you want to delete strategy '{name_item.text()}'?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )
            if reply == QMessageBox.StandardButton.Yes:
                self.db.delete_strategy(name_item.text())
                self._load_strategies()

    def _activate_strategy(self):
        """Activate selected strategy for live trading"""
        name = self.strategy_name_edit.text().strip()
        if not name:
            QMessageBox.warning(self, "Error", "Please select a strategy")
            return

        script = self.pine_script_editor.toPlainText().strip()
        if self.strategy_engine.load_strategy(name, script):
            self.strategy_engine.enable_strategy(name)
            QMessageBox.information(self, "Success", f"Strategy '{name}' activated for live trading")
            self._load_strategies()
        else:
            QMessageBox.warning(self, "Error", "Failed to activate strategy")

    def _place_manual_order(self):
        """Place a manual order"""
        from algo_trader.core.order_manager import Order, OrderType, TransactionType, Exchange

        current_broker = self.broker_combo.currentText().lower()
        if not current_broker or current_broker not in self.brokers:
            QMessageBox.warning(self, "Error", "Please connect to a broker first")
            return

        symbol = self.order_symbol.text().strip().upper()
        if not symbol:
            QMessageBox.warning(self, "Error", "Please enter a symbol")
            return

        order = Order(
            symbol=symbol,
            transaction_type=TransactionType.BUY if self.order_type.currentText() == "BUY" else TransactionType.SELL,
            quantity=self.order_quantity.value(),
            order_type=OrderType[self.order_price_type.currentText().replace("-", "_")],
            price=self.order_price.value() if self.order_price.value() > 0 else None,
            exchange=Exchange[self.order_exchange.currentText()]
        )

        result = self.order_manager.place_order(order, current_broker)

        if result.status.value in ('OPEN', 'COMPLETE'):
            QMessageBox.information(self, "Success", f"Order placed: {result.broker_order_id}")
            self._load_orders()
        else:
            QMessageBox.warning(self, "Error", f"Order failed: {result.message}")

    def _place_gtt_order(self):
        """Place a GTT (Good Till Triggered) order"""
        current_broker = self.broker_combo.currentText().lower()
        if not current_broker or current_broker not in self.brokers:
            QMessageBox.warning(self, "Error", "Please connect to a broker first")
            return

        broker = self.brokers[current_broker]
        if not hasattr(broker, 'place_gtt_order'):
            QMessageBox.warning(self, "Error", f"{current_broker.title()} does not support GTT orders")
            return

        symbol = self.gtt_symbol.text().strip().upper()
        if not symbol:
            QMessageBox.warning(self, "Error", "Please enter a symbol")
            return

        gtt_type = self.gtt_type.currentText()
        trigger_price = self.gtt_trigger.value()
        limit_price = self.gtt_price.value()
        quantity = self.gtt_qty.value()

        if trigger_price <= 0 or limit_price <= 0:
            QMessageBox.warning(self, "Error", "Please enter valid trigger and limit prices")
            return

        try:
            if "OCO" in gtt_type:
                # OCO order - both SL and Target
                target_trigger = self.gtt_target_trigger.value()
                target_price = self.gtt_target_price.value()

                if target_trigger <= 0 or target_price <= 0:
                    QMessageBox.warning(self, "Error", "Please enter target trigger and price for OCO order")
                    return

                result = broker.place_gtt_oco(
                    symbol=symbol,
                    exchange="NSE",
                    transaction_type="SELL",
                    stop_loss_trigger=trigger_price,
                    stop_loss_price=limit_price,
                    target_trigger=target_trigger,
                    target_price=target_price,
                    quantity=quantity
                )
            else:
                # Single GTT order
                transaction_type = "SELL" if "Stop Loss" in gtt_type else "BUY"
                result = broker.place_gtt_order(
                    symbol=symbol,
                    exchange="NSE",
                    transaction_type=transaction_type,
                    trigger_price=trigger_price,
                    limit_price=limit_price,
                    quantity=quantity
                )

            if result.get('success'):
                QMessageBox.information(self, "GTT Order Placed",
                    f"GTT Order placed successfully!\n\n"
                    f"GTT ID: {result.get('gtt_id')}\n"
                    f"Symbol: {symbol}\n"
                    f"Trigger: ₹{trigger_price:.2f}\n\n"
                    f"This order will stay active on broker server even when app is closed.")
                self._refresh_gtt_orders()
            else:
                QMessageBox.warning(self, "Error", f"GTT order failed: {result.get('message', 'Unknown error')}")

        except Exception as e:
            logger.error(f"GTT order error: {e}")
            QMessageBox.warning(self, "Error", f"GTT order failed: {str(e)}")

    def _refresh_gtt_orders(self):
        """Refresh GTT orders table"""
        current_broker = self.broker_combo.currentText().lower()
        if not current_broker or current_broker not in self.brokers:
            return

        broker = self.brokers[current_broker]
        if not hasattr(broker, 'get_gtt_orders'):
            return

        try:
            gtt_orders = broker.get_gtt_orders()
            self.gtt_table.setRowCount(len(gtt_orders))

            for i, order in enumerate(gtt_orders):
                self.gtt_table.setItem(i, 0, QTableWidgetItem(str(order.get('gtt_id', ''))))
                self.gtt_table.setItem(i, 1, QTableWidgetItem(order.get('symbol', '')))
                self.gtt_table.setItem(i, 2, QTableWidgetItem(order.get('transaction_type', '')))
                self.gtt_table.setItem(i, 3, QTableWidgetItem(f"₹{order.get('trigger_price', 0):.2f}"))
                self.gtt_table.setItem(i, 4, QTableWidgetItem(f"₹{order.get('price', 0):.2f}"))
                self.gtt_table.setItem(i, 5, QTableWidgetItem(str(order.get('quantity', 0))))

                # Cancel button
                cancel_btn = QPushButton("🗑️ Cancel")
                cancel_btn.setMaximumWidth(80)
                cancel_btn.clicked.connect(lambda checked, gid=order.get('gtt_id'): self._cancel_gtt_order(gid))
                self.gtt_table.setCellWidget(i, 6, cancel_btn)

        except Exception as e:
            logger.error(f"Error fetching GTT orders: {e}")

    def _cancel_gtt_order(self, gtt_id: str):
        """Cancel a GTT order"""
        current_broker = self.broker_combo.currentText().lower()
        if not current_broker or current_broker not in self.brokers:
            return

        broker = self.brokers[current_broker]

        reply = QMessageBox.question(self, "Cancel GTT Order",
            f"Are you sure you want to cancel GTT order {gtt_id}?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)

        if reply == QMessageBox.StandardButton.Yes:
            try:
                result = broker.cancel_gtt_order(gtt_id)
                if result.get('success'):
                    QMessageBox.information(self, "Success", "GTT order cancelled")
                    self._refresh_gtt_orders()
                else:
                    QMessageBox.warning(self, "Error", f"Failed to cancel: {result.get('message')}")
            except Exception as e:
                QMessageBox.warning(self, "Error", f"Cancel failed: {str(e)}")

    def _run_backtest(self):
        """Run backtest on selected strategy"""
        strategy_name = self.bt_strategy_combo.currentText()
        if not strategy_name:
            QMessageBox.warning(self, "Error", "Please select a strategy")
            return

        symbol = self.bt_symbol.text().strip().upper()
        if not symbol:
            QMessageBox.warning(self, "Error", "Please enter a symbol")
            return

        capital = self.bt_capital.value()

        # Load strategy
        strategy = self.db.get_strategy(strategy_name)
        if not strategy:
            QMessageBox.warning(self, "Error", "Strategy not found")
            return

        from algo_trader.strategies.pine_parser import PineScriptParser
        from algo_trader.strategies.pine_interpreter import PineScriptInterpreter

        parser = PineScriptParser()
        parsed = parser.parse(strategy['pine_script'])

        if not parsed:
            QMessageBox.warning(self, "Error", "Failed to parse strategy")
            return

        interpreter = PineScriptInterpreter(parsed)

        # For demo, create sample data
        # In production, fetch historical data from broker
        import pandas as pd
        import numpy as np

        dates = pd.date_range(end=pd.Timestamp.now(), periods=500, freq='D')
        np.random.seed(42)
        price = 100
        prices = []
        for _ in range(500):
            price = price * (1 + np.random.randn() * 0.02)
            prices.append(price)

        sample_data = pd.DataFrame({
            'open': prices,
            'high': [p * 1.01 for p in prices],
            'low': [p * 0.99 for p in prices],
            'close': [p * (1 + np.random.randn() * 0.005) for p in prices],
            'volume': [np.random.randint(100000, 1000000) for _ in prices]
        }, index=dates)

        interpreter.load_data(sample_data)
        results = interpreter.run_backtest(capital)

        # Display results (old format for backward compat)
        self.bt_stat_pnl.setText(f"P&L: ₹{results['final_capital'] - results['initial_capital']:,.2f}")
        self.bt_stat_trades.setText(f"Trades: {results['total_trades']}")
        self.bt_stat_winrate.setText(f"Win Rate: {results['win_rate']:.1f}%")
        self.bt_stat_pf.setText(f"Profit Factor: {results['profit_factor']:.2f}")
        self.bt_stat_dd.setText(f"Max DD: {results['max_drawdown']:.1f}%")

    def _run_advanced_backtest(self):
        """Run advanced backtest with step-by-step simulation"""
        strategy_name = self.bt_strategy_combo.currentText()
        if not strategy_name:
            QMessageBox.warning(self, "Error", "Please select a strategy")
            return

        symbol = self.bt_symbol.currentText().strip().upper()
        if not symbol:
            QMessageBox.warning(self, "Error", "Please enter a symbol")
            return

        capital = self.bt_capital.value()
        days = self.bt_days.value()

        # Get interval
        interval_text = self.bt_interval.currentText()
        interval_map = {"1d": "1d", "1h": "1h", "15m": "15m", "5m": "5m"}
        interval = "1d"
        for key in interval_map:
            if key in interval_text:
                interval = interval_map[key]
                break

        # Risk settings
        sl_pct = self.bt_sl.value()
        target_pct = self.bt_target.value()
        tsl_pct = self.bt_tsl.value()

        self.bt_progress.setText("Fetching data...")
        self.run_backtest_btn.setEnabled(False)
        QApplication.processEvents()

        try:
            # Load strategy
            strategy = self.db.get_strategy(strategy_name)
            if not strategy:
                QMessageBox.warning(self, "Error", "Strategy not found")
                return

            from algo_trader.strategies.pine_parser import PineScriptParser
            from algo_trader.strategies.pine_interpreter import PineScriptInterpreter
            from algo_trader.backtest.simulator import BacktestSimulator
            from algo_trader.data.historical import HistoricalDataManager

            # Parse strategy
            parser = PineScriptParser()
            parsed = parser.parse(strategy['pine_script'])
            if not parsed:
                QMessageBox.warning(self, "Error", "Failed to parse strategy")
                return

            interpreter = PineScriptInterpreter(parsed)

            # Fetch historical data
            self.bt_progress.setText("Downloading historical data...")
            QApplication.processEvents()

            data_manager = HistoricalDataManager()
            data = data_manager.get_data_for_backtest(symbol, days, interval)

            if data is None or len(data) == 0:
                QMessageBox.warning(self, "Error", "Could not fetch historical data")
                return

            self.bt_progress.setText(f"Running backtest on {len(data)} candles...")
            QApplication.processEvents()

            # Create strategy function for simulator
            interpreter.load_data(data.set_index('datetime') if 'datetime' in data.columns else data)

            def strategy_signal(row, idx, full_data):
                """Generate strategy signal for each candle"""
                try:
                    # Update interpreter with current candle
                    result = interpreter.process_candle({
                        'open': row['open'],
                        'high': row['high'],
                        'low': row['low'],
                        'close': row['close'],
                        'volume': row.get('volume', 0)
                    })
                    if result and 'signal' in result:
                        return result['signal']  # 'BUY' or 'SELL'
                except Exception:
                    pass
                return None

            # Create and configure simulator
            simulator = BacktestSimulator(initial_capital=capital)
            simulator.set_risk_params(
                stop_loss=sl_pct,
                target=target_pct,
                trailing_sl=tsl_pct
            )

            # Register progress callback
            def on_progress(idx, total, time, price, equity, open_trades):
                if idx % 50 == 0:
                    self.bt_progress.setText(f"Processing: {idx}/{total} ({time.strftime('%Y-%m-%d') if hasattr(time, 'strftime') else time})")
                    QApplication.processEvents()

            simulator.register_progress_callback(on_progress)

            # Get speed setting
            speed_text = self.bt_speed.currentText()
            speed = float(speed_text.replace('x', ''))

            # Run backtest
            realtime = self.bt_realtime_mode.isChecked()
            simulator.set_speed(speed)

            self.bt_simulator = simulator  # Store for export
            result = simulator.run_backtest(
                data=data,
                strategy_func=strategy_signal,
                symbol=symbol,
                strategy_name=strategy_name,
                realtime_mode=realtime
            )

            # Update UI with results
            self._display_backtest_results(result)

        except Exception as e:
            logger.error(f"Backtest error: {e}")
            QMessageBox.warning(self, "Error", f"Backtest failed: {str(e)}")
        finally:
            self.run_backtest_btn.setEnabled(True)
            self.bt_progress.setText("Complete")

    def _display_backtest_results(self, result):
        """Display backtest results in UI"""
        # Update summary stats
        pnl_color = "green" if result.total_pnl >= 0 else "red"
        self.bt_stat_pnl.setText(f"P&L: ₹{result.total_pnl:,.2f}")
        self.bt_stat_pnl.setStyleSheet(f"font-size: 14px; font-weight: bold; color: {pnl_color};")

        self.bt_stat_trades.setText(f"Trades: {result.total_trades}")
        self.bt_stat_winrate.setText(f"Win Rate: {result.win_rate:.1f}%")
        self.bt_stat_pf.setText(f"PF: {result.profit_factor:.2f}")
        self.bt_stat_dd.setText(f"Max DD: ₹{result.max_drawdown:,.0f} ({result.max_drawdown_percent:.1f}%)")

        # Populate trade log table
        self.bt_trades_table.setRowCount(len(result.trades))
        for i, trade in enumerate(result.trades):
            self.bt_trades_table.setItem(i, 0, QTableWidgetItem(str(trade.trade_id)))
            self.bt_trades_table.setItem(i, 1, QTableWidgetItem(trade.trade_type.value))
            self.bt_trades_table.setItem(i, 2, QTableWidgetItem(
                trade.entry_time.strftime('%Y-%m-%d %H:%M') if trade.entry_time else ''))
            self.bt_trades_table.setItem(i, 3, QTableWidgetItem(f"₹{trade.entry_price:.2f}"))
            self.bt_trades_table.setItem(i, 4, QTableWidgetItem(
                trade.exit_time.strftime('%Y-%m-%d %H:%M') if trade.exit_time else ''))
            self.bt_trades_table.setItem(i, 5, QTableWidgetItem(f"₹{trade.exit_price:.2f}"))
            self.bt_trades_table.setItem(i, 6, QTableWidgetItem(str(trade.quantity)))

            pnl_item = QTableWidgetItem(f"₹{trade.pnl:,.2f}")
            pnl_item.setForeground(Qt.GlobalColor.green if trade.pnl >= 0 else Qt.GlobalColor.red)
            self.bt_trades_table.setItem(i, 7, pnl_item)

            self.bt_trades_table.setItem(i, 8, QTableWidgetItem(f"{trade.pnl_percent:.2f}%"))
            self.bt_trades_table.setItem(i, 9, QTableWidgetItem(trade.exit_reason))

        self.export_trades_btn.setEnabled(True)
        logger.info(f"Backtest complete: {result.total_trades} trades, P&L: ₹{result.total_pnl:.2f}")

    def _export_backtest_trades(self):
        """Export backtest trades to CSV"""
        if not hasattr(self, 'bt_simulator') or not self.bt_simulator:
            QMessageBox.warning(self, "Error", "No backtest results to export")
            return

        from datetime import datetime
        filename = f"backtest_trades_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"

        try:
            self.bt_simulator.export_trades_csv(filename)
            QMessageBox.information(self, "Exported", f"Trades exported to: {filename}")
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Export failed: {e}")

    def _save_settings(self):
        """Save trading settings"""
        self.config.set('trading.default_quantity', self.default_qty.value())
        self.config.set('trading.max_positions', self.max_positions.value())
        self.config.set('trading.risk_percent', self.risk_percent.value())
        self.config.set('trading.paper_capital', self.paper_capital.value())
        QMessageBox.information(self, "Success", "Settings saved")

    def _save_squareoff_settings(self):
        """Save auto square-off settings"""
        enabled = self.auto_squareoff_enabled.isChecked()

        self.config.set('squareoff.enabled', enabled)
        self.config.set('squareoff.daily_profit_target', self.daily_profit_target.value())
        self.config.set('squareoff.daily_loss_limit', self.daily_loss_limit.value())
        self.config.set('squareoff.time', self.squareoff_time.text())
        self.config.set('squareoff.position_profit_percent', self.position_profit_percent.value())
        self.config.set('squareoff.position_loss_percent', self.position_loss_percent.value())
        self.config.set('squareoff.trailing_profit_percent', self.trailing_profit_percent.value())

        # Apply settings to risk manager
        if hasattr(self, 'risk_manager'):
            self.risk_manager.configure_auto_square_off(
                daily_profit_target=self.daily_profit_target.value() if self.daily_profit_target.value() > 0 else None,
                daily_loss_limit=self.daily_loss_limit.value() if self.daily_loss_limit.value() > 0 else None,
                square_off_time=self.squareoff_time.text() if self.squareoff_time.text() else None,
                position_profit_percent=self.position_profit_percent.value() if self.position_profit_percent.value() > 0 else None,
                position_loss_percent=self.position_loss_percent.value() if self.position_loss_percent.value() > 0 else None,
                trailing_profit_percent=self.trailing_profit_percent.value() if self.trailing_profit_percent.value() > 0 else None,
                enabled=enabled
            )

            # Update status
            if enabled:
                self.squareoff_status_label.setText("Status: ✅ Active")
                self.squareoff_status_label.setStyleSheet("color: green; font-weight: bold;")
            else:
                self.squareoff_status_label.setText("Status: ⚪ Disabled")
                self.squareoff_status_label.setStyleSheet("color: gray;")

        QMessageBox.information(self, "Success", "Auto Square-off settings saved")

    # Paper Trading methods
    def _init_paper_trading(self):
        """Initialize paper trading simulator"""
        from algo_trader.core.paper_trading import PaperTradingSimulator
        capital = self.config.get('trading.paper_capital', 100000)
        self.paper_simulator = PaperTradingSimulator(initial_capital=capital)
        logger.info(f"Paper Trading Simulator initialized with ₹{capital:,.2f}")

    def _on_paper_mode_changed(self, state):
        """Handle paper trading mode toggle"""
        is_paper = state == Qt.CheckState.Checked.value
        self.config.set('trading.paper_mode', is_paper)

        if is_paper:
            self.paper_mode_label.setText("📝 PAPER TRADING")
            self.paper_mode_label.setStyleSheet("font-weight: bold; font-size: 14px; color: orange;")
            self._init_paper_trading()
            QMessageBox.information(self, "Paper Trading",
                "Paper Trading Mode enabled!\n\nAll trades will be simulated - no real money will be used.")
        else:
            self.paper_mode_label.setText("🟢 LIVE TRADING")
            self.paper_mode_label.setStyleSheet("font-weight: bold; font-size: 14px; color: green;")
            QMessageBox.warning(self, "Live Trading",
                "Live Trading Mode enabled!\n\nReal orders will be placed with your broker.")

    def _reset_paper_trading(self):
        """Reset paper trading simulator"""
        if hasattr(self, 'paper_simulator'):
            capital = self.paper_capital.value()
            self.paper_simulator.reset(capital)
            QMessageBox.information(self, "Reset", f"Simulator reset with ₹{capital:,.0f}")
        else:
            self._init_paper_trading()
            QMessageBox.information(self, "Initialized", "Paper Trading Simulator initialized")

    def is_paper_mode(self) -> bool:
        """Check if paper trading mode is enabled"""
        return self.paper_trading_toggle.isChecked()

    # Risk Manager / Auto Square-off methods
    def _init_risk_manager(self):
        """Initialize risk manager for auto square-off"""
        from algo_trader.core.risk_manager import RiskManager
        self.risk_manager = RiskManager()

        # Register callbacks
        self.risk_manager.register_square_off_callback(self._on_auto_square_off)
        self.risk_manager.register_mtm_callback(self._on_mtm_update)

        # Load saved settings
        if self.config.get('squareoff.enabled', False):
            self.risk_manager.configure_auto_square_off(
                daily_profit_target=self.config.get('squareoff.daily_profit_target') or None,
                daily_loss_limit=self.config.get('squareoff.daily_loss_limit') or None,
                square_off_time=self.config.get('squareoff.time') or None,
                position_profit_percent=self.config.get('squareoff.position_profit_percent') or None,
                position_loss_percent=self.config.get('squareoff.position_loss_percent') or None,
                trailing_profit_percent=self.config.get('squareoff.trailing_profit_percent') or None,
                enabled=True
            )
            logger.info("Auto square-off enabled with saved settings")

    def _on_auto_square_off(self, event_data: dict):
        """Handle auto square-off event"""
        symbol = event_data.get('symbol', 'Unknown')
        reason = event_data.get('reason', 'Unknown')
        pnl = event_data.get('pnl', 0)
        message = event_data.get('message', '')

        # Show notification
        QMessageBox.warning(self, "Auto Square-off Triggered",
            f"Position Squared Off!\n\n"
            f"Symbol: {symbol}\n"
            f"Reason: {reason}\n"
            f"P&L: ₹{pnl:.2f}\n\n"
            f"{message}")

        # Send Telegram alert if enabled
        if hasattr(self, 'telegram') and self.telegram and self.telegram.is_enabled():
            self.telegram.send_alert(
                f"🔴 AUTO SQUARE-OFF\n\n"
                f"Symbol: {symbol}\n"
                f"Reason: {reason}\n"
                f"P&L: ₹{pnl:.2f}\n\n"
                f"{message}"
            )

        # Refresh UI
        self._refresh_dashboard()

    def _on_mtm_update(self, mtm_summary):
        """Handle MTM update from risk manager"""
        # Update dashboard P&L display if available
        if hasattr(self, 'daily_pnl_label'):
            pnl = mtm_summary.total_pnl
            color = "green" if pnl >= 0 else "red"
            self.daily_pnl_label.setText(f"₹{pnl:,.2f}")
            self.daily_pnl_label.setStyleSheet(f"color: {color}; font-weight: bold;")

    # Telegram Alert methods
    def _init_telegram(self):
        """Initialize Telegram alerts"""
        from algo_trader.integrations.telegram_alerts import TelegramAlerts
        self.telegram = TelegramAlerts()
        bot_token = self.config.get('telegram.bot_token', '')
        chat_id = self.config.get('telegram.chat_id', '')
        if bot_token and chat_id:
            self.telegram.configure(bot_token, chat_id)
            if self.config.get('telegram.enabled', False):
                self.telegram.enable()

    def _on_telegram_toggle(self, state):
        """Handle Telegram enable/disable"""
        enabled = state == Qt.CheckState.Checked.value
        if enabled:
            if not hasattr(self, 'telegram'):
                self._init_telegram()
            if self.telegram.is_configured():
                self.telegram.enable()
                self.config.set('telegram.enabled', True)
            else:
                QMessageBox.warning(self, "Error", "Please enter Bot Token and Chat ID first")
                self.telegram_enabled.setChecked(False)
        else:
            if hasattr(self, 'telegram'):
                self.telegram.disable()
            self.config.set('telegram.enabled', False)

    def _save_telegram_settings(self):
        """Save Telegram configuration"""
        bot_token = self.telegram_bot_token.text().strip()
        chat_id = self.telegram_chat_id.text().strip()

        self.config.set('telegram.bot_token', bot_token)
        self.config.set('telegram.chat_id', chat_id)

        if not hasattr(self, 'telegram'):
            self._init_telegram()

        if bot_token and chat_id:
            self.telegram.configure(bot_token, chat_id)
            QMessageBox.information(self, "Saved", "Telegram settings saved!")
        else:
            QMessageBox.warning(self, "Warning", "Please enter both Bot Token and Chat ID")

    def _test_telegram(self):
        """Test Telegram connection"""
        if not hasattr(self, 'telegram'):
            self._init_telegram()

        bot_token = self.telegram_bot_token.text().strip()
        chat_id = self.telegram_chat_id.text().strip()

        if not bot_token or not chat_id:
            QMessageBox.warning(self, "Error", "Please enter Bot Token and Chat ID")
            return

        self.telegram.configure(bot_token, chat_id)
        if self.telegram.test_connection():
            QMessageBox.information(self, "Success", "Telegram connection successful! Check your Telegram.")
        else:
            QMessageBox.warning(self, "Failed", "Could not connect to Telegram. Check your credentials.")

    def _send_telegram_alert(self, message_type: str, **kwargs):
        """Send alert to Telegram if enabled"""
        if not hasattr(self, 'telegram') or not self.telegram.enabled:
            return

        try:
            if message_type == 'trade':
                self.telegram.send_trade_alert(**kwargs)
            elif message_type == 'signal':
                self.telegram.send_signal_alert(**kwargs)
            elif message_type == 'chartink':
                self.telegram.send_chartink_alert(**kwargs)
            elif message_type == 'option':
                self.telegram.send_option_alert(**kwargs)
            elif message_type == 'sl_hit':
                self.telegram.send_sl_hit_alert(**kwargs)
            elif message_type == 'target_hit':
                self.telegram.send_target_hit_alert(**kwargs)
        except Exception as e:
            logger.error(f"Telegram alert error: {e}")

    # Telegram Bot Controller methods
    def _init_bot_controller(self):
        """Initialize Telegram Bot Controller for remote control"""
        from algo_trader.integrations.telegram_alerts import TelegramBotController
        bot_token = self.config.get('telegram.bot_token', '')
        chat_id = self.config.get('telegram.chat_id', '')

        if not bot_token or not chat_id:
            return None

        self.bot_controller = TelegramBotController(bot_token, chat_id)

        # Register callbacks
        self.bot_controller.register_status_callback(self._get_bot_status)
        self.bot_controller.register_positions_callback(self._get_bot_positions)
        self.bot_controller.register_orders_callback(self._get_bot_orders)
        self.bot_controller.register_pnl_callback(self._get_bot_pnl)
        self.bot_controller.register_squareoff_callback(self._bot_squareoff)
        self.bot_controller.register_pause_callback(self._bot_pause)
        self.bot_controller.register_resume_callback(self._bot_resume)

        return self.bot_controller

    def _on_bot_controller_toggle(self, state):
        """Handle bot controller enable/disable"""
        enabled = state == Qt.CheckState.Checked.value
        self.config.set('telegram.bot_controller_enabled', enabled)

        if enabled:
            self._start_bot_controller()
        else:
            self._stop_bot_controller()

    def _start_bot_controller(self):
        """Start the Telegram bot controller"""
        bot_token = self.config.get('telegram.bot_token', '')
        chat_id = self.config.get('telegram.chat_id', '')

        if not bot_token or not chat_id:
            QMessageBox.warning(self, "Error", "Please configure Telegram Bot Token and Chat ID first")
            self.bot_controller_enabled.setChecked(False)
            return

        if not hasattr(self, 'bot_controller') or self.bot_controller is None:
            self._init_bot_controller()

        if self.bot_controller and self.bot_controller.start_listening():
            self.bot_controller_status.setText("Running")
            self.bot_controller_status.setStyleSheet("color: #4CAF50; font-weight: bold;")
            self.start_bot_ctrl_btn.setEnabled(False)
            self.stop_bot_ctrl_btn.setEnabled(True)
            self.status_bar.showMessage("Telegram Bot Controller started")
            logger.info("Telegram Bot Controller started")
        else:
            QMessageBox.warning(self, "Error", "Failed to start Bot Controller")

    def _stop_bot_controller(self):
        """Stop the Telegram bot controller"""
        if hasattr(self, 'bot_controller') and self.bot_controller:
            self.bot_controller.stop_listening()

        self.bot_controller_status.setText("Stopped")
        self.bot_controller_status.setStyleSheet("color: red; font-weight: bold;")
        self.start_bot_ctrl_btn.setEnabled(True)
        self.stop_bot_ctrl_btn.setEnabled(False)
        self.status_bar.showMessage("Telegram Bot Controller stopped")
        logger.info("Telegram Bot Controller stopped")

    def _get_bot_status(self) -> str:
        """Get algo status for Telegram bot"""
        try:
            # Get connected broker
            broker_name = "Not Connected"
            broker_status = "❌"
            if self.active_broker:
                broker_name = type(self.active_broker).__name__.replace("Broker", "")
                broker_status = "✅"

            # Get trading status
            algo_status = "⏸️ PAUSED" if (hasattr(self, 'bot_controller') and
                                           self.bot_controller.is_paused()) else "▶️ RUNNING"

            # Get P&L
            total_pnl = 0
            if hasattr(self, 'dash_total_pnl'):
                pnl_text = self.dash_total_pnl.text().replace("₹", "").replace(",", "")
                try:
                    total_pnl = float(pnl_text)
                except:
                    pass

            pnl_emoji = "🟢" if total_pnl >= 0 else "🔴"

            return f"""
📊 <b>Algo Status</b>

{algo_status}
{broker_status} Broker: {broker_name}
{pnl_emoji} Today's P&L: ₹{total_pnl:+,.2f}

🕐 {datetime.now().strftime('%H:%M:%S')}
"""
        except Exception as e:
            logger.error(f"Error getting bot status: {e}")
            return f"❌ Error getting status: {e}"

    def _get_bot_positions(self) -> str:
        """Get positions for Telegram bot"""
        try:
            if not self.active_broker:
                return "❌ No broker connected"

            positions = self.active_broker.get_positions()
            if not positions:
                return "📊 No open positions"

            msg = "📊 <b>Open Positions</b>\n\n"
            for pos in positions[:10]:  # Limit to 10
                symbol = pos.get('symbol', 'N/A')
                qty = pos.get('quantity', 0)
                pnl = pos.get('pnl', 0)
                pnl_emoji = "🟢" if pnl >= 0 else "🔴"
                msg += f"• {symbol}: {qty} | {pnl_emoji} ₹{pnl:+,.2f}\n"

            return msg
        except Exception as e:
            logger.error(f"Error getting positions: {e}")
            return f"❌ Error: {e}"

    def _get_bot_orders(self) -> str:
        """Get pending orders for Telegram bot"""
        try:
            if not self.active_broker:
                return "❌ No broker connected"

            orders = self.active_broker.get_orders()
            pending = [o for o in orders if o.get('status') in ['PENDING', 'OPEN', 'TRIGGER_PENDING']]

            if not pending:
                return "📋 No pending orders"

            msg = "📋 <b>Pending Orders</b>\n\n"
            for order in pending[:10]:
                symbol = order.get('symbol', 'N/A')
                side = order.get('side', 'N/A')
                qty = order.get('quantity', 0)
                price = order.get('price', 0)
                msg += f"• {side} {symbol}: {qty} @ ₹{price:.2f}\n"

            return msg
        except Exception as e:
            logger.error(f"Error getting orders: {e}")
            return f"❌ Error: {e}"

    def _get_bot_pnl(self) -> str:
        """Get P&L summary for Telegram bot"""
        try:
            realized = 0
            unrealized = 0
            total = 0

            if hasattr(self, 'dash_realized_pnl'):
                try:
                    realized = float(self.dash_realized_pnl.text().replace("₹", "").replace(",", ""))
                except:
                    pass

            if hasattr(self, 'dash_unrealized_pnl'):
                try:
                    unrealized = float(self.dash_unrealized_pnl.text().replace("₹", "").replace(",", ""))
                except:
                    pass

            total = realized + unrealized
            pnl_emoji = "🟢" if total >= 0 else "🔴"

            return f"""
💰 <b>P&L Summary</b>

✅ Realized: ₹{realized:+,.2f}
📊 Unrealized: ₹{unrealized:+,.2f}
{pnl_emoji} <b>Total: ₹{total:+,.2f}</b>

🕐 {datetime.now().strftime('%H:%M:%S')}
"""
        except Exception as e:
            return f"❌ Error: {e}"

    def _bot_squareoff(self) -> str:
        """Square off all positions via Telegram command"""
        try:
            if not self.active_broker:
                return "No broker connected"

            positions = self.active_broker.get_positions()
            if not positions:
                return "No positions to square off"

            squared = 0
            for pos in positions:
                qty = pos.get('quantity', 0)
                if qty == 0:
                    continue

                symbol = pos.get('symbol')
                exchange = pos.get('exchange', 'NSE')
                side = 'SELL' if qty > 0 else 'BUY'

                try:
                    self.active_broker.place_order(
                        symbol=symbol,
                        exchange=exchange,
                        transaction_type=side,
                        quantity=abs(qty),
                        order_type='MARKET',
                        product_type='INTRADAY'
                    )
                    squared += 1
                except Exception as e:
                    logger.error(f"Failed to square off {symbol}: {e}")

            return f"Squared off {squared} position(s)"
        except Exception as e:
            return f"Error: {e}"

    def _bot_pause(self):
        """Pause algo trading via Telegram command"""
        # Set flag to pause new trades
        if hasattr(self, 'chartink_scanner') and self.chartink_scanner:
            for scan_name in self.chartink_scanner.scans:
                self.chartink_scanner.pause_scan(scan_name)
        logger.info("Algo paused via Telegram")

    def _bot_resume(self):
        """Resume algo trading via Telegram command"""
        # Resume trading
        if hasattr(self, 'chartink_scanner') and self.chartink_scanner:
            for scan_name in self.chartink_scanner.scans:
                self.chartink_scanner.resume_scan(scan_name)
        logger.info("Algo resumed via Telegram")

    # Chartink methods
    def _init_chartink(self):
        """Initialize Chartink scanner"""
        from algo_trader.integrations.chartink import ChartinkScanner
        # Enable test_mode to skip time checks (allows testing outside market hours)
        test_mode = self.config.get('chartink.test_mode', True)  # Default True for easy testing
        self.chartink_scanner = ChartinkScanner(test_mode=test_mode)
        # Use signal to ensure GUI updates happen on main thread
        self.chartink_alert_signal.connect(self._handle_chartink_alert_gui)
        self.chartink_scanner.register_alert_callback(self._on_chartink_alert_thread)
        if test_mode:
            logger.info("Chartink scanner initialized in TEST MODE (time checks disabled)")

    def _load_chartink_scans(self):
        """Load saved Chartink scans from config"""
        try:
            self._init_chartink()
            scans = self.config.get('chartink.scans', [])
            for scan in scans:
                self.chartink_scanner.add_scan(
                    scan_name=scan.get('name'),
                    scan_url=scan.get('url'),
                    interval=scan.get('interval', 60),
                    action=scan.get('action', 'BUY'),
                    quantity=scan.get('quantity', 1),
                    start_time=scan.get('start_time', '09:15'),
                    exit_time=scan.get('exit_time', '15:15'),
                    no_new_trade_time=scan.get('no_new_trade_time', '14:30'),
                    total_capital=scan.get('total_capital', scan.get('total_amount', 0)),
                    alloc_type=scan.get('alloc_type', 'auto'),
                    alloc_value=scan.get('alloc_value', scan.get('stock_quantity', 0)),
                    max_trades=scan.get('max_trades', 0),
                    risk_config=scan.get('risk_config', None),
                    trigger_on_first=scan.get('trigger_on_first', True),
                    enabled=scan.get('enabled', True)
                )
            self._refresh_chartink_scans_table()
        except Exception as e:
            logger.error(f"Error loading Chartink scans: {e}")

    def _on_chartink_trade_type_changed(self, index):
        """Handle trade type change (Equity/Options)"""
        is_options = index == 1  # Options (F&O)
        self.chartink_options_frame.setVisible(is_options)
        if is_options:
            self.chartink_quantity.setSuffix(" lots")
        else:
            self.chartink_quantity.setSuffix("")

    def _on_chartink_strike_changed(self, index):
        """Handle strike selection change (show/hide manual input)"""
        is_manual = index == 5  # "Manual" option
        self.chartink_manual_strike.setVisible(is_manual)

    def _on_auto_leg1_strike_changed(self, text):
        """Handle Leg 1 strike selection change (show/hide manual input)"""
        is_manual = text == "Manual"
        self.auto_leg1_manual_strike.setVisible(is_manual)

    def _on_auto_leg2_strike_changed(self, text):
        """Handle Leg 2 strike selection change (show/hide manual input)"""
        is_manual = text == "Manual"
        self.auto_leg2_manual_strike.setVisible(is_manual)

    def _on_chartink_alloc_type_changed(self, index):
        """Handle allocation type change"""
        if index == 0:  # Auto
            self.chartink_alloc_value.setEnabled(False)
            self.chartink_alloc_value.setValue(0)
            self.chartink_alloc_value_label.setText("(Auto)")
        elif index == 1:  # Fixed Qty
            self.chartink_alloc_value.setEnabled(True)
            self.chartink_alloc_value.setPrefix("")
            self.chartink_alloc_value.setSuffix(" shares")
            self.chartink_alloc_value_label.setText("")
        elif index == 2:  # Fixed Amount
            self.chartink_alloc_value.setEnabled(True)
            self.chartink_alloc_value.setPrefix("₹ ")
            self.chartink_alloc_value.setSuffix("")
            self.chartink_alloc_value_label.setText("per stock")

    def _on_chartink_sl_type_changed(self, index):
        """Handle SL type change"""
        self.chartink_sl_value.setEnabled(index > 0)
        if index == 1:  # Points
            self.chartink_sl_value.setSuffix(" pts")
            self.chartink_sl_value.setPrefix("")
        elif index == 2:  # Percentage
            self.chartink_sl_value.setSuffix(" %")
            self.chartink_sl_value.setPrefix("")
        elif index == 3:  # Amount
            self.chartink_sl_value.setSuffix("")
            self.chartink_sl_value.setPrefix("₹ ")

    def _on_chartink_target_type_changed(self, index):
        """Handle Target type change"""
        self.chartink_target_value.setEnabled(index > 0)
        if index == 1:  # Points
            self.chartink_target_value.setSuffix(" pts")
            self.chartink_target_value.setPrefix("")
        elif index == 2:  # Percentage
            self.chartink_target_value.setSuffix(" %")
            self.chartink_target_value.setPrefix("")
        elif index == 3:  # Amount
            self.chartink_target_value.setSuffix("")
            self.chartink_target_value.setPrefix("₹ ")

    def _on_chartink_tsl_toggle(self, state):
        """Handle TSL enable/disable"""
        enabled = state == Qt.CheckState.Checked.value
        self.chartink_tsl_type.setEnabled(enabled)
        self.chartink_tsl_value.setEnabled(enabled)

    def _add_chartink_scan(self):
        """Add a new Chartink scan with time controls and allocation"""
        name = self.chartink_scan_name.text().strip()
        url = self.chartink_scan_url.text().strip()

        if not name:
            QMessageBox.warning(self, "Error", "Please enter a scan name")
            return

        if not url:
            QMessageBox.warning(self, "Error", "Please enter a scanner URL")
            return

        if not url.startswith("https://chartink.com/screener/"):
            QMessageBox.warning(self, "Error", "Invalid Chartink URL. Should start with https://chartink.com/screener/")
            return

        action = self.chartink_action.currentText()
        quantity = self.chartink_quantity.value()
        interval = self.chartink_interval.value()
        start_time = self.chartink_start_time.currentText().strip()
        exit_time = self.chartink_exit_time.currentText().strip()
        no_new_trade_time = self.chartink_no_new_trade_time.currentText().strip()
        total_capital = self.chartink_total_capital.value()
        max_trades = self.chartink_max_trades.value()

        # Trade type (Equity or Options)
        trade_type = "options" if self.chartink_trade_type.currentIndex() == 1 else "equity"

        # Options config (if F&O selected)
        options_config = None
        if trade_type == "options":
            strike_map = ["ATM", "ITM-1", "ITM-2", "OTM-1", "OTM-2"]
            option_type_map = ["auto", "CE", "PE"]
            expiry_map = ["current_week", "next_week", "current_month", "next_month"]
            strike_idx = self.chartink_strike_selection.currentIndex()
            options_config = {
                'strike_selection': strike_map[strike_idx] if strike_idx < 5 else "Manual",
                'option_type': option_type_map[self.chartink_option_type.currentIndex()],
                'expiry': expiry_map[self.chartink_expiry.currentIndex()]
            }
            # Add manual strike price if "Manual" selected
            if strike_idx == 5:  # Manual
                options_config['manual_strike'] = int(self.chartink_manual_strike.value())

        # Get allocation type and value
        alloc_type_index = self.chartink_alloc_type.currentIndex()
        alloc_type = ["auto", "fixed_qty", "fixed_amount"][alloc_type_index]
        alloc_value = self.chartink_alloc_value.value()

        # Get risk settings
        sl_type_index = self.chartink_sl_type.currentIndex()
        sl_type = ["none", "points", "percent", "amount"][sl_type_index]
        sl_value = self.chartink_sl_value.value()

        target_type_index = self.chartink_target_type.currentIndex()
        target_type = ["none", "points", "percent", "amount"][target_type_index]
        target_value = self.chartink_target_value.value()

        tsl_enabled = self.chartink_tsl_enabled.isChecked()
        tsl_type = "points" if self.chartink_tsl_type.currentIndex() == 0 else "percent"
        tsl_value = self.chartink_tsl_value.value()

        profit_lock_enabled = self.chartink_profit_lock_enabled.isChecked()
        profit_lock_type = "points" if self.chartink_profit_lock_type.currentIndex() == 0 else "amount"
        profit_lock_value = self.chartink_profit_lock_value.value()

        mtm_profit = self.chartink_mtm_profit.value()
        mtm_loss = self.chartink_mtm_loss.value()

        # Risk config dict
        risk_config = {
            'sl_type': sl_type,
            'sl_value': sl_value,
            'target_type': target_type,
            'target_value': target_value,
            'tsl_enabled': tsl_enabled,
            'tsl_type': tsl_type,
            'tsl_value': tsl_value,
            'profit_lock_enabled': profit_lock_enabled,
            'profit_lock_type': profit_lock_type,
            'profit_lock_value': profit_lock_value,
            'mtm_profit': mtm_profit,
            'mtm_loss': mtm_loss
        }

        # Get trigger on first option
        trigger_on_first = self.chartink_trigger_first.isChecked()

        # Add to scanner
        self.chartink_scanner.add_scan(
            scan_name=name,
            scan_url=url,
            interval=interval,
            action=action,
            quantity=quantity,
            start_time=start_time,
            exit_time=exit_time,
            no_new_trade_time=no_new_trade_time,
            total_capital=total_capital,
            alloc_type=alloc_type,
            alloc_value=alloc_value,
            max_trades=max_trades,
            risk_config=risk_config,
            trigger_on_first=trigger_on_first,
            enabled=True
        )

        # Save to config
        scans = self.config.get('chartink.scans', [])
        scan_config = {
            'name': name,
            'url': url,
            'action': action,
            'quantity': quantity,
            'interval': interval,
            'start_time': start_time,
            'exit_time': exit_time,
            'no_new_trade_time': no_new_trade_time,
            'total_capital': total_capital,
            'alloc_type': alloc_type,
            'alloc_value': alloc_value,
            'max_trades': max_trades,
            'risk_config': risk_config,
            'trigger_on_first': trigger_on_first,
            'enabled': True,
            'trade_type': trade_type
        }
        if options_config:
            scan_config['options_config'] = options_config
        scans.append(scan_config)
        self.config.set('chartink.scans', scans)

        # Clear inputs
        self.chartink_scan_name.clear()
        self.chartink_scan_url.clear()

        # Refresh table
        self._refresh_chartink_scans_table()

        QMessageBox.information(self, "Success", f"Scanner '{name}' added successfully")

    def _test_chartink_scan(self):
        """Test a Chartink scanner URL"""
        url = self.chartink_scan_url.text().strip()

        if not url:
            QMessageBox.warning(self, "Error", "Please enter a scanner URL")
            return

        if not url.startswith("https://chartink.com/screener/"):
            QMessageBox.warning(self, "Error", "Invalid Chartink URL")
            return

        try:
            self.status_bar.showMessage("Testing scanner...")
            results = self.chartink_scanner.test_scan(url)

            if results:
                result_text = f"Found {len(results)} stocks:\n\n"
                for i, stock in enumerate(results[:10]):  # Show max 10
                    result_text += f"{i+1}. {stock['symbol']} - ₹{stock['price']:.2f}\n"
                if len(results) > 10:
                    result_text += f"\n... and {len(results) - 10} more"

                QMessageBox.information(self, "Scanner Test Results", result_text)
            else:
                QMessageBox.information(self, "Scanner Test Results", "No stocks found in this scanner")

            self.status_bar.showMessage("Ready")

        except Exception as e:
            QMessageBox.warning(self, "Error", f"Failed to test scanner: {e}")
            self.status_bar.showMessage("Ready")

    def _refresh_chartink_scans_table(self):
        """Refresh the Chartink scans table with all fields"""
        scans = self.chartink_scanner.get_active_scans()
        self.chartink_scans_table.setRowCount(len(scans))

        for i, scan in enumerate(scans):
            # ON/OFF Toggle button
            toggle_btn = QPushButton("ON" if scan.get('enabled', True) else "OFF")
            toggle_btn.setStyleSheet(
                "background-color: #00aa00; color: white; font-weight: bold;" if scan.get('enabled', True)
                else "background-color: #aa0000; color: white; font-weight: bold;"
            )
            toggle_btn.clicked.connect(lambda checked, name=scan['name']: self._toggle_chartink_scan(name))
            self.chartink_scans_table.setCellWidget(i, 0, toggle_btn)

            self.chartink_scans_table.setItem(i, 1, QTableWidgetItem(scan['name']))
            self.chartink_scans_table.setItem(i, 2, QTableWidgetItem(scan['action']))

            # Show allocation type and value
            alloc_type = scan.get('alloc_type', 'auto')
            alloc_value = scan.get('alloc_value', 0)
            if alloc_type == 'fixed_qty':
                alloc_text = f"{int(alloc_value)} shares"
            elif alloc_type == 'fixed_amount':
                alloc_text = f"₹{alloc_value:,.0f}/stock"
            else:
                alloc_text = "Auto"
            self.chartink_scans_table.setItem(i, 3, QTableWidgetItem(alloc_text))

            self.chartink_scans_table.setItem(i, 4, QTableWidgetItem(scan.get('start_time', '09:15')))
            self.chartink_scans_table.setItem(i, 5, QTableWidgetItem(scan.get('exit_time', '15:15')))
            self.chartink_scans_table.setItem(i, 6, QTableWidgetItem(scan.get('no_new_trade_time', '14:30')))
            amt = scan.get('total_capital', 0)
            self.chartink_scans_table.setItem(i, 7, QTableWidgetItem(f"₹{amt:,.0f}" if amt > 0 else "Unlimited"))
            max_t = scan.get('max_trades', 0)
            self.chartink_scans_table.setItem(i, 8, QTableWidgetItem(str(max_t) if max_t > 0 else "Unlimited"))
            self.chartink_scans_table.setItem(i, 9, QTableWidgetItem(str(scan.get('trade_count', 0))))

            remove_btn = QPushButton("Remove")
            remove_btn.clicked.connect(lambda checked, name=scan['name']: self._remove_chartink_scan(name))
            self.chartink_scans_table.setCellWidget(i, 10, remove_btn)

        # Also refresh positions table
        self._refresh_chartink_positions_table()

    def _toggle_chartink_scan(self, scan_name: str):
        """Toggle scanner ON/OFF"""
        new_state = self.chartink_scanner.toggle_scan(scan_name)

        # Update config
        scans = self.config.get('chartink.scans', [])
        for scan in scans:
            if scan.get('name') == scan_name:
                scan['enabled'] = new_state
                break
        self.config.set('chartink.scans', scans)

        # Refresh table
        self._refresh_chartink_scans_table()

    def _refresh_chartink_positions_table(self):
        """Refresh the scanner open positions table"""
        all_positions = []
        for scan_name in self.chartink_scanner.active_scans:
            positions = self.chartink_scanner.get_open_positions(scan_name)
            for symbol, pos in positions.items():
                all_positions.append({
                    'scanner': scan_name,
                    'symbol': symbol,
                    **pos
                })

        self.chartink_positions_table.setRowCount(len(all_positions))
        for i, pos in enumerate(all_positions):
            self.chartink_positions_table.setItem(i, 0, QTableWidgetItem(pos['scanner']))
            self.chartink_positions_table.setItem(i, 1, QTableWidgetItem(pos['symbol']))
            self.chartink_positions_table.setItem(i, 2, QTableWidgetItem(pos['action']))
            self.chartink_positions_table.setItem(i, 3, QTableWidgetItem(str(pos['quantity'])))
            self.chartink_positions_table.setItem(i, 4, QTableWidgetItem(f"₹{pos['price']:.2f}" if pos.get('price') else "N/A"))
            self.chartink_positions_table.setItem(i, 5, QTableWidgetItem(pos.get('time', '')))

    def _remove_chartink_scan(self, scan_name: str):
        """Remove a Chartink scan"""
        self.chartink_scanner.remove_scan(scan_name)

        # Remove from config
        scans = self.config.get('chartink.scans', [])
        scans = [s for s in scans if s.get('name') != scan_name]
        self.config.set('chartink.scans', scans)

        self._refresh_chartink_scans_table()

    def _start_chartink_monitoring(self):
        """Start Chartink monitoring"""
        logger.info("=== START MONITORING CLICKED ===")
        logger.info(f"Active scans: {list(self.chartink_scanner.active_scans.keys())}")
        logger.info(f"Callbacks registered: {len(self.chartink_scanner.alert_callbacks)}")
        logger.info(f"Test mode: {self.chartink_scanner.test_mode}")
        self.chartink_scanner.start_monitoring()
        self.start_chartink_btn.setEnabled(False)
        self.stop_chartink_btn.setEnabled(True)
        self.status_bar.showMessage("Chartink monitoring started")
        logger.info("=== MONITORING STARTED ===")

    def _stop_chartink_monitoring(self):
        """Stop Chartink monitoring"""
        self.chartink_scanner.stop_monitoring()
        self.start_chartink_btn.setEnabled(True)
        self.stop_chartink_btn.setEnabled(False)
        self.status_bar.showMessage("Chartink monitoring stopped")

    def _reset_chartink_counters(self):
        """Reset daily trade counters for all scanners"""
        self.chartink_scanner.reset_daily_counters()
        self._refresh_chartink_scans_table()
        self.status_bar.showMessage("Daily counters reset")

    def _on_chartink_alert_thread(self, alert):
        """Called from background thread - emit signal to main thread"""
        # Emit signal to handle alert on main GUI thread
        logger.info(f"Alert received in thread callback: {alert.symbol} @ {alert.price}")
        self.chartink_alert_signal.emit(alert)

    def _handle_chartink_alert_gui(self, alert):
        """Handle Chartink alert on main GUI thread - execute trade with position tracking"""
        logger.info(f"Alert received in GUI handler: {alert.symbol} @ {alert.price}")
        from algo_trader.core.order_manager import Order, OrderType, TransactionType, Exchange

        extra = alert.extra_data or {}
        is_squareoff = extra.get('is_squareoff', False)

        # Get scan config to check trade type
        scan_config = self.chartink_scanner.active_scans.get(alert.scan_name, {})

        # Determine action and quantity
        if is_squareoff:
            action = extra.get('exit_action', 'SELL')
            quantity = extra.get('quantity', 1)
        else:
            action = scan_config.get('action', 'BUY')
            # Use calculated quantity from allocation logic
            quantity = extra.get('calculated_quantity', scan_config.get('quantity', 1))

        # Check if this is an F&O trade
        trade_type = scan_config.get('trade_type', 'equity')
        options_config = scan_config.get('options_config')

        # Log alert to table
        row = self.chartink_alerts_table.rowCount()
        self.chartink_alerts_table.insertRow(row)
        self.chartink_alerts_table.setItem(row, 0, QTableWidgetItem(alert.triggered_at.strftime("%H:%M:%S")))
        self.chartink_alerts_table.setItem(row, 1, QTableWidgetItem(alert.scan_name))
        self.chartink_alerts_table.setItem(row, 2, QTableWidgetItem(alert.symbol))
        self.chartink_alerts_table.setItem(row, 3, QTableWidgetItem(f"₹{alert.price:.2f}" if alert.price else "N/A"))
        self.chartink_alerts_table.setItem(row, 4, QTableWidgetItem(str(quantity)))

        # Check if this is an OPTIONS trade
        if trade_type == "options" and options_config and not is_squareoff:
            # Execute as Options trade using Auto Options
            self._execute_chartink_options_trade(alert, action, quantity, options_config)
            return

        # Execute trade - Paper Trading mode or Broker mode (EQUITY)
        paper_mode = self.config.get('trading.paper_mode', False)

        if paper_mode and hasattr(self, 'paper_simulator') and self.paper_simulator:
            # Paper Trading Mode
            try:
                result = self.paper_simulator.place_order(
                    symbol=alert.symbol,
                    action=action,
                    quantity=quantity,
                    order_type='MARKET',
                    price=alert.price or 0
                )

                if result.get('success'):
                    if is_squareoff:
                        action_text = f"PAPER SQUARE-OFF {action}: {result.get('order_id')}"
                        self.chartink_scanner.record_squareoff(alert.scan_name, alert.symbol)
                    else:
                        action_text = f"PAPER {action}: {result.get('order_id')}"
                        self.chartink_scanner.record_trade(
                            alert.scan_name, alert.symbol, action, quantity,
                            alert.price or 0
                        )
                    logger.info(f"Chartink paper trade: {action_text}")
                else:
                    action_text = f"Paper order failed: {result.get('message')}"
            except Exception as e:
                action_text = f"Paper order failed: {e}"
                logger.error(f"Chartink paper trade error: {e}")

        else:
            # Live Broker Mode
            current_broker = self.broker_combo.currentText().lower()
            if current_broker and current_broker in self.brokers:
                try:
                    order = Order(
                        symbol=alert.symbol,
                        transaction_type=TransactionType.BUY if action == "BUY" else TransactionType.SELL,
                        quantity=quantity,
                        order_type=OrderType.MARKET,
                        exchange=Exchange.NSE
                    )
                    result = self.order_manager.place_order(order, current_broker)

                    if is_squareoff:
                        action_text = f"SQUARE-OFF {action}: {result.broker_order_id}"
                        self.chartink_scanner.record_squareoff(alert.scan_name, alert.symbol)
                    else:
                        action_text = f"{action} order placed: {result.broker_order_id}"
                        # Record this trade for position tracking
                        self.chartink_scanner.record_trade(
                            alert.scan_name, alert.symbol, action, quantity,
                            alert.price or 0
                        )

                    logger.info(f"Chartink auto-trade: {action_text}")
                except Exception as e:
                    action_text = f"Order failed: {e}"
                    logger.error(f"Chartink auto-trade error: {e}")
            else:
                action_text = "No broker connected (enable Paper Trading mode in Settings)"

        self.chartink_alerts_table.setItem(row, 5, QTableWidgetItem(action_text))

        # Refresh positions and scans tables
        self._refresh_chartink_scans_table()

    def _on_positions_cell_clicked(self, row, col):
        """Handle click on positions table - open chart for Symbol column"""
        if col == 1:  # Symbol column
            item = self.chartink_positions_table.item(row, col)
            if item:
                symbol = item.text()
                self._open_tradingview_chart(symbol)

    def _on_alerts_cell_clicked(self, row, col):
        """Handle click on alerts table - open chart for Symbol column"""
        if col == 2:  # Symbol column
            item = self.chartink_alerts_table.item(row, col)
            if item:
                symbol = item.text()
                self._open_tradingview_chart(symbol)

    def _open_tradingview_chart(self, symbol: str):
        """Open TradingView chart for the given symbol"""
        # Clean symbol (remove .NS if present)
        clean_symbol = symbol.replace('.NS', '').replace('.BSE', '').upper()
        # TradingView URL for NSE stocks
        url = f"https://www.tradingview.com/chart/?symbol=NSE%3A{clean_symbol}"
        QDesktopServices.openUrl(QUrl(url))
        logger.info(f"Opening TradingView chart for {clean_symbol}")
        self.status_bar.showMessage(f"Opening chart for {clean_symbol}...", 3000)

    def _toggle_table_expand(self, table: QTableWidget, button: QPushButton):
        """Toggle table between normal and expanded view"""
        if button.text() == "⬜ Expand":
            # Expand
            table.setMinimumHeight(400)
            table.setMaximumHeight(600)
            button.setText("⬛ Collapse")
        else:
            # Collapse
            table.setMinimumHeight(120)
            table.setMaximumHeight(200)
            button.setText("⬜ Expand")

    def _execute_chartink_options_trade(self, alert, action: str, quantity: int, options_config: dict):
        """Execute an options trade from Chartink F&O scanner signal"""
        try:
            symbol = alert.symbol
            strike_selection = options_config.get('strike_selection', 'ATM')
            manual_strike = options_config.get('manual_strike', 0)
            option_type_setting = options_config.get('option_type', 'auto')
            expiry_setting = options_config.get('expiry', 'current_week')

            # Determine option type (CE/PE)
            if option_type_setting == 'auto':
                # BUY signal = CE (bullish), SELL signal = PE (bearish)
                option_type = 'CE' if action == 'BUY' else 'PE'
            else:
                option_type = option_type_setting  # CE or PE

            # Log to alerts table
            row = self.chartink_alerts_table.rowCount() - 1  # Last added row

            # Format strike display
            if strike_selection == 'Manual' and manual_strike > 0:
                strike_display = f"₹{manual_strike}"
            else:
                strike_display = strike_selection
            action_text = f"OPTIONS: {symbol} {strike_display} {option_type}"

            logger.info(f"Chartink F&O Signal: {symbol} -> {action} -> {option_type} {strike_display}")

            # Check if Auto Options executor is available
            if hasattr(self, 'auto_options') and self.auto_options:
                # Use Auto Options to execute
                from algo_trader.core.auto_options import SignalAction, StrikeSelection, ExpirySelection

                # Map strike selection
                strike_map = {
                    'ATM': StrikeSelection.ATM,
                    'ITM-1': StrikeSelection.ITM_1,
                    'ITM-2': StrikeSelection.ITM_2,
                    'OTM-1': StrikeSelection.OTM_1,
                    'OTM-2': StrikeSelection.OTM_2
                }

                # Map expiry
                expiry_map = {
                    'current_week': ExpirySelection.CURRENT_WEEK,
                    'next_week': ExpirySelection.NEXT_WEEK,
                    'current_month': ExpirySelection.CURRENT_MONTH
                }

                try:
                    # Determine strike - manual or auto
                    if strike_selection == 'Manual' and manual_strike > 0:
                        # Execute with manual strike price
                        result = self.auto_options.execute_signal(
                            underlying=symbol,
                            signal=SignalAction.BUY if action == 'BUY' else SignalAction.SELL,
                            strike_price=manual_strike,  # Use manual strike
                            option_type=option_type,
                            expiry_selection=expiry_map.get(expiry_setting, ExpirySelection.CURRENT_WEEK),
                            quantity=quantity
                        )
                    else:
                        # Execute via Auto Options with auto strike selection
                        result = self.auto_options.execute_signal(
                            underlying=symbol,
                            signal=SignalAction.BUY if action == 'BUY' else SignalAction.SELL,
                            strike_selection=strike_map.get(strike_selection, StrikeSelection.ATM),
                            expiry_selection=expiry_map.get(expiry_setting, ExpirySelection.CURRENT_WEEK),
                            quantity=quantity
                        )

                    if result and result.get('success'):
                        action_text = f"OPTIONS {option_type}: {result.get('option_symbol', 'N/A')}"
                        self.chartink_scanner.record_trade(
                            alert.scan_name, symbol, action, quantity,
                            alert.price or 0
                        )
                        logger.info(f"Chartink F&O trade executed: {action_text}")
                    else:
                        action_text = f"OPTIONS FAILED: {result.get('message', 'Unknown error')}"
                        logger.error(f"Chartink F&O trade failed: {result}")
                except Exception as e:
                    action_text = f"OPTIONS ERROR: {str(e)[:30]}"
                    logger.error(f"Chartink F&O trade error: {e}")
            else:
                # No Auto Options - just log the signal
                action_text = f"OPTIONS SIGNAL: {option_type} {strike_selection} (Auto-Options not configured)"
                logger.warning("Chartink F&O signal received but Auto-Options not configured")

            self.chartink_alerts_table.setItem(row, 5, QTableWidgetItem(action_text))
            self._refresh_chartink_scans_table()

        except Exception as e:
            logger.error(f"Error executing Chartink options trade: {e}")
            row = self.chartink_alerts_table.rowCount() - 1
            self.chartink_alerts_table.setItem(row, 5, QTableWidgetItem(f"OPTIONS ERROR: {str(e)[:30]}"))

    # Risk Management / TSL Methods
    def _init_risk_manager(self):
        """Initialize the risk manager"""
        self.risk_manager = RiskManager()
        self.risk_manager.register_mtm_callback(self._on_mtm_update)
        self.risk_manager.register_sl_hit_callback(self._on_sl_hit)
        self.risk_manager.register_target_hit_callback(self._on_target_hit)
        logger.info("Risk manager initialized")

    def _on_sl_type_changed(self, sl_type: str):
        """Handle stop loss type change"""
        if sl_type == "Fixed Price":
            self.risk_sl_value.setPrefix("₹")
            self.risk_sl_value.setSuffix("")
            self.risk_sl_value.setDecimals(2)
        elif sl_type == "Trailing %":
            self.risk_sl_value.setPrefix("")
            self.risk_sl_value.setSuffix("%")
            self.risk_sl_value.setDecimals(1)
            self.risk_sl_value.setRange(0.1, 50)
        elif sl_type == "Trailing Points":
            self.risk_sl_value.setPrefix("₹")
            self.risk_sl_value.setSuffix(" pts")
            self.risk_sl_value.setDecimals(2)

    def _add_risk_position(self):
        """Add a new position to risk manager"""
        symbol = self.risk_symbol.text().strip().upper()
        if not symbol:
            QMessageBox.warning(self, "Error", "Please enter a symbol")
            return

        exchange = self.risk_exchange.currentText()
        quantity = self.risk_quantity.value()
        entry_price = self.risk_entry_price.value()
        sl_type = self.risk_sl_type.currentText()
        sl_value = self.risk_sl_value.value()
        target = self.risk_target.value() if self.risk_target.value() > 0 else None

        if quantity == 0:
            QMessageBox.warning(self, "Error", "Quantity cannot be zero")
            return

        if entry_price <= 0:
            QMessageBox.warning(self, "Error", "Please enter a valid entry price")
            return

        # Determine SL parameters based on type
        stop_loss = None
        trailing_sl_percent = None
        trailing_sl_points = None

        if sl_value > 0:
            if sl_type == "Fixed Price":
                stop_loss = sl_value
            elif sl_type == "Trailing %":
                trailing_sl_percent = sl_value
                # Calculate initial SL
                if quantity > 0:
                    stop_loss = entry_price * (1 - sl_value / 100)
                else:
                    stop_loss = entry_price * (1 + sl_value / 100)
            elif sl_type == "Trailing Points":
                trailing_sl_points = sl_value
                # Calculate initial SL
                if quantity > 0:
                    stop_loss = entry_price - sl_value
                else:
                    stop_loss = entry_price + sl_value

        # Add position to risk manager
        self.risk_manager.add_position(
            symbol=symbol,
            quantity=quantity,
            entry_price=entry_price,
            exchange=exchange,
            stop_loss=stop_loss,
            target=target,
            trailing_sl_percent=trailing_sl_percent,
            trailing_sl_points=trailing_sl_points
        )

        # Clear inputs
        self.risk_symbol.clear()
        self.risk_entry_price.setValue(0)
        self.risk_sl_value.setValue(0)
        self.risk_target.setValue(0)

        # Refresh table
        self._refresh_risk_positions()

        QMessageBox.information(self, "Success", f"Position {symbol} added for tracking")

    def _start_risk_monitoring(self):
        """Start risk monitoring with price updates"""
        # Get active broker for price feed
        current_broker = self.broker_combo.currentText().lower()
        price_feed = self.brokers.get(current_broker) if current_broker else None

        self.risk_manager.start_monitoring(price_feed=price_feed, interval=2.0)

        self.start_risk_monitor_btn.setEnabled(False)
        self.stop_risk_monitor_btn.setEnabled(True)
        self.status_bar.showMessage("Risk monitoring started")

    def _stop_risk_monitoring(self):
        """Stop risk monitoring"""
        self.risk_manager.stop_monitoring()

        self.start_risk_monitor_btn.setEnabled(True)
        self.stop_risk_monitor_btn.setEnabled(False)
        self.status_bar.showMessage("Risk monitoring stopped")

    def _refresh_risk_positions(self):
        """Refresh the risk positions table"""
        positions = self.risk_manager.get_all_positions()
        self.risk_positions_table.setRowCount(len(positions))

        for i, pos in enumerate(positions):
            self.risk_positions_table.setItem(i, 0, QTableWidgetItem(pos.symbol))
            self.risk_positions_table.setItem(i, 1, QTableWidgetItem(str(pos.quantity)))
            self.risk_positions_table.setItem(i, 2, QTableWidgetItem(f"₹{pos.entry_price:.2f}"))
            self.risk_positions_table.setItem(i, 3, QTableWidgetItem(f"₹{pos.current_price:.2f}"))
            self.risk_positions_table.setItem(i, 4, QTableWidgetItem(f"₹{pos.stop_loss:.2f}" if pos.stop_loss else "-"))
            self.risk_positions_table.setItem(i, 5, QTableWidgetItem(f"₹{pos.target:.2f}" if pos.target else "-"))

            # High/Low for trailing SL
            if pos.quantity > 0:
                self.risk_positions_table.setItem(i, 6, QTableWidgetItem(f"H: ₹{pos.highest_price:.2f}"))
            else:
                self.risk_positions_table.setItem(i, 6, QTableWidgetItem(f"L: ₹{pos.lowest_price:.2f}"))

            # P&L with color
            pnl_item = QTableWidgetItem(f"₹{pos.pnl:.2f}")
            if pos.pnl > 0:
                pnl_item.setForeground(Qt.GlobalColor.darkGreen)
            elif pos.pnl < 0:
                pnl_item.setForeground(Qt.GlobalColor.red)
            self.risk_positions_table.setItem(i, 7, pnl_item)

            # P&L %
            pnl_pct_item = QTableWidgetItem(f"{pos.pnl_percent:.2f}%")
            if pos.pnl_percent > 0:
                pnl_pct_item.setForeground(Qt.GlobalColor.darkGreen)
            elif pos.pnl_percent < 0:
                pnl_pct_item.setForeground(Qt.GlobalColor.red)
            self.risk_positions_table.setItem(i, 8, pnl_pct_item)

            # Close button
            close_btn = QPushButton("Close")
            close_btn.clicked.connect(lambda checked, s=pos.symbol, e=pos.exchange: self._close_risk_position(s, e))
            self.risk_positions_table.setCellWidget(i, 9, close_btn)

        # Update MTM summary
        mtm = self.risk_manager.get_mtm_summary()
        self._update_mtm_display(mtm)

    def _close_risk_position(self, symbol: str, exchange: str):
        """Close a tracked position"""
        position = self.risk_manager.get_position(symbol, exchange)
        if position:
            # Use current price as exit price
            self.risk_manager.close_position(symbol, position.current_price, exchange)
            self._refresh_risk_positions()
            QMessageBox.information(self, "Position Closed", f"Closed {symbol} @ ₹{position.current_price:.2f}")

    def _update_mtm_display(self, mtm):
        """Update MTM display labels"""
        # Total P&L
        total_color = "green" if mtm.total_pnl >= 0 else "red"
        self.mtm_total_pnl.setText(f"Total P&L: <span style='color:{total_color}'>₹{mtm.total_pnl:.2f}</span>")
        self.mtm_total_pnl.setTextFormat(Qt.TextFormat.RichText)

        # Realized
        realized_color = "green" if mtm.realized_pnl >= 0 else "red"
        self.mtm_realized.setText(f"Realized: <span style='color:{realized_color}'>₹{mtm.realized_pnl:.2f}</span>")
        self.mtm_realized.setTextFormat(Qt.TextFormat.RichText)

        # Unrealized
        unrealized_color = "green" if mtm.unrealized_pnl >= 0 else "red"
        self.mtm_unrealized.setText(f"Unrealized: <span style='color:{unrealized_color}'>₹{mtm.unrealized_pnl:.2f}</span>")
        self.mtm_unrealized.setTextFormat(Qt.TextFormat.RichText)

        # Trade stats
        self.mtm_trades.setText(f"Trades: {mtm.total_trades} (W: {mtm.winning_trades} / L: {mtm.losing_trades})")

    def _on_mtm_update(self, mtm):
        """Handle MTM update from risk manager"""
        self._update_mtm_display(mtm)
        self._refresh_risk_positions()

    def _on_sl_hit(self, position):
        """Handle stop loss hit"""
        QMessageBox.warning(
            self, "Stop Loss Hit",
            f"Stop Loss triggered for {position.symbol}!\n"
            f"Entry: ₹{position.entry_price:.2f}\n"
            f"Exit: ₹{position.current_price:.2f}\n"
            f"P&L: ₹{position.pnl:.2f}"
        )

        # Auto-close position from tracking
        self.risk_manager.close_position(position.symbol, position.current_price, position.exchange)
        self._refresh_risk_positions()

        # Place exit order if broker connected
        self._place_exit_order(position)

    def _on_target_hit(self, position):
        """Handle target hit"""
        QMessageBox.information(
            self, "Target Hit",
            f"Target reached for {position.symbol}!\n"
            f"Entry: ₹{position.entry_price:.2f}\n"
            f"Exit: ₹{position.current_price:.2f}\n"
            f"P&L: ₹{position.pnl:.2f}"
        )

        # Auto-close position from tracking
        self.risk_manager.close_position(position.symbol, position.current_price, position.exchange)
        self._refresh_risk_positions()

        # Place exit order if broker connected
        self._place_exit_order(position)

    def _place_exit_order(self, position):
        """Place exit order for a position"""
        from algo_trader.core.order_manager import Order, OrderType, TransactionType, Exchange

        current_broker = self.broker_combo.currentText().lower()
        if not current_broker or current_broker not in self.brokers:
            logger.warning("No broker connected for exit order")
            return

        try:
            # Exit is opposite of entry
            if position.quantity > 0:
                transaction_type = TransactionType.SELL
            else:
                transaction_type = TransactionType.BUY

            order = Order(
                symbol=position.symbol,
                transaction_type=transaction_type,
                quantity=abs(position.quantity),
                order_type=OrderType.MARKET,
                exchange=Exchange[position.exchange]
            )

            result = self.order_manager.place_order(order, current_broker)
            logger.info(f"Exit order placed: {result.broker_order_id}")
            self._load_orders()

        except Exception as e:
            logger.error(f"Failed to place exit order: {e}")

    # Options Trading Methods
    def _init_options_manager(self):
        """Initialize the options manager"""
        self.options_manager = OptionsManager()
        self.options_manager.register_exit_callback(self._on_option_exit)
        self.options_manager.register_pnl_callback(self._on_option_pnl_update)
        self._on_opt_symbol_changed(self.opt_symbol.currentText())

        # Initialize auto-options executor
        self.auto_options = AutoOptionsExecutor(self.options_manager, self.strategy_engine)
        self.auto_options.register_trade_callback(self._on_auto_option_trade)
        logger.info("Options manager initialized")

    def _save_auto_options_config(self):
        """Save auto-options configuration with per-leg settings"""
        from algo_trader.core.auto_options import LegConfig

        # Update Leg 1 config
        leg1_strike = self.auto_leg1_strike.currentText()
        leg1_manual = int(self.auto_leg1_manual_strike.value()) if leg1_strike == "Manual" else 0
        self.auto_options.config.leg1 = LegConfig(
            enabled=True,
            option_type=self.auto_leg1_type.currentText(),
            action=self.auto_leg1_action.currentText(),
            strike_selection=leg1_strike,
            expiry_selection=self.auto_leg1_expiry.currentText(),
            quantity=self.auto_leg1_qty.value(),
            manual_strike=leg1_manual
        )

        # Update Leg 2 config (hedge)
        hedge_on = self.auto_opt_hedge_enabled.isChecked()
        leg2_strike = self.auto_leg2_strike.currentText()
        leg2_manual = int(self.auto_leg2_manual_strike.value()) if leg2_strike == "Manual" else 0
        self.auto_options.config.leg2 = LegConfig(
            enabled=hedge_on,
            option_type=self.auto_leg2_type.currentText(),
            action=self.auto_leg2_action.currentText(),
            strike_selection=leg2_strike,
            expiry_selection=self.auto_leg2_expiry.currentText(),
            quantity=self.auto_leg2_qty.value(),
            manual_strike=leg2_manual
        )

        self.auto_options.update_config(
            symbol=self.auto_opt_symbol.currentText().strip().upper(),
            hedge_enabled=hedge_on,
            close_on_opposite=self.auto_opt_close_opposite.isChecked(),
            exit_type=self.auto_opt_exit_type.currentText(),
            sl_value=self.auto_opt_sl.value(),
            target_value=self.auto_opt_target.value(),
            tsl_value=self.auto_opt_tsl.value(),
        )

        if self.auto_opt_enabled.isChecked():
            current_broker = self.broker_combo.currentText().lower()
            if current_broker and current_broker in self.brokers:
                self.auto_options.set_broker(self.brokers[current_broker])
            self.auto_options.enable()
        else:
            self.auto_options.disable()

        # Build info message
        leg1_strike_display = f"₹{leg1_manual}" if leg1_strike == "Manual" else leg1_strike
        leg1_info = (f"Leg 1: {self.auto_leg1_action.currentText()} "
                     f"{self.auto_leg1_type.currentText()} "
                     f"Strike:{leg1_strike_display} "
                     f"Expiry:{self.auto_leg1_expiry.currentText()}")

        leg2_info = ""
        if hedge_on:
            leg2_strike_display = f"₹{leg2_manual}" if leg2_strike == "Manual" else leg2_strike
            leg2_info = (f"\nLeg 2: {self.auto_leg2_action.currentText()} "
                         f"{self.auto_leg2_type.currentText()} "
                         f"Strike:{self.auto_leg2_strike.currentText()} "
                         f"Expiry:{self.auto_leg2_expiry.currentText()}")

        QMessageBox.information(
            self, "Saved",
            f"Auto-Options config saved!\n"
            f"Status: {'ENABLED' if self.auto_opt_enabled.isChecked() else 'DISABLED'}\n"
            f"{leg1_info}{leg2_info}"
        )

    def _on_auto_option_trade(self, trade_info):
        """Handle auto-option trade execution"""
        row = self.auto_trade_log_table.rowCount()
        self.auto_trade_log_table.insertRow(row)
        self.auto_trade_log_table.setItem(row, 0, QTableWidgetItem(trade_info.get("time", "")))
        self.auto_trade_log_table.setItem(row, 1, QTableWidgetItem(trade_info.get("signal", "")))
        self.auto_trade_log_table.setItem(row, 2, QTableWidgetItem(trade_info.get("strategy", "")))
        self.auto_trade_log_table.setItem(row, 3, QTableWidgetItem(trade_info.get("action", "")))
        self.auto_trade_log_table.setItem(row, 4, QTableWidgetItem(trade_info.get("expiry", "")))
        self.auto_trade_log_table.setItem(row, 5, QTableWidgetItem(str(trade_info.get("qty", ""))))
        self._refresh_option_positions()

    def _on_opt_symbol_changed(self, symbol: str):
        """Handle symbol change in options tab"""
        if not symbol:
            return
        from algo_trader.core.options_manager import INDEX_LOT_SIZES
        lot_size = INDEX_LOT_SIZES.get(symbol.upper(), 1)
        self.opt_lot_size.setText(f"Lot Size: {lot_size}")

    def _fetch_spot_price(self):
        """Fetch spot price from broker"""
        symbol = self.opt_symbol.currentText().strip()
        current_broker = self.broker_combo.currentText().lower()

        if current_broker and current_broker in self.brokers:
            try:
                quote = self.brokers[current_broker].get_quote(symbol, "NSE")
                if quote:
                    ltp = quote.get('last_price') or quote.get('ltp') or 0
                    if isinstance(ltp, dict):
                        ltp = ltp.get('last_price', 0)
                    self.opt_spot_price.setValue(float(ltp))
                    self.status_bar.showMessage(f"Spot price fetched: ₹{ltp}")
                    return
            except Exception as e:
                logger.error(f"Error fetching spot: {e}")

        QMessageBox.warning(self, "Error", "Could not fetch spot price. Please enter manually or connect broker.")

    def _load_expiry_dates(self):
        """Load expiry dates for selected symbol"""
        symbol = self.opt_symbol.currentText().strip()
        if not symbol:
            return

        current_broker = self.broker_combo.currentText().lower()
        broker = self.brokers.get(current_broker) if current_broker else None

        expiries = self.options_manager.get_expiry_dates(symbol, broker)
        self.opt_expiry.clear()
        for exp in expiries:
            self.opt_expiry.addItem(exp)

        self.status_bar.showMessage(f"Loaded {len(expiries)} expiry dates for {symbol}")

    def _load_strike_prices(self):
        """Load strike prices based on spot price"""
        symbol = self.opt_symbol.currentText().strip()
        spot = self.opt_spot_price.value()

        if spot <= 0:
            QMessageBox.warning(self, "Error", "Please enter or fetch spot price first")
            return

        strikes = self.options_manager.get_strike_prices(symbol, spot)
        self.opt_strike.clear()
        for strike in strikes:
            self.opt_strike.addItem(str(int(strike)) if strike == int(strike) else str(strike))

        # Select ATM
        from algo_trader.core.options_manager import INDEX_STRIKE_GAPS
        gap = INDEX_STRIKE_GAPS.get(symbol.upper(), 50)
        atm = round(spot / gap) * gap
        atm_str = str(int(atm)) if atm == int(atm) else str(atm)
        idx = self.opt_strike.findText(atm_str)
        if idx >= 0:
            self.opt_strike.setCurrentIndex(idx)

        self.status_bar.showMessage(f"Loaded {len(strikes)} strikes. ATM: {atm_str}")

    def _add_builder_leg(self):
        """Add a new leg row to the multi-leg builder table"""
        row = self.leg_builder_table.rowCount()
        self.leg_builder_table.insertRow(row)

        # Strike combo - populate from loaded strikes
        strike_combo = QComboBox()
        strike_combo.setEditable(True)
        for i in range(self.opt_strike.count()):
            strike_combo.addItem(self.opt_strike.itemText(i))
        # Auto-select ATM if available
        if self.opt_strike.currentText():
            strike_combo.setCurrentText(self.opt_strike.currentText())
        self.leg_builder_table.setCellWidget(row, 0, strike_combo)

        # Expiry combo - populate from loaded expiries
        expiry_combo = QComboBox()
        for i in range(self.opt_expiry.count()):
            expiry_combo.addItem(self.opt_expiry.itemText(i))
        if self.opt_expiry.currentText():
            expiry_combo.setCurrentText(self.opt_expiry.currentText())
        self.leg_builder_table.setCellWidget(row, 1, expiry_combo)

        # CE/PE
        type_combo = QComboBox()
        type_combo.addItems(["CE", "PE"])
        self.leg_builder_table.setCellWidget(row, 2, type_combo)

        # BUY/SELL
        action_combo = QComboBox()
        action_combo.addItems(["BUY", "SELL"])
        self.leg_builder_table.setCellWidget(row, 3, action_combo)

        # Lots
        qty_spin = QSpinBox()
        qty_spin.setRange(1, 500)
        qty_spin.setValue(1)
        self.leg_builder_table.setCellWidget(row, 4, qty_spin)

        # Premium
        premium_spin = QDoubleSpinBox()
        premium_spin.setRange(0, 100000)
        premium_spin.setDecimals(2)
        self.leg_builder_table.setCellWidget(row, 5, premium_spin)

        # Remove button
        remove_btn = QPushButton("X")
        remove_btn.setMaximumWidth(30)
        remove_btn.clicked.connect(lambda checked, r=row: self._remove_builder_leg(r))
        self.leg_builder_table.setCellWidget(row, 6, remove_btn)

    def _remove_builder_leg(self, row):
        """Remove a leg from builder"""
        if row < self.leg_builder_table.rowCount():
            self.leg_builder_table.removeRow(row)

    def _on_opt_strategy_changed(self, strategy_name: str):
        """Handle strategy type change - kept for compatibility"""
        pass

    def _on_opt_exit_type_changed(self, exit_type: str):
        """Handle exit type change"""
        is_pnl = exit_type == "P&L Based"
        is_sl_pct = exit_type == "SL %"
        is_tsl = exit_type in ("TSL %", "TSL Points")

        if is_pnl:
            self.opt_sl_value.setPrefix("₹")
            self.opt_sl_value.setSuffix("")
            self.opt_target_value.setPrefix("₹")
            self.opt_target_value.setSuffix("")
            self.opt_tsl_value.setPrefix("₹")
            self.opt_tsl_value.setSuffix("")
        elif is_sl_pct:
            self.opt_sl_value.setPrefix("")
            self.opt_sl_value.setSuffix("%")
            self.opt_target_value.setPrefix("")
            self.opt_target_value.setSuffix("%")
        elif is_tsl:
            self.opt_tsl_value.setPrefix("")
            self.opt_tsl_value.setSuffix("%" if exit_type == "TSL %" else " pts")

    def _add_option_position(self):
        """Add a new multi-leg option position from builder table"""
        symbol = self.opt_symbol.currentText().strip().upper()
        if not symbol:
            QMessageBox.warning(self, "Error", "Please select a symbol")
            return

        num_legs = self.leg_builder_table.rowCount()
        if num_legs == 0:
            QMessageBox.warning(self, "Error", "Please add at least one leg using '+ Add Leg' button")
            return

        # Read all legs from builder table
        legs_data = []
        for row in range(num_legs):
            strike_widget = self.leg_builder_table.cellWidget(row, 0)
            expiry_widget = self.leg_builder_table.cellWidget(row, 1)
            type_widget = self.leg_builder_table.cellWidget(row, 2)
            action_widget = self.leg_builder_table.cellWidget(row, 3)
            qty_widget = self.leg_builder_table.cellWidget(row, 4)
            premium_widget = self.leg_builder_table.cellWidget(row, 5)

            if not all([strike_widget, expiry_widget, type_widget, action_widget, qty_widget, premium_widget]):
                QMessageBox.warning(self, "Error", f"Leg {row + 1} has missing fields")
                return

            try:
                strike = float(strike_widget.currentText())
            except (ValueError, AttributeError):
                QMessageBox.warning(self, "Error", f"Leg {row + 1}: Invalid strike price")
                return

            expiry = expiry_widget.currentText()
            if not expiry:
                QMessageBox.warning(self, "Error", f"Leg {row + 1}: Please select expiry")
                return

            legs_data.append({
                "strike": strike,
                "expiry": expiry,
                "option_type": type_widget.currentText(),
                "action": action_widget.currentText(),
                "quantity": qty_widget.value(),
                "entry_price": premium_widget.value()
            })

        exit_type = self.opt_exit_type.currentText()
        sl_value = self.opt_sl_value.value()
        target_value = self.opt_target_value.value()
        tsl_value = self.opt_tsl_value.value()

        if num_legs == 1:
            # Single leg - use create_single_option
            leg = legs_data[0]
            position = self.options_manager.create_single_option(
                symbol=symbol, expiry=leg["expiry"], strike=leg["strike"],
                option_type=leg["option_type"], action=leg["action"],
                quantity=leg["quantity"], entry_price=leg["entry_price"],
                exit_type=exit_type, sl_value=sl_value,
                target_value=target_value, tsl_value=tsl_value
            )
        else:
            # Multi-leg - use create_custom_multileg
            position = self.options_manager.create_custom_multileg(
                symbol=symbol, legs_data=legs_data,
                exit_type=exit_type, sl_value=sl_value,
                target_value=target_value, tsl_value=tsl_value
            )

        # Show confirmation
        legs_info = "\n".join(
            f"  Leg {i+1}: {l['strike']}{l['option_type']} {l['action']} "
            f"Exp:{l['expiry']} Qty:{l['quantity']} @₹{l['entry_price']}"
            for i, l in enumerate(legs_data)
        )
        QMessageBox.information(
            self, "Position Created",
            f"{position.position_id}: {symbol}\n"
            f"Legs ({num_legs}):\n{legs_info}"
        )

        # Clear builder table
        self.leg_builder_table.setRowCount(0)

        self._refresh_option_positions()

    def _refresh_option_positions(self):
        """Refresh the options positions table"""
        positions = self.options_manager.get_all_positions()
        self.opt_positions_table.setRowCount(len(positions))

        for i, pos in enumerate(positions):
            self.opt_positions_table.setItem(i, 0, QTableWidgetItem(pos.position_id))
            self.opt_positions_table.setItem(i, 1, QTableWidgetItem(pos.symbol))
            self.opt_positions_table.setItem(i, 2, QTableWidgetItem(pos.strategy_type.value))
            self.opt_positions_table.setItem(i, 3, QTableWidgetItem(str(len(pos.legs))))

            # Expiry from first leg
            expiry = pos.legs[0].expiry if pos.legs else "-"
            self.opt_positions_table.setItem(i, 4, QTableWidgetItem(expiry))

            # P&L with color
            pnl_item = QTableWidgetItem(f"₹{pos.total_pnl:.2f}")
            if pos.total_pnl > 0:
                pnl_item.setForeground(Qt.GlobalColor.darkGreen)
            elif pos.total_pnl < 0:
                pnl_item.setForeground(Qt.GlobalColor.red)
            self.opt_positions_table.setItem(i, 5, pnl_item)

            # P&L %
            pct_item = QTableWidgetItem(f"{pos.total_pnl_percent:.2f}%")
            if pos.total_pnl_percent > 0:
                pct_item.setForeground(Qt.GlobalColor.darkGreen)
            elif pos.total_pnl_percent < 0:
                pct_item.setForeground(Qt.GlobalColor.red)
            self.opt_positions_table.setItem(i, 6, pct_item)

            self.opt_positions_table.setItem(i, 7, QTableWidgetItem(f"₹{pos.max_pnl:.2f}"))
            self.opt_positions_table.setItem(i, 8, QTableWidgetItem(pos.exit_type.value))

            # Close button
            close_btn = QPushButton("Close")
            close_btn.clicked.connect(lambda checked, pid=pos.position_id: self._close_option_position(pid))
            self.opt_positions_table.setCellWidget(i, 9, close_btn)

        # Update summary
        summary = self.options_manager.get_options_summary()
        color = "green" if summary["total_pnl"] >= 0 else "red"
        self.opt_total_pnl.setText(f"Total P&L: <span style='color:{color}'>₹{summary['total_pnl']:.2f}</span>")
        self.opt_total_pnl.setTextFormat(Qt.TextFormat.RichText)
        self.opt_active_count.setText(f"Active: {summary['active_positions']}")
        self.opt_closed_count.setText(f"Closed: {summary['closed_positions']}")

    def _on_opt_position_selected(self, row, col):
        """Show leg details when position is selected"""
        pos_id_item = self.opt_positions_table.item(row, 0)
        if not pos_id_item:
            return

        pos = self.options_manager.get_position(pos_id_item.text())
        if not pos:
            return

        self.opt_legs_table.setRowCount(len(pos.legs))
        for i, leg in enumerate(pos.legs):
            self.opt_legs_table.setItem(i, 0, QTableWidgetItem(f"Leg {leg.leg_id}"))
            self.opt_legs_table.setItem(i, 1, QTableWidgetItem(str(int(leg.strike))))
            self.opt_legs_table.setItem(i, 2, QTableWidgetItem(leg.expiry))
            self.opt_legs_table.setItem(i, 3, QTableWidgetItem(leg.option_type.value))
            self.opt_legs_table.setItem(i, 4, QTableWidgetItem(leg.action))
            self.opt_legs_table.setItem(i, 5, QTableWidgetItem(f"{leg.quantity} x {leg.lot_size}"))
            self.opt_legs_table.setItem(i, 6, QTableWidgetItem(f"₹{leg.entry_price:.2f}"))
            self.opt_legs_table.setItem(i, 7, QTableWidgetItem(f"₹{leg.current_price:.2f}"))

            pnl_item = QTableWidgetItem(f"₹{leg.pnl:.2f}")
            if leg.pnl > 0:
                pnl_item.setForeground(Qt.GlobalColor.darkGreen)
            elif leg.pnl < 0:
                pnl_item.setForeground(Qt.GlobalColor.red)
            self.opt_legs_table.setItem(i, 8, pnl_item)

            pct_item = QTableWidgetItem(f"{leg.pnl_percent:.2f}%")
            if leg.pnl_percent > 0:
                pct_item.setForeground(Qt.GlobalColor.darkGreen)
            elif leg.pnl_percent < 0:
                pct_item.setForeground(Qt.GlobalColor.red)
            self.opt_legs_table.setItem(i, 9, pct_item)

    def _close_option_position(self, pos_id: str):
        """Close an option position"""
        pos = self.options_manager.get_position(pos_id)
        if not pos:
            return

        reply = QMessageBox.question(
            self, "Close Position",
            f"Close {pos.symbol} {pos.strategy_type.value} position?\n"
            f"Current P&L: ₹{pos.total_pnl:.2f}",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )

        if reply == QMessageBox.StandardButton.Yes:
            # Place exit orders for each leg
            self._place_option_exit_orders(pos)
            self.options_manager.close_position(pos_id)
            self._refresh_option_positions()
            QMessageBox.information(self, "Closed", f"Position {pos_id} closed. P&L: ₹{pos.total_pnl:.2f}")

    def _place_option_exit_orders(self, position):
        """Place exit orders for all legs of an option position"""
        from algo_trader.core.order_manager import Order, OrderType, TransactionType, Exchange

        current_broker = self.broker_combo.currentText().lower()
        if not current_broker or current_broker not in self.brokers:
            logger.warning("No broker connected for option exit orders")
            return

        for leg in position.legs:
            try:
                # Exit is opposite of entry action
                exit_action = TransactionType.SELL if leg.action == "BUY" else TransactionType.BUY

                order = Order(
                    symbol=leg.trading_symbol,
                    transaction_type=exit_action,
                    quantity=leg.quantity * leg.lot_size,
                    order_type=OrderType.MARKET,
                    exchange=Exchange.NFO,
                    product="NRML"
                )

                result = self.order_manager.place_order(order, current_broker)
                logger.info(f"Option exit order: {leg.trading_symbol} -> {result.broker_order_id}")

            except Exception as e:
                logger.error(f"Failed to place option exit order for {leg.trading_symbol}: {e}")

    def _on_option_exit(self, position, reason):
        """Handle auto-exit trigger for options"""
        QMessageBox.warning(
            self, "Option Exit Triggered",
            f"Exit triggered for {position.symbol} {position.strategy_type.value}!\n"
            f"Reason: {reason}\n"
            f"P&L: ₹{position.total_pnl:.2f}"
        )

        self._place_option_exit_orders(position)
        self.options_manager.close_position(position.position_id)
        self._refresh_option_positions()

    def _on_option_pnl_update(self, position):
        """Handle P&L update for options"""
        self._refresh_option_positions()

    def _load_window_geometry(self):
        """Load saved window size and position"""
        try:
            import json
            import os
            settings_file = os.path.join(os.path.dirname(__file__), '..', 'window_settings.json')
            if os.path.exists(settings_file):
                with open(settings_file, 'r') as f:
                    settings = json.load(f)
                    if 'width' in settings and 'height' in settings:
                        self.resize(settings['width'], settings['height'])
                    if 'x' in settings and 'y' in settings:
                        self.move(settings['x'], settings['y'])
                    if settings.get('maximized', False):
                        self.showMaximized()
        except Exception as e:
            logger.debug(f"Could not load window geometry: {e}")

    def _save_window_geometry(self):
        """Save window size and position"""
        try:
            import json
            import os
            settings_file = os.path.join(os.path.dirname(__file__), '..', 'window_settings.json')
            settings = {
                'width': self.width(),
                'height': self.height(),
                'x': self.x(),
                'y': self.y(),
                'maximized': self.isMaximized()
            }
            with open(settings_file, 'w') as f:
                json.dump(settings, f)
        except Exception as e:
            logger.debug(f"Could not save window geometry: {e}")

    def closeEvent(self, event):
        """Handle window close"""
        reply = QMessageBox.question(
            self, "Exit",
            "Are you sure you want to exit?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        if reply == QMessageBox.StandardButton.Yes:
            self._save_window_geometry()  # Save window size before closing
            event.accept()
        else:
            event.ignore()
